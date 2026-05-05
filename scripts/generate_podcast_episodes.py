#!/usr/bin/env python3
"""
generate_podcast_episodes.py
Build-time script: generates first-party episode pages under /podcast/episodes/<slug>/
from live podcast RSS data merged with any existing manual enrichment stored in
data/podcast-episodes.json.

Each page includes:
  - Episode summary and key takeaways
  - Transcript text or external transcript URL
  - Related topic guide and book links
  - PodcastEpisode JSON-LD schema
  - A compatibility redirect page for bare session IDs like /podcast/TT-2026-04-10/
"""
from __future__ import annotations

import html as html_mod
import json
import os
import re
import sys
import urllib.request

import openpyxl
import xml.etree.ElementTree as ET
from email.utils import parsedate_to_datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
HEADER_PARTIAL = ROOT / "assets" / "partials" / "header.html"
FOOTER_PARTIAL = ROOT / "assets" / "partials" / "footer.html"
EPISODES_DATA = ROOT / "data" / "podcast-episodes.json"
EPISODES_DIR = ROOT / "podcast" / "episodes"
PODCAST_COMPAT_DIR = ROOT / "podcast"
REDIRECTS_FILE = ROOT / "_redirects"
RSS_URL = os.environ.get(
    "PODCAST_RSS_URL",
    "https://podcast-rss-feeds.jonathan-harris.online/turing-torch.xml",
)
SITE = os.environ.get("PODCAST_SITE_BASE_URL", "https://jonathan-harris.online").rstrip("/")
SERIES_NAME = "Turing's Torch: AI Weekly"
SERIES_URL = f"{SITE}/podcast/"
NS = {"podcast": "https://podcastindex.org/namespace/1.0"}
PAGE_HEADER_ROW = 5
GENERIC_PODCAST_SLUGS = {"artificial-intelligence-weekly"}


def load_partial(path: Path, label: str) -> str:
    if not path.exists():
        raise FileNotFoundError(f"{label} partial missing: {path}")
    return path.read_text(encoding="utf-8").rstrip("\n")


def slugify(title: str) -> str:
    slug = (title or "").lower()
    slug = re.sub(r"['’]", "", slug)
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    return slug.strip("-")[:80]


def is_absolute_http_url(value: str | None) -> bool:
    return bool(value and re.match(r"^https?://", value.strip(), flags=re.IGNORECASE))


def first_non_empty(*values: object) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def clean_description(raw: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw or "")
    return re.sub(r"\s+", " ", text).strip()


def format_date_iso(pub_date: str) -> str:
    if not pub_date:
        return ""
    try:
        return parsedate_to_datetime(pub_date).strftime("%Y-%m-%d")
    except Exception:
        return ""


def is_first_party_episode_url(value: str | None) -> bool:
    if not is_absolute_http_url(value):
        return False
    candidate = str(value).strip().rstrip("/") + "/"
    return candidate.startswith(f"{SITE}/podcast/episodes/")


def parse_episode_page_url(link: str, title: str) -> str:
    if is_first_party_episode_url(link):
        return link
    return f"{SITE}/podcast/episodes/{slugify(title)}/"

def episode_slug_suffix(ep: dict) -> str:
    date_value = first_non_empty(ep.get("date"))
    session_id = first_non_empty(ep.get("session_id"))
    date_match = re.search(r"\d{4}-\d{2}-\d{2}", first_non_empty(date_value, session_id))
    if date_match:
        return date_match.group(0)
    return slugify(first_non_empty(session_id, ep.get("title")))[:40]


def ensure_unique_episode_slugs(episodes: list[dict]) -> list[dict]:
    """Guarantee one stable canonical episode URL per episode record.

    RSS titles can collapse into generic slugs such as
    ``artificial-intelligence-weekly``. That creates multiple records with one
    canonical URL, so append the ISO episode date or session identifier whenever
    the base slug is generic or already taken.
    """
    seen: set[str] = set()
    normalised: list[dict] = []

    for raw_ep in episodes:
        ep = dict(raw_ep)
        base_slug = slugify(first_non_empty(ep.get("slug"), ep.get("title")))
        if not base_slug:
            base_slug = slugify(first_non_empty(ep.get("session_id"), "podcast-episode")) or "podcast-episode"

        slug = base_slug
        if slug in GENERIC_PODCAST_SLUGS or slug in seen:
            suffix = slugify(episode_slug_suffix(ep))
            slug = f"{base_slug}-{suffix}" if suffix else base_slug

        counter = 2
        candidate = slug
        while candidate in seen:
            candidate = f"{slug}-{counter}"
            counter += 1

        ep["slug"] = candidate
        ep["page_url"] = f"{SITE}/podcast/episodes/{candidate}/"
        seen.add(candidate)
        normalised.append(ep)

    return normalised


def persist_episode_data(episodes: list[dict]) -> None:
    EPISODES_DATA.parent.mkdir(parents=True, exist_ok=True)
    EPISODES_DATA.write_text(json.dumps(episodes, indent=2, ensure_ascii=False), encoding="utf-8")



def parse_audio_url(item: ET.Element) -> str:
    enclosure = item.find("enclosure")
    if enclosure is None:
        return ""
    candidate = first_non_empty(enclosure.get("url"), enclosure.get("href"))
    return candidate if is_absolute_http_url(candidate) else ""


def parse_transcript_url(item: ET.Element) -> str:
    tx = item.find("podcast:transcript", NS)
    if tx is None:
        return ""
    candidate = first_non_empty(tx.get("url"), tx.get("href"), tx.text)
    return candidate if is_absolute_http_url(candidate) else ""


def fetch_rss_episodes(limit: int = 20) -> list[dict]:
    request = urllib.request.Request(RSS_URL, headers={"User-Agent": "PodcastEpisodePageSync/1.0"})
    with urllib.request.urlopen(request, timeout=20) as resp:
        tree = ET.parse(resp)

    episodes: list[dict] = []
    for item in tree.findall(".//item")[:limit]:
        title = (item.findtext("title", "") or "").strip()
        if not title:
            continue

        pub_date = (item.findtext("pubDate", "") or "").strip()
        link = (item.findtext("link", "") or "").strip()
        guid = (item.findtext("guid", "") or "").strip()
        description = clean_description((item.findtext("description", "") or "").strip())
        session_id = guid if guid and not is_absolute_http_url(guid) else ""
        slug = slugify(title)

        episodes.append(
            {
                "title": title,
                "slug": slug,
                "session_id": session_id,
                "date": format_date_iso(pub_date),
                "page_url": parse_episode_page_url(link, title),
                "audio_url": parse_audio_url(item),
                "external_url": "",
                "summary": description[:500] if description else "",
                "key_takeaways": [],
                "transcript_url": parse_transcript_url(item),
                "transcript_text": "",
                "related_topic": {},
                "related_books": [],
            }
        )
    return episodes


def load_existing_episodes() -> dict[str, dict]:
    if not EPISODES_DATA.exists():
        return {}
    try:
        items = json.loads(EPISODES_DATA.read_text(encoding="utf-8"))
    except Exception:
        return {}
    existing: dict[str, dict] = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        key = first_non_empty(item.get("session_id"), item.get("slug"), item.get("title"))
        if key:
            existing[key] = item
    return existing


def merge_episode(rss_ep: dict, existing_lookup: dict[str, dict]) -> dict:
    existing = existing_lookup.get(first_non_empty(rss_ep.get("session_id"), rss_ep.get("slug"), rss_ep.get("title")), {})
    merged = dict(existing)
    merged.update(rss_ep)

    merged["title"] = rss_ep["title"]
    merged["slug"] = rss_ep["slug"]
    merged["session_id"] = first_non_empty(rss_ep.get("session_id"), existing.get("session_id"))
    merged["date"] = first_non_empty(rss_ep.get("date"), existing.get("date"))
    merged["page_url"] = first_non_empty(rss_ep.get("page_url"), existing.get("page_url"), f"{SITE}/podcast/episodes/{rss_ep['slug']}/")
    merged["audio_url"] = first_non_empty(rss_ep.get("audio_url"), existing.get("audio_url"))
    merged["external_url"] = first_non_empty(existing.get("external_url"), rss_ep.get("external_url"))
    merged["summary"] = first_non_empty(existing.get("summary"), rss_ep.get("summary"))
    merged["transcript_url"] = first_non_empty(rss_ep.get("transcript_url"), existing.get("transcript_url"))
    merged["transcript_text"] = first_non_empty(existing.get("transcript_text"), rss_ep.get("transcript_text"))
    merged["key_takeaways"] = existing.get("key_takeaways") or rss_ep.get("key_takeaways") or []
    merged["related_topic"] = existing.get("related_topic") or rss_ep.get("related_topic") or {}
    merged["related_books"] = existing.get("related_books") or rss_ep.get("related_books") or []
    return merged


def trim_to_sentence(text: str, limit: int = 155) -> str:
    cleaned = clean_description(text)
    if len(cleaned) <= limit:
        return cleaned
    candidate = cleaned[:limit].rsplit(" ", 1)[0].strip(" ,;:-")
    if not candidate:
        candidate = cleaned[:limit].strip()
    return candidate + "…"


def fallback_episode_summary(title: str) -> str:
    clean_title = clean_description(title) or "this episode"
    return (
        f"Jonathan Harris examines {clean_title} in plain English, cutting through AI hype and focusing on what the story means for work, policy, business, and everyday users."
    )


def normalise_episode_summary(ep: dict) -> str:
    return first_non_empty(ep.get("summary"), fallback_episode_summary(ep.get("title", "")))


def direct_answer_for_episode(ep: dict) -> str:
    summary = normalise_episode_summary(ep)
    title = clean_description(ep.get("title", "this episode"))
    return (
        f"This episode of {SERIES_NAME} explains {title} through the practical lens Jonathan Harris uses across his AI books and commentary. "
        f"The direct answer: {summary}"
    )


def default_takeaways(ep: dict) -> list[str]:
    title = clean_description(ep.get("title", "the episode"))
    return [
        f"Why {title} matters beyond the usual AI headline noise.",
        "What the story means for businesses, creators, public services, and ordinary users.",
        "Where the technology is useful, where the claims need testing, and what to watch next.",
    ]


def faq_schema_for_episode(ep: dict, canonical: str, summary: str, takeaways: list[str]) -> dict:
    title = clean_description(ep.get("title", "this episode"))
    return {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "mainEntity": [
            {
                "@type": "Question",
                "name": f"What is {title} about?",
                "acceptedAnswer": {"@type": "Answer", "text": summary},
            },
            {
                "@type": "Question",
                "name": "What should listeners take from this episode?",
                "acceptedAnswer": {"@type": "Answer", "text": " ".join(takeaways[:3])},
            },
        ],
        "url": canonical,
    }


def load_existing_episode_list() -> list[dict]:
    if not EPISODES_DATA.exists():
        return []
    try:
        items = json.loads(EPISODES_DATA.read_text(encoding="utf-8"))
    except Exception:
        return []
    return [item for item in items if isinstance(item, dict)]


def load_and_merge_episodes() -> list[dict]:
    existing_lookup = load_existing_episodes()
    try:
        rss_episodes = fetch_rss_episodes()
    except Exception as exc:
        print(f"WARNING: RSS fetch failed - {exc}. Rebuilding episode pages from committed episode data.", file=sys.stderr)
        committed = ensure_unique_episode_slugs(load_existing_episode_list())
        if not committed:
            raise
        persist_episode_data(committed)
        print(f"Normalised {len(committed)} committed episode record(s) in {EPISODES_DATA.relative_to(ROOT)}")
        return committed

    merged = ensure_unique_episode_slugs([merge_episode(ep, existing_lookup) for ep in rss_episodes])
    persist_episode_data(merged)
    print(f"Saved {len(merged)} merged episode record(s) to {EPISODES_DATA.relative_to(ROOT)}")
    return merged


def build_episode_page(ep: dict) -> str:
    title = ep["title"]
    slug = ep["slug"]
    date = ep.get("date", "")
    summary = normalise_episode_summary(ep)
    meta_description = trim_to_sentence(summary, 155)
    takeaways = ep.get("key_takeaways", []) or default_takeaways(ep)
    transcript_text = ep.get("transcript_text", "")
    transcript_url = ep.get("transcript_url", "")
    audio_url = ep.get("audio_url", "")
    external_url = ep.get("external_url", "")
    related_topic = ep.get("related_topic", {}) or {}
    related_books = ep.get("related_books", []) or []
    canonical = ep.get("page_url") or f"{SITE}/podcast/episodes/{slug}/"
    direct_answer = direct_answer_for_episode({**ep, "summary": summary})

    items = "".join(f"<li>{html_mod.escape(t)}</li>" for t in takeaways)
    takeaways_html = f'<section class="card u-s21"><h2>What should you take from this episode?</h2><ul>{items}</ul></section>'

    transcript_html = ""
    if transcript_text:
        paras = "".join(f"<p>{html_mod.escape(p.strip())}</p>" for p in transcript_text.split("\n\n") if p.strip())
        transcript_html = f'<section class="card u-s22"><h2>What does the transcript cover?</h2><div class="episode-transcript">{paras}</div></section>'
    elif transcript_url:
        transcript_html = (
            f'<section class="card u-s22"><h2>Where can I read the transcript?</h2>'
            f'<p><a href="{html_mod.escape(transcript_url)}" rel="noopener noreferrer" target="_blank">Read the full transcript</a></p>'
            f'</section>'
        )

    related_topic_html = ""
    if isinstance(related_topic, dict) and related_topic.get("url"):
        related_topic_html = (
            f'<p>Related topic guide: <a href="{html_mod.escape(related_topic["url"])}">'
            f'{html_mod.escape(related_topic.get("name", "Related topic"))}</a></p>'
        )

    related_books_html = ""
    if related_books:
        links = " · ".join(
            f'<a href="{html_mod.escape(b["url"])}">{html_mod.escape(b["title"])}</a>'
            for b in related_books
            if isinstance(b, dict) and b.get("url") and b.get("title")
        )
        if links:
            related_books_html = f'<section class="card u-s21"><h2>Which Jonathan Harris books connect to this episode?</h2><p>{links}</p></section>'

    schema: dict = {
        "@context": "https://schema.org",
        "@type": "PodcastEpisode",
        "name": title,
        "url": canonical,
        "datePublished": date,
        "description": summary,
        "partOfSeries": {"@type": "PodcastSeries", "name": SERIES_NAME, "url": SERIES_URL},
        "author": {"@id": f"{SITE}/#person"},
    }
    if transcript_url:
        schema["transcript"] = transcript_url
    if audio_url:
        schema["associatedMedia"] = {"@type": "MediaObject", "contentUrl": audio_url}

    breadcrumb = {
        "@context": "https://schema.org",
        "@type": "BreadcrumbList",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "Home", "item": SITE + "/"},
            {"@type": "ListItem", "position": 2, "name": "Podcast", "item": SERIES_URL},
            {"@type": "ListItem", "position": 3, "name": title, "item": canonical},
        ],
    }
    faq_schema = faq_schema_for_episode({**ep, "summary": summary}, canonical, summary, takeaways)

    display_date = date or ""
    listen_href = first_non_empty(audio_url, external_url, SERIES_URL)
    listen_label = "Listen to episode" if listen_href else "Open podcast"
    listen_link = f'<p><a class="button" href="{html_mod.escape(listen_href)}" rel="noopener noreferrer" target="_blank">{listen_label}</a></p>' if listen_href else ""
    audio_player = f'<section class="card u-s21"><h2>How can I listen?</h2><audio controls preload="none" src="{html_mod.escape(audio_url)}"></audio></section>' if audio_url else ""
    related_section = related_topic_html or related_books_html
    if related_topic_html and related_books_html:
        related_section = related_topic_html + related_books_html

    return f"""<!DOCTYPE html>
<html lang="en-GB">
<head>
<meta charset="utf-8"/>
<meta content="width=device-width, initial-scale=1" name="viewport"/>
<title>{html_mod.escape(title)} | Turing&#39;s Torch</title>
<meta content="{html_mod.escape(meta_description)}" name="description"/>
<meta content="#0D1420" name="theme-color"/>
<meta content="index,follow" name="robots"/>
<link crossorigin="" href="https://fonts.gstatic.com" rel="preconnect"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,400;0,600;0,700;0,800&display=swap" rel="stylesheet"/>
<link as="style" href="/assets/css/site.css" rel="preload"/>
<script>document.documentElement.classList.add('js-enabled');</script><link href="/assets/css/site.css" rel="stylesheet"/>
<meta content="article" property="og:type"/>
<meta content="{html_mod.escape(canonical)}" property="og:url"/>
<meta content="{html_mod.escape(title)} | Turing&#39;s Torch" property="og:title"/>
<meta content="{html_mod.escape(meta_description)}" property="og:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" property="og:image"/>
<meta content="summary_large_image" name="twitter:card"/>
<meta content="{html_mod.escape(title)} | Turing&#39;s Torch" name="twitter:title"/>
<meta content="{html_mod.escape(meta_description)}" name="twitter:description"/>
<link href="{html_mod.escape(canonical)}" rel="canonical"/>
<script type="application/ld+json">{json.dumps(schema, ensure_ascii=False)}</script>
<script type="application/ld+json">{json.dumps(breadcrumb, ensure_ascii=False)}</script>
<script type="application/ld+json">{json.dumps(faq_schema, ensure_ascii=False)}</script>
<link href="https://cdn-cookieyes.com" rel="dns-prefetch"/>
<link href="https://tracker.metricool.com" rel="dns-prefetch"/>
<link href="https://botsailor.com" rel="dns-prefetch"/>
<script async="" id="cookieyes" src="https://cdn-cookieyes.com/client_data/c981d18033783598d2216add/script.js" type="text/javascript"></script>
<script defer="" data-cookieyes="ignore" data-cookieconsent="ignore" src="/assets/js/script-governance.min.js"></script>
</head>
<body class="page-podcast-episode jh-no-hero-page">
{load_partial(HEADER_PARTIAL, "header")}
<main class="main" id="main" role="main">
  <div class="wrap ebook-shell">
    <section class="card ebook-index-intro">
      <p class="breadcrumb-hint"><a href="/podcast/">← Podcast</a></p>
      <h1>{html_mod.escape(title)}</h1>
      {f'<p class="episode-meta">Published {html_mod.escape(display_date)}</p>' if display_date else ""}
      <p>{html_mod.escape(summary)}</p>
      {listen_link}
    </section>
    <section class="card u-s21 answer-first">
      <h2>What is this episode about?</h2>
      <p>{html_mod.escape(direct_answer)}</p>
    </section>
    {audio_player}
    {takeaways_html}
    {transcript_html}
    {related_section}
  </div>
</main>
{load_partial(FOOTER_PARTIAL, "footer")}
<script defer="" src="/assets/js/site-ui.min.js"></script>
</body>
</html>"""


def build_compat_redirect(session_id: str, slug: str) -> str:
    canonical = f"/podcast/episodes/{slug}/"
    return f"""<!DOCTYPE html>
<html lang="en-GB">
<head>
<meta charset="utf-8"/>
<meta name="robots" content="noindex,follow"/>
<meta http-equiv="refresh" content="0; url={canonical}"/>
<link rel="canonical" href="{canonical}"/>
<title>Redirecting to podcast episode</title>
<script>window.location.replace({json.dumps(canonical)});</script>
</head>
<body>
<p>Redirecting to <a href="{canonical}">{canonical}</a>.</p>
</body>
</html>"""


def sync_podcast_episode_redirects(episodes: list[dict]) -> int:
    if not REDIRECTS_FILE.exists():
        return 0
    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for ep in episodes:
        session_id = first_non_empty(ep.get("session_id"))
        slug = first_non_empty(ep.get("slug"))
        if not session_id or not slug:
            continue
        target = f"/podcast/episodes/{slug}/"
        for source in (f"/podcast/{session_id}", f"/podcast/{session_id}/"):
            pair = (source, target)
            if pair not in seen:
                seen.add(pair)
                pairs.append(pair)
    if not pairs:
        return 0
    block = "# BEGIN GENERATED PODCAST EPISODE REDIRECTS\n" + "\n".join(
        f"{source}  {target}  301" for source, target in pairs
    ) + "\n# END GENERATED PODCAST EPISODE REDIRECTS"
    text = REDIRECTS_FILE.read_text(encoding="utf-8")
    pattern = re.compile(r"\n?# BEGIN GENERATED PODCAST EPISODE REDIRECTS\n.*?\n# END GENERATED PODCAST EPISODE REDIRECTS", re.S)
    if pattern.search(text):
        text = pattern.sub("\n" + block, text)
    else:
        anchor = "# Transcript archive hardening"
        if anchor in text:
            text = text.replace(anchor, block + "\n\n" + anchor, 1)
        else:
            text = text.rstrip() + "\n\n" + block + "\n"
    REDIRECTS_FILE.write_text(text.rstrip() + "\n", encoding="utf-8")
    return len(pairs)

def _validate_transcript_archive_contract(path: Path, label: str) -> list[str]:
    if not path.exists():
        return [f"Missing {label}: {path}"]

    src = path.read_text(encoding="utf-8")
    failures: list[str] = []

    if 'id="transcriptSearch"' not in src:
        failures.append(f"Transcript search input missing from {label}")

    transcript_list_match = re.search(r'<ul[^>]*id="transcriptList"[^>]*>(?P<body>.*?)</ul>', src, flags=re.DOTALL)
    if not transcript_list_match:
        failures.append(f"Transcript list missing from {label}")
    else:
        count = transcript_list_match.group("body").count("<li")
        if count < 4:
            failures.append(f"{label} should contain at least 4 transcript entries, found {count}")

    if 'class="transcript-list"' not in src:
        failures.append(f"Transcript archive class missing from {label}")

    return failures


def _validate_podcast_transcript_route(path: Path, label: str) -> list[str]:
    if not path.exists():
        return [f"Missing {label}: {path}"]

    src = path.read_text(encoding="utf-8")
    failures: list[str] = []

    if 'elfsight-app-76cc65a0-0bcf-4dc0-ad36-1046c5a20e3d' not in src:
        failures.append("Elfsight podcast player embed missing from podcast/index.html")

    if 'href="/transcripts/"' not in src:
        failures.append("Transcript archive route missing from podcast/index.html")

    if 'Transcript Archive' not in src:
        failures.append("Transcript archive CTA missing from podcast/index.html")

    if 'id="transcriptSearch"' in src or 'id="transcriptList"' in src or 'class="transcript-list"' in src:
        failures.append(
            "podcast/index.html should link to the transcript archive instead of embedding the searchable transcript list"
        )

    return failures


def validate_podcast_index_contract() -> list[str]:
    failures = _validate_podcast_transcript_route(HTML_INDEX, "podcast/index.html")
    failures.extend(_validate_transcript_archive_contract(TRANSCRIPTS_INDEX, "transcripts/index.html"))
    return failures


HTML_INDEX = ROOT / "podcast" / "index.html"
TRANSCRIPTS_INDEX = ROOT / "transcripts" / "index.html"


def validate_generated_pages(episodes: list[dict]) -> list[str]:
    failures: list[str] = []
    expected_recent = min(4, len(episodes))
    failures.extend(validate_podcast_index_contract())

    for ep in episodes[:expected_recent]:
        slug = ep.get("slug", "")
        if not slug:
            failures.append(f"Episode missing slug: {ep.get('title', '(untitled)')}")
            continue

        page_path = EPISODES_DIR / slug / "index.html"
        if not page_path.exists():
            failures.append(f"Missing generated page: {page_path.relative_to(ROOT)}")
            continue

        page_html = page_path.read_text(encoding="utf-8")
        audio_url = (ep.get("audio_url") or "").strip()
        if audio_url:
            if '<audio' not in page_html or 'controls' not in page_html:
                failures.append(f"Audio player missing from episode page: {slug}")
            if html_mod.escape(audio_url) not in page_html:
                failures.append(f"Audio URL missing from episode page: {slug}")

        transcript_url = (ep.get("transcript_url") or "").strip()
        if transcript_url and html_mod.escape(transcript_url) not in page_html and "<h2>Transcript</h2>" not in page_html:
            failures.append(f"Transcript handling missing from episode page: {slug}")

        session_id = (ep.get("session_id") or "").strip()
        if session_id:
            compat_path = PODCAST_COMPAT_DIR / session_id / "index.html"
            if not compat_path.exists():
                failures.append(f"Missing compatibility redirect: {compat_path.relative_to(ROOT)}")

    return failures


def register_in_workbook(slugs_generated: list[str]) -> None:
    wb_candidates = sorted(ROOT.glob("*.xlsm")) + sorted(ROOT.glob("*.xlsx"))
    if not wb_candidates:
        print("  [WARN] No workbook found — skipping workbook registration.")
        return
    wb_path = wb_candidates[0]
    wb = openpyxl.load_workbook(wb_path, keep_vba=True)
    if "Pages" not in wb.sheetnames:
        print("  [WARN] Workbook has no Pages sheet — skipping registration.")
        return
    ws = wb["Pages"]

    governed: set[str] = set()
    last_row = PAGE_HEADER_ROW
    for row in range(PAGE_HEADER_ROW + 1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val:
            governed.add(str(val).strip())
            last_row = row

    next_row = last_row + 1
    added = 0
    for slug in sorted(set(slugs_generated)):
        rel = f"podcast/episodes/{slug}/index.html"
        if rel in governed:
            continue
        page = ROOT / rel
        if not page.exists():
            continue
        title_m = re.search(r"<title>(.*?)</title>", page.read_text(encoding="utf-8"))
        title = html_mod.unescape(title_m.group(1).strip()) if title_m else slug
        url_path = f"/podcast/episodes/{slug}/"
        ws.cell(next_row, 1).value = rel
        ws.cell(next_row, 2).value = url_path
        ws.cell(next_row, 3).value = SITE + url_path
        ws.cell(next_row, 4).value = title
        ws.cell(next_row, 5).value = "Podcast episode"
        print(f"  Registered in workbook: {rel}")
        next_row += 1
        added += 1

    if added:
        wb.save(wb_path)
        print(f"  Workbook saved — {added} new episode(s) registered.")
    else:
        print("  Workbook: all episode pages already governed.")


def main() -> int:
    try:
        episodes = load_and_merge_episodes()
    except Exception as exc:
        print(f"Failed to fetch or merge podcast episodes: {exc}", file=sys.stderr)
        return 1

    if not episodes:
        print("No episode data found.", file=sys.stderr)
        return 1

    EPISODES_DIR.mkdir(parents=True, exist_ok=True)
    created = 0
    generated_slugs: list[str] = []
    for ep in episodes:
        slug = ep.get("slug")
        if not slug:
            continue
        ep_dir = EPISODES_DIR / slug
        ep_dir.mkdir(parents=True, exist_ok=True)
        (ep_dir / "index.html").write_text(build_episode_page(ep), encoding="utf-8")
        created += 1
        generated_slugs.append(slug)
        print(f"  Generated: podcast/episodes/{slug}/")

        session_id = ep.get("session_id")
        if session_id:
            compat_dir = PODCAST_COMPAT_DIR / session_id
            compat_dir.mkdir(parents=True, exist_ok=True)
            (compat_dir / "index.html").write_text(build_compat_redirect(session_id, slug), encoding="utf-8")
            print(f"  Generated compatibility redirect: podcast/{session_id}/")

    print(f"Generated {created} episode page(s) under podcast/episodes/")
    redirect_count = sync_podcast_episode_redirects(episodes)
    if redirect_count:
        print(f"  Synced {redirect_count} podcast compatibility redirect rule(s) in _redirects.")

    validation_failures = validate_generated_pages(episodes)
    if validation_failures:
        for failure in validation_failures:
            print(f"[FAIL] {failure}", file=sys.stderr)
        return 1

    print("Podcast episode generation checks passed.")
    register_in_workbook(generated_slugs)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
