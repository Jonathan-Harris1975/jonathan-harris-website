#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import re
import sys
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib import error, request

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.ebook_pipeline import ROOT, SITE_URL, clean_paragraph, load_master  # noqa: E402

DEFAULT_TIMEOUT = 15.0
DEFAULT_RETRIES = 2
DEFAULT_RETRY_DELAY_SECONDS = 2.0
DEFAULT_USER_AGENT = "JonathanHarrisPageSmoke/1.0 (+https://jonathan-harris.online)"
PAGE_HEADER_ROW = 5
SCRIPT_RE = re.compile(r"<script\b.*?</script>", re.I | re.S)
STYLE_RE = re.compile(r"<style\b.*?</style>", re.I | re.S)
TAG_RE = re.compile(r"<[^>]+>")
WHITESPACE_RE = re.compile(r"\s+")
TITLE_RE = re.compile(r"<title>(.*?)</title>", re.I | re.S)
LINK_TAG_RE = re.compile(r"<link\b[^>]*>", re.I)
REL_ATTR_RE = re.compile(r'\brel=["\']([^"\']+)["\']', re.I)
HREF_ATTR_RE = re.compile(r'\bhref=["\']([^"\']+)["\']', re.I)
H1_RE = re.compile(r"<h1[^>]*>(.*?)</h1>", re.I | re.S)


@dataclass
class PageCheck:
    label: str
    local_path: Path
    live_url: str
    extractors: list[tuple[str, str]]


@dataclass
class HtmlContractCheck:
    label: str
    local_path: Path
    live_url: str
    expected_status: int = 200


@dataclass
class PageResult:
    label: str
    ok: bool
    message: str
    live_url: str
    missing_markers: list[str]
    status_code: int | None = None


@dataclass
class FetchResult:
    body: str
    status_code: int | None
    final_url: str | None
    error: str | None = None


def fetch_url(url: str, *, timeout: float) -> FetchResult:
    """Fetch a live URL for CI smoke checks without letting transient network errors crash the gate."""
    last_error: str | None = None

    for attempt in range(1, DEFAULT_RETRIES + 2):
        req = request.Request(
            url,
            headers={
                "User-Agent": DEFAULT_USER_AGENT,
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Accept": "text/html,application/xhtml+xml,application/json;q=0.9,*/*;q=0.8",
            },
            method="GET",
        )

        try:
            with request.urlopen(req, timeout=timeout) as response:
                return FetchResult(
                    body=response.read().decode("utf-8", errors="replace"),
                    status_code=getattr(response, "status", None) or response.getcode(),
                    final_url=response.geturl(),
                )
        except error.HTTPError as exc:
            return FetchResult(
                body=exc.read().decode("utf-8", errors="replace"),
                status_code=exc.code,
                final_url=getattr(exc, "url", url),
                error=f"HTTP {exc.code} for {url}",
            )
        except TimeoutError:
            last_error = f"Request timed out for {url} after {timeout:g}s"
        except error.URLError as exc:
            last_error = f"Request failed for {url}: {exc.reason}"
        except OSError as exc:
            last_error = f"Request failed for {url}: {exc}"

        if attempt <= DEFAULT_RETRIES:
            time.sleep(DEFAULT_RETRY_DELAY_SECONDS)

    return FetchResult(body="", status_code=None, final_url=None, error=last_error or f"Request failed for {url}")


def normalise_visible_text(raw_html: str) -> str:
    without_scripts = SCRIPT_RE.sub(" ", raw_html)
    without_styles = STYLE_RE.sub(" ", without_scripts)
    stripped = TAG_RE.sub(" ", without_styles)
    return WHITESPACE_RE.sub(" ", html.unescape(stripped)).strip()


def normalise_fragment(raw_html: str) -> str:
    stripped = TAG_RE.sub(" ", raw_html or "")
    return WHITESPACE_RE.sub(" ", html.unescape(stripped)).strip()


def extract_required_snippets(page: PageCheck) -> list[str]:
    source = page.local_path.read_text(encoding="utf-8")
    snippets: list[str] = []
    for label, pattern in page.extractors:
        match = re.search(pattern, source, flags=re.I | re.S)
        if not match:
            raise ValueError(f"Could not extract {label} from {page.local_path.relative_to(ROOT)}")
        snippet = normalise_fragment(match.group(1))
        if snippet:
            snippets.append(snippet)
    return snippets


def extract_canonical_href(raw_html: str) -> str | None:
    for link_tag in LINK_TAG_RE.findall(raw_html):
        rel_match = REL_ATTR_RE.search(link_tag)
        if not rel_match:
            continue
        rel_values = {value.strip().lower() for value in rel_match.group(1).split()}
        if "canonical" not in rel_values:
            continue
        href_match = HREF_ATTR_RE.search(link_tag)
        if href_match:
            return clean_paragraph(href_match.group(1))
    return None


def extract_contract_markers(raw_html: str, *, label: str) -> dict[str, str]:
    markers: dict[str, str] = {}
    title_match = TITLE_RE.search(raw_html)
    if not title_match:
        raise ValueError(f"Could not extract title from {label}")
    markers["title"] = normalise_fragment(title_match.group(1))

    canonical_href = extract_canonical_href(raw_html)
    if not canonical_href:
        raise ValueError(f"Could not extract canonical href from {label}")
    markers["canonical"] = canonical_href

    h1_match = H1_RE.search(raw_html)
    if not h1_match:
        raise ValueError(f"Could not extract h1 from {label}")
    markers["h1"] = normalise_fragment(h1_match.group(1))
    return markers


def load_workbook_page_checks(workbook_path: Path) -> list[HtmlContractCheck]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Pages" not in wb.sheetnames:
        raise ValueError("Workbook is missing the Pages sheet")

    ws = wb["Pages"]
    headers = {
        clean_paragraph(ws.cell(PAGE_HEADER_ROW, col).value).lower(): col
        for col in range(1, ws.max_column + 1)
        if clean_paragraph(ws.cell(PAGE_HEADER_ROW, col).value)
    }
    required_headers = {"relative file path", "public url path", "full url"}
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise ValueError(f"Pages sheet is missing required columns: {', '.join(missing_headers)}")

    checks: list[HtmlContractCheck] = []
    for row_idx in range(PAGE_HEADER_ROW + 1, ws.max_row + 1):
        relative_path = clean_paragraph(ws.cell(row_idx, headers["relative file path"]).value)
        if not relative_path or not relative_path.endswith(".html"):
            continue
        local_path = ROOT / relative_path
        if not local_path.exists():
            continue
        public_path = clean_paragraph(ws.cell(row_idx, headers["public url path"]).value)
        full_url = clean_paragraph(ws.cell(row_idx, headers["full url"]).value) or f"{SITE_URL}{public_path}"
        # The workbook may govern /404 as a published compatibility route that
        # returns HTTP 200, while the real unknown-route 404 behaviour is
        # verified separately by run_not_found_contract_checks().
        expected_status = 200
        checks.append(
            HtmlContractCheck(
                label=relative_path,
                local_path=local_path,
                live_url=full_url,
                expected_status=expected_status,
            )
        )
    if not checks:
        raise ValueError("Pages sheet did not yield any governed HTML checks")
    return checks


def selected_page_checks() -> list[PageCheck]:
    books = {book["slug"]: book for book in load_master()}
    key_book_slug = "the-dumbening-how-ai-is-reshaping-our-minds"
    ai_book_slug = "the-artificial-intelligence-revolution-from-algorithms-to-consciousness"
    pages = [
        PageCheck(
            label="homepage featured merch block",
            local_path=ROOT / "index.html",
            live_url=f"{SITE_URL}/",
            extractors=[
                ("featured title", r'id="featuredEbookTitle">(.*?)</h3>'),
                ("featured description", r'id="featuredEbookDesc">(.*?)</p>'),
            ],
        ),
        PageCheck(
            label="The Dumbening canonical page",
            local_path=ROOT / "ebooks" / key_book_slug / "index.html",
            live_url=books[key_book_slug]["canonical_url"],
            extractors=[
                ("book title", r"<h1>(.*?)</h1>"),
                ("hero summary", r"<h1>.*?</h1>\s*<p>(.*?)</p>"),
                ("faq audience", r"<summary>Who is this book for\?</summary><div><p>(.*?)</p></div>"),
            ],
        ),
        PageCheck(
            label="Artificial intelligence topic page",
            local_path=ROOT / "catalogue" / "artificial-intelligence" / "index.html",
            live_url=f"{SITE_URL}/catalogue/artificial-intelligence/",
            extractors=[
                ("topic heading", r"<h1>(.*?)</h1>"),
                ("topic intro", r"<h1>.*?</h1>\s*<p>(.*?)</p>"),
            ],
        ),
        PageCheck(
            label="Artificial intelligence core book page",
            local_path=ROOT / "ebooks" / ai_book_slug / "index.html",
            live_url=books[ai_book_slug]["canonical_url"],
            extractors=[
                ("book title", r"<h1>(.*?)</h1>"),
                ("faq audience", r"<summary>Who is this book for\?</summary><div><p>(.*?)</p></div>"),
            ],
        ),
    ]
    return pages


def run_selected_page_checks(*, timeout: float) -> list[PageResult]:
    results: list[PageResult] = []
    for page in selected_page_checks():
        fetch_result = fetch_url(page.live_url, timeout=timeout)
        if fetch_result.error or fetch_result.status_code != 200:
            results.append(
                PageResult(
                    page.label,
                    False,
                    fetch_result.error or f"Unexpected HTTP {fetch_result.status_code} for {page.live_url}",
                    page.live_url,
                    [],
                    status_code=fetch_result.status_code,
                )
            )
            continue

        live_text = normalise_visible_text(fetch_result.body)
        snippets = extract_required_snippets(page)
        missing = [snippet for snippet in snippets if snippet not in live_text]
        if missing:
            results.append(
                PageResult(
                    label=page.label,
                    ok=False,
                    message="Live page is reachable but missing one or more governed content markers",
                    live_url=page.live_url,
                    missing_markers=missing,
                    status_code=fetch_result.status_code,
                )
            )
            continue
        results.append(PageResult(page.label, True, "Live page contains the governed markers", page.live_url, [], status_code=fetch_result.status_code))
    return results


def run_workbook_page_checks(*, workbook_path: Path, timeout: float) -> list[PageResult]:
    results: list[PageResult] = []
    for page in load_workbook_page_checks(workbook_path):
        fetch_result = fetch_url(page.live_url, timeout=timeout)
        if fetch_result.error or fetch_result.status_code != page.expected_status:
            results.append(
                PageResult(
                    page.label,
                    False,
                    fetch_result.error or f"Unexpected HTTP {fetch_result.status_code} for {page.live_url}; expected {page.expected_status}",
                    page.live_url,
                    [],
                    status_code=fetch_result.status_code,
                )
            )
            continue

        expected_markers = extract_contract_markers(page.local_path.read_text(encoding="utf-8"), label=str(page.local_path.relative_to(ROOT)))
        live_markers = extract_contract_markers(fetch_result.body, label=page.live_url)
        mismatches = [
            f"{field}: expected '{expected_markers[field]}' but saw '{live_markers.get(field, '')}'"
            for field in expected_markers
            if live_markers.get(field) != expected_markers[field]
        ]
        if mismatches:
            results.append(
                PageResult(
                    label=page.label,
                    ok=False,
                    message="Live page returned the expected status but drifted from the governed HTML contract",
                    live_url=page.live_url,
                    missing_markers=mismatches,
                    status_code=fetch_result.status_code,
                )
            )
            continue

        results.append(
            PageResult(
                label=page.label,
                ok=True,
                message="Live page matches the governed title, canonical, and H1 contract",
                live_url=page.live_url,
                missing_markers=[],
                status_code=fetch_result.status_code,
            )
        )
    return results


def run_not_found_contract_checks(*, timeout: float, include_explicit_route: bool = True) -> list[PageResult]:
    checks = []
    if include_explicit_route:
        checks.append(
            HtmlContractCheck(
                label="Explicit /404 route contract",
                local_path=ROOT / "404.html",
                live_url=f"{SITE_URL}/404",
                expected_status=404,
            )
        )
    checks.append(
        HtmlContractCheck(
            label="Unknown-route 404 contract",
            local_path=ROOT / "404.html",
            live_url=f"{SITE_URL}/__release-smoke-missing-{uuid.uuid4().hex}/",
            expected_status=404,
        )
    )

    results: list[PageResult] = []
    expected_markers = extract_contract_markers((ROOT / "404.html").read_text(encoding="utf-8"), label="404.html")
    for page in checks:
        fetch_result = fetch_url(page.live_url, timeout=timeout)
        if fetch_result.status_code != page.expected_status:
            results.append(
                PageResult(
                    page.label,
                    False,
                    fetch_result.error or f"Expected HTTP {page.expected_status} but received {fetch_result.status_code}",
                    page.live_url,
                    [],
                    status_code=fetch_result.status_code,
                )
            )
            continue
        live_markers = extract_contract_markers(fetch_result.body, label=page.live_url)
        mismatches = [
            f"{field}: expected '{expected_markers[field]}' but saw '{live_markers.get(field, '')}'"
            for field in expected_markers
            if live_markers.get(field) != expected_markers[field]
        ]
        if mismatches:
            results.append(
                PageResult(
                    label=page.label,
                    ok=False,
                    message="Live 404 response is present but drifted from the governed 404 contract",
                    live_url=page.live_url,
                    missing_markers=mismatches,
                    status_code=fetch_result.status_code,
                )
            )
            continue
        results.append(PageResult(page.label, True, "Live 404 contract is correct", page.live_url, [], status_code=fetch_result.status_code))
    return results


def run_api_contract_check(*, timeout: float) -> PageResult:
    live_url = f"{SITE_URL}/api/v1/books.json"
    fetch_result = fetch_url(live_url, timeout=timeout)
    if fetch_result.error or fetch_result.status_code != 200:
        return PageResult(
            label="Public books API parity",
            ok=False,
            message=fetch_result.error or f"Unexpected HTTP {fetch_result.status_code} for {live_url}",
            live_url=live_url,
            missing_markers=[],
            status_code=fetch_result.status_code,
        )

    try:
        live_payload = json.loads(fetch_result.body)
    except json.JSONDecodeError as exc:
        return PageResult(
            label="Public books API parity",
            ok=False,
            message=f"Live API returned invalid JSON: {exc}",
            live_url=live_url,
            missing_markers=[],
            status_code=fetch_result.status_code,
        )

    governed_payload = json.loads((ROOT / "api" / "v1" / "books.json").read_text(encoding="utf-8"))
    if live_payload != governed_payload:
        return PageResult(
            label="Public books API parity",
            ok=False,
            message="Live API payload does not match the governed repo artifact",
            live_url=live_url,
            missing_markers=[],
            status_code=fetch_result.status_code,
        )
    return PageResult(
        label="Public books API parity",
        ok=True,
        message="Live API payload matches the governed repo artifact",
        live_url=live_url,
        missing_markers=[],
        status_code=fetch_result.status_code,
    )



def run_growth_surface_checks(*, timeout: float) -> list[PageResult]:
    """Fail a release when the high-value live surfaces drift from growth contracts."""
    results: list[PageResult] = []
    book_count = len(load_master())

    def result(label: str, path: str, failures: list[str], ok_message: str) -> None:
        live_url = f"{SITE_URL}{path}"
        fetched = fetch_url(live_url, timeout=timeout)
        if fetched.error or fetched.status_code != 200:
            results.append(PageResult(label, False, fetched.error or f"Unexpected HTTP {fetched.status_code}", live_url, [], status_code=fetched.status_code))
            return
        raw = fetched.body
        visible = normalise_visible_text(raw)
        found = [failure for failure in failures if failure]
        # Callers pass sentinel expressions evaluated against their fetched body
        # through the helper-specific closures below, so this branch is unused.
        results.append(PageResult(label, not found, ok_message if not found else "Live growth contract failed", live_url, found, status_code=fetched.status_code))

    # Homepage: generated catalogue count and current evidence routing.
    home_url = f"{SITE_URL}/"
    home = fetch_url(home_url, timeout=timeout)
    if home.error or home.status_code != 200:
        results.append(PageResult("Homepage growth contract", False, home.error or f"Unexpected HTTP {home.status_code}", home_url, [], status_code=home.status_code))
    else:
        visible = normalise_visible_text(home.body)
        failures = []
        if re.search(r"\b36(?:-book|\s+(?:plain-English\s+)?(?:AI\s+)?(?:books|ebooks))\b", visible, re.I):
            failures.append("stale 36-book catalogue copy remains live")
        if str(book_count) not in visible:
            failures.append(f"generated catalogue count {book_count} is not visible")
        if "/evidence/eu-ai-act-article-50-transparency/" not in home.body:
            failures.append("Article 50 evidence route is not surfaced on the homepage")
        results.append(PageResult(
            "Homepage growth contract",
            not failures,
            "Homepage catalogue/evidence contract is current" if not failures else "Homepage growth contract failed",
            home_url,
            failures,
            status_code=home.status_code,
        ))

    # AI Edge: one collection system and no cadence/read-time promise.
    news_url = f"{SITE_URL}/newsletter/"
    news = fetch_url(news_url, timeout=timeout)
    if news.error or news.status_code != 200:
        results.append(PageResult("AI Edge signup contract", False, news.error or f"Unexpected HTTP {news.status_code}", news_url, [], status_code=news.status_code))
    else:
        visible = normalise_visible_text(news.body)
        failures = []
        if "form.jotform.com/260277027608054" not in news.body:
            failures.append("governed AI Edge Jotform is missing")
        if "/api/newsletter/subscribe" in news.body or "newsletter-native-form" in news.body or "data-newsletter-form" in news.body:
            failures.append("a second/native newsletter collection path is still exposed")
        forbidden = re.search(r"(?:three[- ]minute|3[- ]minute|weekday|daily\s+(?:AI\s+)?(?:newsletter|briefing))", visible, re.I)
        if forbidden:
            failures.append(f"newsletter time/cadence wording remains live: {forbidden.group(0)}")
        results.append(PageResult(
            "AI Edge signup contract",
            not failures,
            "AI Edge uses one Jotform path with no time promise" if not failures else "AI Edge signup contract failed",
            news_url,
            failures,
            status_code=news.status_code,
        ))

    # Podcast: current episodes must be crawlable before any third-party enhancement.
    pod_url = f"{SITE_URL}/podcast/"
    pod = fetch_url(pod_url, timeout=timeout)
    if pod.error or pod.status_code != 200:
        results.append(PageResult("Podcast crawlable contract", False, pod.error or f"Unexpected HTTP {pod.status_code}", pod_url, [], status_code=pod.status_code))
    else:
        failures = []
        if "data-episode-slug=" not in pod.body:
            failures.append("no server-visible current episode card was returned")
        if re.search(r"Latest episode details are temporarily unavailable|loading current episode|Loading the latest Turing", pod.body, re.I):
            failures.append("podcast placeholder copy is still being served")
        if "elfsight-app-76cc65a0-0bcf-4dc0-ad36-1046c5a20e3d" not in pod.body or "https://elfsightcdn.com/platform.js" not in pod.body:
            failures.append("Elfsight six-episode player is missing")
        if re.search(r'<details[^>]*class=["\'][^"\']*podcast-archive-widget', pod.body, re.I):
            failures.append("Elfsight player is still hidden inside a disclosure")
        results.append(PageResult(
            "Podcast crawlable contract",
            not failures,
            (
                "Podcast returns current crawlable episodes plus the visible six-episode player"
                if not failures
                else "Podcast crawlable contract failed"
            ),
            pod_url,
            failures,
            status_code=pod.status_code,
        ))

    return results

def run_checks(*, timeout: float, workbook_path: Path | None = None) -> list[PageResult]:
    if workbook_path:
        results = run_workbook_page_checks(workbook_path=workbook_path, timeout=timeout)
        results.extend(run_not_found_contract_checks(timeout=timeout, include_explicit_route=False))
    else:
        results = run_selected_page_checks(timeout=timeout)
        results.extend(run_not_found_contract_checks(timeout=timeout))
    results.append(run_api_contract_check(timeout=timeout))
    results.extend(run_growth_surface_checks(timeout=timeout))
    return results


def print_results(results: Iterable[PageResult]) -> None:
    for result in results:
        status = "PASS" if result.ok else "FAIL"
        status_text = f" HTTP {result.status_code}" if result.status_code is not None else ""
        print(f"[{status}] {result.label}:{status_text} {result.message}")
        print(f"      {result.live_url}")
        for marker in result.missing_markers[:3]:
            print(f"      missing: {marker}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare governed live page markers, 404 behaviour, and API parity against the published site.")
    parser.add_argument("--workbook", help="Optional workbook path. When supplied, validate every governed HTML route from the Pages sheet.")
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT, help="HTTP timeout in seconds. Default: 15")
    parser.add_argument("--strict", action="store_true", help="Exit non-zero if any governed page, 404 contract, or API parity check fails.")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve() if args.workbook else None
    results = run_checks(timeout=max(args.timeout, 1.0), workbook_path=workbook_path)
    print_results(results)
    failures = [result for result in results if not result.ok]
    if failures and args.strict:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
