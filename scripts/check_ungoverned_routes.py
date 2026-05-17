#!/usr/bin/env python3
"""
check_ungoverned_routes.py
Release gate: fails when any governed routed HTML file in the repo is absent
from the workbook Pages sheet.

Excludes:
  - non-routed source / utility directories
  - hidden directories
  - 404.html
  - podcast compatibility redirects
  - generated dynamic leaves only when they are present in data/dynamic-route-manifest.json

Called from validate_release.py with a workbook path.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PAGE_HEADER_ROW = 5

EXCLUDED_PREFIXES = (
    "assets/",
    "scripts/",
    "functions/",
    "config/",
    "docs/",
    "data/",
    ".github/",
    "node_modules/",
)

EXCLUDED_FILENAMES = {"404.html"}

COMPAT_REDIRECT_PREFIX = "podcast/TT-"

DYNAMIC_ROUTE_MANIFEST = ROOT / "data" / "dynamic-route-manifest.json"


def dynamic_manifest_paths() -> set[str]:
    try:
        payload = json.loads(DYNAMIC_ROUTE_MANIFEST.read_text(encoding="utf-8"))
    except Exception:
        return set()
    routes = payload.get("routes") if isinstance(payload, dict) else []
    governed: set[str] = set()
    for route in routes if isinstance(routes, list) else []:
        if not isinstance(route, dict):
            continue
        repo_path = str(route.get("repo_path") or "").strip()
        if repo_path.endswith(".html"):
            governed.add(repo_path)
    return governed


def should_exclude_route(rel: Path) -> bool:
    rel_str = rel.as_posix()

    if rel_str.startswith(COMPAT_REDIRECT_PREFIX) and rel.name == "index.html":
        return True

    return False


def routed_html_files(root: Path) -> list[Path]:
    results: list[Path] = []

    for p in root.rglob("*.html"):
        rel = p.relative_to(root)
        rel_str = rel.as_posix()

        if any(rel_str.startswith(prefix) for prefix in EXCLUDED_PREFIXES):
            continue
        if any(part.startswith(".") for part in rel.parts):
            continue
        if p.name in EXCLUDED_FILENAMES:
            continue
        if should_exclude_route(rel):
            continue

        results.append(p)

    return sorted(results)


def governed_paths(workbook_path: Path) -> set[str]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    if "Pages" not in wb.sheetnames:
        raise ValueError("Workbook is missing the Pages sheet")

    ws = wb["Pages"]
    paths: set[str] = set()

    for row in range(PAGE_HEADER_ROW + 1, ws.max_row + 1):
        val = ws.cell(row, 1).value
        if val and str(val).strip().endswith(".html"):
            paths.add(str(val).strip())

    return paths


def run_checks(workbook_path: Path) -> list[str]:
    governed = governed_paths(workbook_path)
    governed.update(dynamic_manifest_paths())
    errors: list[str] = []

    for html_path in routed_html_files(ROOT):
        rel = html_path.relative_to(ROOT).as_posix()
        if rel not in governed:
            errors.append(
                f"Ungoverned route: {rel} exists in the repo but is absent from the workbook Pages sheet and data/dynamic-route-manifest.json."
            )

    return errors


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser(
        description="Fail when governed routed HTML files are absent from the workbook Pages sheet."
    )
    parser.add_argument("--workbook", required=True, help="Path to the .xlsm workbook")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        print(f"ERROR: Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    errors = run_checks(workbook_path)
    if errors:
        for e in errors:
            print(f"[FAIL] {e}")
        print(f"\nUngoverned route check failed: {len(errors)} unregistered governed route(s).")
        return 1

    print("[PASS] Ungoverned route check: all governed routed HTML files are governed in the workbook.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
