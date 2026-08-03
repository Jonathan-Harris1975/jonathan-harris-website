#!/usr/bin/env python3
from __future__ import annotations

import json
import mimetypes
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import boto3
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EXCLUDES = ["/podcast", "/blog"]


@dataclass(slots=True)
class WorkbookInfo:
  path: str
  sheet_names: list[str]
  header_row: int | None
  primary_sheet: str | None
  url_count: int
  first_rows: list[list[Any]]
  urls: list[str]
  rows: list[dict[str, Any]]


def utc_now() -> str:
  return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def safe_slug(value: str) -> str:
  slug = re.sub(r"[^a-zA-Z0-9._-]+", "-", str(value).strip()).strip("-._")
  return slug or "audit"


def ensure_dir(path: Path) -> Path:
  path.mkdir(parents=True, exist_ok=True)
  return path


def write_json(path: Path, payload: Any) -> Path:
  ensure_dir(path.parent)
  def audit_json_default(value: Any) -> Any:
    if isinstance(value, set):
      return sorted(value, key=lambda item: str(item))
    if isinstance(value, Path):
      return str(value)
    raise TypeError(f"Object of type {value.__class__.__name__} is not JSON serializable")

  path.write_text(json.dumps(payload, indent=2, ensure_ascii=False, default=audit_json_default), encoding="utf-8")
  return path


def write_text(path: Path, text: str) -> Path:
  ensure_dir(path.parent)
  path.write_text(text, encoding="utf-8")
  return path


def find_workbook(repo_root: Path = REPO_ROOT) -> Path:
  candidates = sorted(repo_root.glob("*site-url-inventory*.xls*"))
  if not candidates:
    raise FileNotFoundError("Workbook not found in repo root")
  return candidates[0]


HEADER_ALIASES: dict[str, set[str]] = {
  "full_url": {"url", "page url", "live url", "public url", "full url"},
  "public_url_path": {"public url path", "url path", "public path", "route path"},
  "relative_file_path": {"relative file path", "file path", "relative path"},
  "page_title": {"page title", "title"},
  "notes": {"notes"},
}


def _find_header_row(rows: list[list[Any]]) -> tuple[int | None, dict[str, int]]:
  for idx, row in enumerate(rows[:12], start=1):
    normalised = [str(cell).strip().lower() if cell is not None else "" for cell in row]
    mapping: dict[str, int] = {}
    for col_index, value in enumerate(normalised):
      for key, aliases in HEADER_ALIASES.items():
        if value in aliases and key not in mapping:
          mapping[key] = col_index
    if any(key in mapping for key in ("full_url", "public_url_path", "relative_file_path")):
      return idx, mapping
  return None, {}


def load_workbook_info(path: Path) -> WorkbookInfo:
  wb = load_workbook(path, read_only=True, data_only=True)
  sheet_names = list(wb.sheetnames)
  primary_sheet = "Pages" if "Pages" in wb.sheetnames else (sheet_names[0] if sheet_names else None)
  if primary_sheet is None:
    return WorkbookInfo(str(path), [], None, None, 0, [], [], [])

  ws = wb[primary_sheet]

  # Some generated workbooks omit the worksheet dimension metadata. In
  # openpyxl read-only mode that leaves max_row/max_column as None and causes
  # min(ws.max_row, 12) to raise before either audit can start. Force a bounded
  # dimension scan once, then continue with the normal header probe.
  if not isinstance(ws.max_row, int) or ws.max_row < 1:
    ws.calculate_dimension(force=True)
  max_row = ws.max_row if isinstance(ws.max_row, int) and ws.max_row > 0 else 12

  rows = list(ws.iter_rows(min_row=1, max_row=min(max_row, 12), values_only=True))
  header_row, header_map = _find_header_row(rows)

  discovered_urls: list[str] = []
  discovered_rows: list[dict[str, Any]] = []
  if header_row:
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
      relative_file_path = ""
      public_url_path = ""
      full_url = ""
      page_title = ""
      notes = ""

      if (idx := header_map.get("relative_file_path")) is not None and idx < len(row) and row[idx] is not None:
        relative_file_path = str(row[idx]).strip()
      if (idx := header_map.get("public_url_path")) is not None and idx < len(row) and row[idx] is not None:
        public_url_path = str(row[idx]).strip()
      if (idx := header_map.get("full_url")) is not None and idx < len(row) and row[idx] is not None:
        full_url = str(row[idx]).strip()
      if (idx := header_map.get("page_title")) is not None and idx < len(row) and row[idx] is not None:
        page_title = str(row[idx]).strip()
      if (idx := header_map.get("notes")) is not None and idx < len(row) and row[idx] is not None:
        notes = str(row[idx]).strip()

      value = full_url or public_url_path or relative_file_path
      if not value:
        continue

      discovered_urls.append(value)
      discovered_rows.append({
        "relative_file_path": relative_file_path,
        "public_url_path": public_url_path,
        "full_url": full_url,
        "page_title": page_title,
        "notes": notes,
      })

  first_rows = [list(row) for row in rows[:5]]
  return WorkbookInfo(
    path=str(path),
    sheet_names=sheet_names,
    header_row=header_row,
    primary_sheet=primary_sheet,
    url_count=len(discovered_urls),
    first_rows=first_rows,
    urls=discovered_urls,
    rows=discovered_rows,
  )


def normalise_route(value: str, base_url: str | None = None) -> str:
  raw = str(value or "").strip()
  if not raw:
    return "/"
  if raw.startswith("http://") or raw.startswith("https://"):
    parsed = urlparse(raw)
    raw = parsed.path or "/"
  elif base_url and raw.startswith(base_url):
    parsed = urlparse(raw)
    raw = parsed.path or "/"
  if not raw.startswith("/"):
    raw = f"/{raw}"
  if raw != "/" and raw.endswith("index.html"):
    raw = raw[: -len("index.html")]
  if raw == "/404.html":
    raw = "/404"
  if raw != "/" and raw.endswith("/"):
    raw = raw[:-1]
  return raw or "/"


def route_to_url(base_url: str, route: str) -> str:
  route = normalise_route(route)
  return urljoin(base_url.rstrip("/") + "/", route.lstrip("/"))


def should_exclude(route: str, excludes: Iterable[str] | None = None) -> bool:
  check = normalise_route(route)
  prefixes = DEFAULT_EXCLUDES if excludes is None else [str(prefix).strip() for prefix in excludes]
  for prefix in prefixes:
    prefix = prefix.rstrip("/")
    if not prefix:
      continue
    if check == prefix or check.startswith(f"{prefix}/"):
      return True
  return False



def repo_html_routes(repo_root: Path = REPO_ROOT, excludes: Iterable[str] | None = None) -> list[str]:
  routes: list[str] = []
  for file_path in repo_root.rglob("index.html"):
    relative = file_path.relative_to(repo_root)
    if relative.parts[:2] == ("assets", "partials"):
      continue
    route = "/" + str(relative.parent).replace(os.sep, "/")
    if route == "/.":
      route = "/"
    if route.endswith("/index"):
      route = route[: -len("/index")]
    if should_exclude(route, excludes):
      continue
    routes.append(route or "/")

  root_404 = repo_root / "404.html"
  if root_404.exists():
    routes.append("/404")

  return sorted(set(route.rstrip("/") or "/" for route in routes))


def fetch_html(url: str, timeout: float = 20.0, extra_headers: dict[str, str] | None = None, allow_redirects: bool = True) -> dict[str, Any]:
  headers = {
    "User-Agent": "AI-management-suite/audits (+https://jonathan-harris.online)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
  }
  if extra_headers:
    headers.update({str(k): str(v) for k, v in extra_headers.items() if str(v).strip()})
  try:
    response = requests.get(url, timeout=timeout, headers=headers, allow_redirects=allow_redirects)
    text = response.text or ""
    return {
      "ok": response.ok,
      "status": response.status_code,
      "text": text,
      "url": response.url,
      "headers": dict(response.headers),
      "history": [{"status": item.status_code, "url": item.url} for item in response.history],
    }
  except Exception as exc:  # pragma: no cover - defensive runtime path
    return {"ok": False, "status": 0, "text": "", "url": url, "error": str(exc), "headers": {}, "history": []}


def parse_html(html: str) -> BeautifulSoup:
  return BeautifulSoup(html or "", "html.parser")


def detect_challenge_page(status: int, html: str) -> str | None:
  soup = parse_html(html)
  title = soup.title.get_text(" ", strip=True) if soup.title else ""
  body = soup.get_text(" ", strip=True)[:6000].lower()
  title_lower = title.lower()

  challenge_markers = (
    "just a moment",
    "attention required",
    "checking your browser",
    "enable javascript and cookies to continue",
    "cf-browser-verification",
    "challenge-platform",
    "cdn-cgi/challenge-platform",
    "ddos protection by cloudflare",
  )

  if any(marker in title_lower or marker in body for marker in challenge_markers):
    return f"challenge/interstitial detected (status={status}, title={title or '(missing title)'})"

  if status in {401, 403, 429, 503} and title_lower in {"just a moment...", "attention required! | cloudflare"}:
    return f"challenge/interstitial detected (status={status}, title={title})"

  return None


def extract_meta(soup: BeautifulSoup) -> dict[str, Any]:
  title = soup.title.get_text(" ", strip=True) if soup.title else ""
  canonical = ""
  canonical_tag = soup.select_one("link[rel='canonical']")
  if canonical_tag and canonical_tag.get("href"):
    canonical = canonical_tag["href"].strip()

  meta_description = ""
  description_tag = soup.select_one("meta[name='description']")
  if description_tag and description_tag.get("content"):
    meta_description = description_tag["content"].strip()

  viewport = ""
  viewport_tag = soup.select_one("meta[name='viewport']")
  if viewport_tag and viewport_tag.get("content"):
    viewport = viewport_tag["content"].strip()

  h1 = soup.select_one("h1")
  h1_text = h1.get_text(" ", strip=True) if h1 else ""
  og_tags = {
    tag.get("property") or tag.get("name"): tag.get("content", "").strip()
    for tag in soup.select("meta[property^='og:'], meta[name^='og:']")
    if tag.get("content")
  }
  schema_count = len(soup.select("script[type='application/ld+json']"))

  return {
    "title": title,
    "canonical": canonical,
    "metaDescription": meta_description,
    "viewport": viewport,
    "h1": h1_text,
    "og": og_tags,
    "schemaCount": schema_count,
  }


def html_report_shell(title: str, body: str) -> str:
  generated = utc_now()
  return f"""<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>{title}</title><style>
body{{font-family:Arial,Helvetica,sans-serif;margin:0;background:#f4f7fb;color:#111827;line-height:1.55;}}
main{{max-width:1280px;margin:0 auto;padding:40px 24px 72px;}}
header{{background:#0d1420;color:#fff;padding:28px 24px;}}
h1,h2,h3{{line-height:1.2;margin:0 0 12px;}}
section{{background:#fff;border:1px solid #e5e7eb;border-radius:18px;padding:24px;margin:20px 0;box-shadow:0 12px 30px rgba(13,20,32,.06);}}
table{{width:100%;border-collapse:collapse;font-size:14px;}}
th,td{{border-bottom:1px solid #e5e7eb;padding:10px 8px;text-align:left;vertical-align:top;}}
small,.muted{{color:#6b7280;}}
.badge{{display:inline-block;padding:4px 10px;border-radius:999px;font-size:12px;font-weight:700;background:#eef2ff;color:#4338ca;}}
.badge.fail{{background:#fee2e2;color:#991b1b;}}
.badge.pass{{background:#dcfce7;color:#166534;}}
.badge.warn{{background:#fef3c7;color:#92400e;}}
code{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;background:#f3f4f6;padding:2px 6px;border-radius:6px;}}
ul{{padding-left:20px;}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:16px;}}
.kpi{{background:#0d1420;color:#fff;border-radius:18px;padding:18px;}}
.toc a{{text-decoration:none;}}
.pill{{display:inline-block;margin:0 6px 6px 0;padding:6px 10px;border-radius:999px;background:#eef2ff;color:#4338ca;font-size:12px;font-weight:700;}}
.section-note{{margin-top:8px;font-size:13px;color:#4b5563;}}
.mono{{font-family:ui-monospace,SFMono-Regular,Menlo,monospace;}}
a{{color:#4338ca;}}
.tight td,.tight th{{font-size:13px;padding:8px 6px;}}
</style></head><body><header><h1>{title}</h1><div class="muted">Generated {generated}</div></header><main>{body}</main></body></html>"""


def build_r2_client() -> Any:
  return boto3.client(
    "s3",
    endpoint_url=os.environ["R2_ENDPOINT"],
    aws_access_key_id=os.environ["R2_ACCESS_KEY_ID"],
    aws_secret_access_key=os.environ["R2_SECRET_ACCESS_KEY"],
    region_name=os.environ.get("R2_REGION", "auto"),
  )


def resolve_r2_public_base_for_bucket(bucket: str, explicit_public_base: str | None = None) -> str:
  if explicit_public_base and explicit_public_base.strip():
    return explicit_public_base.rstrip("/")

  bucket = str(bucket or "").strip()
  audits_bucket = os.environ.get("R2_BUCKET_AUDITS", "").strip()
  brand_assets_bucket = os.environ.get("R2_BUCKET_BRAND_ASSETS", "").strip()

  if audits_bucket and bucket == audits_bucket:
    public_base = os.environ.get("R2_PUBLIC_BASE_URL_AUDITS", "").strip()
    if public_base:
      return public_base.rstrip("/")

  if brand_assets_bucket and bucket == brand_assets_bucket:
    public_base = os.environ.get("R2_PUBLIC_BASE_URL_BRAND_ASSETS", "").strip()
    if public_base:
      return public_base.rstrip("/")

  public_base = os.environ.get("R2_PUBLIC_BASE_URL_AUDITS", "").strip() or os.environ.get("R2_PUBLIC_BASE_URL_BRAND_ASSETS", "").strip()
  if not public_base:
    raise RuntimeError("No public R2 base URL is configured for the selected audit bucket")
  return public_base.rstrip("/")


def upload_file_to_r2(client: Any, bucket: str, prefix: str, file_path: Path, public_base_url: str | None = None) -> str:
  key = f"{prefix.rstrip('/')}/{file_path.name}"
  content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
  client.upload_file(
    str(file_path),
    bucket,
    key,
    ExtraArgs={"ContentType": content_type},
  )
  public_base = resolve_r2_public_base_for_bucket(bucket, public_base_url)
  return f"{public_base}/{key}"


def upload_directory_to_r2(client: Any, bucket: str, prefix: str, directory: Path, public_base_url: str | None = None) -> dict[str, str]:
  uploaded: dict[str, str] = {}
  public_base = resolve_r2_public_base_for_bucket(bucket, public_base_url)
  for file_path in sorted(directory.rglob("*")):
    if file_path.is_dir():
      continue
    relative = file_path.relative_to(directory).as_posix()
    key = f"{prefix.rstrip('/')}/{relative}"
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    client.upload_file(
      str(file_path),
      bucket,
      key,
      ExtraArgs={"ContentType": content_type},
    )
    uploaded[relative] = f"{public_base}/{key}"
  return uploaded


def upload_selected_files_to_r2(client: Any, bucket: str, prefix: str, files: dict[str, Path], public_base_url: str | None = None) -> dict[str, str]:
  uploaded: dict[str, str] = {}
  public_base = resolve_r2_public_base_for_bucket(bucket, public_base_url)
  for relative_name, file_path in sorted(files.items()):
    key = f"{prefix.rstrip('/')}/{relative_name}"
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    client.upload_file(str(file_path), bucket, key, ExtraArgs={"ContentType": content_type})
    uploaded[relative_name] = f"{public_base}/{key}"
  return uploaded


def validate_public_json_artifacts(
  uploaded: dict[str, str],
  names: Iterable[str],
  *,
  attempts: int = 8,
  delay_seconds: float = 1.5,
  timeout_seconds: float = 20,
) -> dict[str, Any]:
  """Verify that required uploaded JSON artefacts are publicly readable.

  A completed callback is the hand-off boundary to AIMS.  Do not cross that
  boundary until every required machine-readable object can be fetched and
  parsed from the same public URL advertised in the callback.
  """
  required = [str(name).strip() for name in names if str(name).strip()]
  if not required:
    raise ValueError("At least one public JSON artefact name is required")

  results: dict[str, Any] = {}
  for name in required:
    url = str(uploaded.get(name) or "").strip()
    if not url:
      raise RuntimeError(f"public JSON validation missing URL for {name}")

    last_error: Exception | None = None
    for attempt in range(1, max(1, int(attempts)) + 1):
      try:
        response = requests.get(url, headers={"Accept": "application/json"}, timeout=timeout_seconds)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict):
          raise RuntimeError(f"{name} returned {type(payload).__name__}; expected a JSON object")
        results[name] = {
          "url": url,
          "status": "PASS",
          "attempt": attempt,
          "contentLength": len(response.content),
        }
        break
      except Exception as exc:  # pragma: no cover - depends on the live R2 endpoint
        last_error = exc
        if attempt < max(1, int(attempts)):
          time.sleep(max(0.0, float(delay_seconds)))
    else:
      raise RuntimeError(f"public JSON artefact validation failed for {name}: {last_error}")

  return {"status": "PASS", "checked": results, "generatedAt": utc_now()}


def post_callback(callback_url: str | None, callback_token: str | None, payload: dict[str, Any]) -> None:
  if not callback_url:
    return
  headers = {"Content-Type": "application/json"}
  if callback_token:
    headers["Authorization"] = f"Bearer {callback_token}"
    headers["X-Audit-Callback-Token"] = callback_token
  response = requests.post(callback_url, headers=headers, data=json.dumps(payload), timeout=20)
  response.raise_for_status()
