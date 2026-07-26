#!/usr/bin/env python3
"""Generate an ephemeral manuscript manifest from the governed ebook workbook.

The workbook remains the source of truth. The generated manifest is build-time only
and must never be published because it contains private manuscript source URLs.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "scripts" / "data" / "manuscripts.json"
REQUIRED_COLUMNS = ("Slug", "Title", "Manuscript filename", "Manuscript URL")
DRIVE_ID_RE = re.compile(r"/file/d/([^/]+)")


def drive_file_id(url: str) -> str:
    match = DRIVE_ID_RE.search(url)
    if match:
        return match.group(1)
    parsed = urlparse(url)
    if parsed.netloc.endswith("google.com"):
        from urllib.parse import parse_qs

        value = parse_qs(parsed.query).get("id", [""])[0]
        if value:
            return value
    raise ValueError(f"Unsupported Google Drive manuscript URL: {url}")


def build_manifest(workbook_path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Ebooks Master" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'Ebooks Master' sheet")
    sheet = workbook["Ebooks Master"]
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(next(sheet.iter_rows()), start=1) if cell.value}
    missing = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing:
        raise ValueError("Ebooks Master is missing required manuscript fields: " + ", ".join(missing))

    books: list[dict[str, str]] = []
    seen_slugs: set[str] = set()
    seen_ids: set[str] = set()
    errors: list[str] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        slug = str(row[headers["Slug"] - 1] or "").strip()
        title = str(row[headers["Title"] - 1] or "").strip()
        if not slug and not title:
            continue
        filename = str(row[headers["Manuscript filename"] - 1] or "").strip()
        source_url = str(row[headers["Manuscript URL"] - 1] or "").strip()
        if not slug or not title or not filename or not source_url:
            errors.append(f"row {row_number}: active ebook is missing slug/title/manuscript filename/manuscript URL")
            continue
        try:
            file_id = drive_file_id(source_url)
        except ValueError as exc:
            errors.append(f"row {row_number} ({slug}): {exc}")
            continue
        if slug in seen_slugs:
            errors.append(f"row {row_number}: duplicate ebook slug {slug}")
        if file_id in seen_ids:
            errors.append(f"row {row_number}: duplicate manuscript file id {file_id}")
        seen_slugs.add(slug)
        seen_ids.add(file_id)
        books.append(
            {
                "slug": slug,
                "title": title,
                "filename": filename,
                "source_url": source_url,
                "file_id": file_id,
                "download_url": f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t",
            }
        )

    if errors:
        raise ValueError("Manuscript manifest validation failed:\n- " + "\n- ".join(errors))
    if not books:
        raise ValueError("No active ebook manuscript records were found")

    return {
        "schema_version": 1,
        "source": workbook_path.name,
        "book_count": len(books),
        "books": books,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate the private build-time manuscript manifest")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    manifest = build_manifest(workbook_path)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Generated {args.output.relative_to(ROOT)} for {manifest['book_count']} manuscripts.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
