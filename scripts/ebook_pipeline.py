from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import re
import shutil
import hashlib
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
from urllib.parse import urlencode, urlparse

import openpyxl

from scripts.ebook_content_helpers import (
    audience_faq_answer,
    build_same_source_srcset,
    cover_sizes,
    default_short as helper_default_short,
    normalise_audience_copy,
    normalise_topic_copy,
    topic_intro as helper_topic_intro,
)

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
EBOOKS_DIR = ROOT / "ebooks"
CATALOGUE_DIR = ROOT / "catalogue"
TOPICS_DIR = ROOT / "topics"
MASTER_PATH = DATA_DIR / "ebooks-master.json"
WORKBOOK_NORMALISATIONS_PATH = ROOT / "config" / "workbook-normalisations.json"
CRAWLER_CHECKSUMS_PATH = ROOT / "config" / "crawler-checksums.json"
CRAWLER_SNAPSHOTS_DIR = ROOT / "config" / "crawler-snapshots"
DYNAMIC_ROUTE_MANIFEST_PATH = DATA_DIR / "dynamic-route-manifest.json"
SEARCH_VISIBILITY_SURFACES_PATH = DATA_DIR / "search-visibility-surfaces.json"
AMAZON_BOOK_SIGNALS_PATH = DATA_DIR / "amazon-book-signals.json"
BOOK_SAMPLE_CHAPTERS_PATH = DATA_DIR / "book-sample-chapters.json"
SITE_FACTS_PATH = DATA_DIR / "site-facts.json"
RELATED_BOOK_CURATION_PATH = DATA_DIR / "related-book-curation.json"
HEADER_PARTIAL = ROOT / "assets" / "partials" / "header.html"
FOOTER_PARTIAL = ROOT / "assets" / "partials" / "footer.html"
EBOOK_TEMPLATE_CSS = ROOT / "assets" / "css" / "ebook-template.css"
SITE_NAME = "Jonathan Harris"
SITE_URL = "https://jonathan-harris.online"
DEFAULT_AUDIENCE = "Readers who want practical, plain-English AI insight without the buzzwords."
DEFAULT_TONE = "Plain-English, practical, sceptical, no-hype"
BOOK_COVER_WIDTH = 2480
BOOK_COVER_HEIGHT = 3508
VALIDATION_REPORT = ROOT / "VALIDATION_OUTPUT.txt"
SHARED_INTER_FONT_HEAD_BLOCK = """<link href="https://fonts.googleapis.com" rel="preconnect"/>
<link crossorigin="" href="https://fonts.gstatic.com" rel="preconnect"/>
<link href="https://fonts.googleapis.com/css2?family=Inter:ital,wght@0,400;0,600;0,700;0,800&display=swap" rel="stylesheet"/>"""

STOPWORDS = {
    "a", "an", "and", "are", "as", "at", "be", "book", "by", "for", "from", "how", "in", "into",
    "is", "it", "its", "of", "on", "or", "our", "that", "the", "this", "through", "to", "what",
    "when", "where", "why", "with", "your", "you", "ai", "artificial", "intelligence",
}


LEGACY_DETAIL_REDIRECT_LINES = [
    "/ebooks/*/detail  /ebooks/:splat/  301",
    "/ebooks/*/detail.html  /ebooks/:splat/  301",
    "/ebooks/*/details.html  /ebooks/:splat/  301",
]

LOCALE_ALIAS_REDIRECT_LINES = [
    f"/en-gb/*  {SITE_URL}/:splat  301",
    f"/en-us/*  {SITE_URL}/:splat  301",
    f"/en-ca/*  {SITE_URL}/:splat  301",
    f"/en-au/*  {SITE_URL}/:splat  301",
]

MALFORMED_SLUG_FIXES = [
    {
        "source": "/ebooks/artificial-intelligence-in-pharmaceuticalsrevolutionizing-healthcare",
        "target": "/ebooks/artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare/",
    },
    {
        "source": "/ebooks/artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement-k",
        "target": "/ebooks/artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement/",
    },
    {
        "source": "/ebooks/lights-camera-algorithm-ais-role-in-modern-filmmaking",
        "target": "/ebooks/lights-camera-algorithm-ai-s-role-in-modern-filmmaking/",
    },
    {
        "source": "/ebooks/the-architects-of-ai_-pioneers_-breakthroughs_-and-the-road-ahead",
        "target": "/ebooks/the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead/",
    },
]

EXTERNAL_CRAWLER_FILES = {
    "robots": f"{SITE_URL}/robots.txt",
    "sitemap": f"{SITE_URL}/sitemap.xml",
    "llms": f"{SITE_URL}/llms.txt",
}

CRAWLER_SNAPSHOT_FILENAMES = {
    "robots": "robots.txt",
    "sitemap": "sitemap.xml",
    "llms": "llms.txt",
}

TEMPLATE_REQUIRED_FRAGMENTS = [
    '<body class="ebook-detail"',
    '<nav aria-label="Breadcrumb" class="breadcrumbs">',
    '<section class="card ebook-section quick-facts">',
    '<section class="ebook-showcase">',
    '<section class="card ebook-section" id="deeper-overview">',
    '<section class="related-books card">',
    '<section class="faq card" aria-label="Frequently asked questions">',
    '<section class="jh-journey-panel">',
]

TEMPLATE_ORDERED_MARKERS = [
    '<section class="card ebook-section quick-facts">',
    '<section class="ebook-showcase">',
    '<h2>Who this book is for</h2>',
    '<h2>Key themes</h2>',
    '<h2>What you’ll learn</h2>',
    '<section class="card ebook-section" id="deeper-overview">',
    '<section class="related-books card">',
    '<section class="faq card" aria-label="Frequently asked questions">',
]

BROKEN_PHRASE_FRAGMENTS = [
    "AI in how ai shapes attention and thinking",
    "how AI changes the day-to-day reality of how ai shapes attention and thinking",
    "how AI changes how ai shapes attention and thinking in practice",
    "where AI fits inside how ai shapes attention and thinking",
    "AI is not arriving in how ai shapes attention and thinking",
]

TITLE_SUBSTITUTION_FIELD_NAMES = (
    "short",
    "description",
    "summary",
    "audience",
    "who_for",
    "what_this_book_covers",
    "why_it_matters",
)

SLUG_SUBJECT_PREFIXES = (
    "ai-revolution-in-",
    "ai-in-",
    "artificial-intelligence-revolution-in-",
    "artificial-intelligence-in-",
    "artificial-intelligence-for-",
    "artificial-intelligence-powered-",
    "digital-diagnosis-how-ai-is-revolutionizing-",
    "digital-defense-the-role-of-ai-in-",
    "from-reporters-to-robots-how-ai-is-reshaping-",
    "lights-camera-algorithm-ai-s-role-in-",
    "smart-buildings-ai-powered-",
    "beyond-earth-how-ai-is-transforming-",
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-",
    "the-future-of-",
    "the-ai-behind-your-feed-",
    "the-ai-music-revolution-",
    "the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-",
)

SLUG_SUBJECT_STOP_WORDS = {
    "revolutionizing",
    "transforming",
    "modernizing",
    "redefining",
    "reimagining",
    "leveraging",
    "harnessing",
    "optimizing",
    "building",
    "navigating",
    "reshaping",
    "enhance",
    "enhancing",
    "through",
    "with",
    "for",
    "a",
    "an",
    "the",
    "future",
    "how",
    "what",
}

SLUG_SUBJECT_OVERRIDES = {
    "smart-buildings-ai-powered-efficiency-and-sustainability": "smart buildings",
    "digital-diagnosis-how-ai-is-revolutionizing-healthcare": "healthcare",
    "digital-defense-the-role-of-ai-in-modern-warfare": "modern warfare",
    "from-reporters-to-robots-how-ai-is-reshaping-journalism": "journalism",
    "lights-camera-algorithm-ai-s-role-in-modern-filmmaking": "modern filmmaking",
    "beyond-earth-how-ai-is-transforming-space-exploration": "space exploration",
    "the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media": "social media",
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information": "government",
    "the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming": "gaming",
}


def infer_subject_phrase(slug: str, title: str, topic: str) -> str:
    slug_clean = slugify(slug)
    if slug_clean in SLUG_SUBJECT_OVERRIDES:
        return SLUG_SUBJECT_OVERRIDES[slug_clean]
    for prefix in SLUG_SUBJECT_PREFIXES:
        if slug_clean.startswith(prefix):
            remainder = slug_clean[len(prefix):]
            words = []
            for part in remainder.split("-"):
                if words and part in SLUG_SUBJECT_STOP_WORDS:
                    break
                words.append(part)
            candidate = clean_paragraph(" ".join(words))
            if candidate:
                return candidate

    title_core = clean_paragraph(title).split(":", 1)[0].strip()
    patterns = [
        r"^AI Revolution in (.+)$",
        r"^AI in (.+)$",
        r"^Artificial Intelligence Revolution in (.+)$",
        r"^Artificial Intelligence in (.+)$",
        r"^Artificial Intelligence for (.+)$",
        r"^Artificial Intelligence Powered (.+)$",
        r"^The Future of (.+)$",
    ]
    for title_pattern in patterns:
        title_match = re.match(title_pattern, title_core, flags=re.I)
        if title_match:
            candidate = clean_paragraph(title_match.group(1))
            if candidate:
                return candidate.lower()

    topic_candidate = clean_paragraph(topic)
    if topic_candidate:
        return topic_candidate.lower()
    return title_core.lower()



def normalise_title_substitution_text(value: str, *, slug: str, title: str, topic: str) -> str:
    cleaned = clean_paragraph(value)
    if not cleaned:
        return cleaned

    title_core = clean_paragraph(title).split(":", 1)[0].strip()
    if title_core:
        subject = infer_subject_phrase(slug, title, topic)
        replacements = [
            (rf"\bAI in {re.escape(title_core)}\b", f"AI in {subject}"),
            (rf"\bA practical overview of AI in {re.escape(title_core)}\b", f"A practical overview of AI in {subject}"),
            (rf"\bhow AI changes the day-to-day reality of {re.escape(title_core)}\b", f"how AI changes the day-to-day reality of {subject}"),
            (rf"\bhow AI changes {re.escape(title_core)} in practice\b", f"how AI changes {subject} in practice"),
            (rf"\bwhere AI fits inside {re.escape(title_core)}\b", f"where AI fits inside {subject}"),
            (rf"\bAI is not arriving in {re.escape(title_core)}\b", f"AI is not arriving in {subject}"),
        ]
        for phrase_pattern, replacement_text in replacements:
            cleaned = re.sub(phrase_pattern, replacement_text, cleaned, flags=re.I)
    return normalise_topic_copy(cleaned, topic)



def sanitise_record_copy(record: Dict[str, Any]) -> Dict[str, Any]:
    slug = clean_paragraph(record.get("slug", ""))
    title = clean_paragraph(record.get("title", ""))
    topic = clean_paragraph(record.get("topic", ""))
    for field in TITLE_SUBSTITUTION_FIELD_NAMES:
        if field in record:
            record[field] = replace_banned_ebook_phrases(
                normalise_title_substitution_text(record.get(field, ""), slug=slug, title=title, topic=topic),
                slug,
            )
    if "what_youll_learn" in record and isinstance(record.get("what_youll_learn"), list):
        record["what_youll_learn"] = [
            replace_banned_ebook_phrases(
                normalise_title_substitution_text(item, slug=slug, title=title, topic=topic),
                slug,
            )
            for item in record.get("what_youll_learn", [])
            if clean_paragraph(item)
        ]
    if "short" in record:
        record["short_description"] = record.get("short", "")
    if "audience" in record:
        record["audience"] = normalise_audience_copy(record.get("audience", ""), topic) or record.get("audience", "")
    if "who_for" in record:
        record["who_for"] = replace_banned_ebook_phrases(
            normalise_title_substitution_text(record.get("who_for", ""), slug=slug, title=title, topic=topic),
            slug,
        )
    faq_items = record.get("faq")
    if isinstance(faq_items, list):
        cleaned_faq = []
        for item in faq_items:
            if not isinstance(item, dict):
                continue
            question = clean_paragraph(item.get("name", ""))
            answer = clean_paragraph((item.get("acceptedAnswer") or {}).get("text", ""))
            if question.lower() == "who is this book for?":
                answer = audience_faq_answer(record.get("audience", ""), topic)
            else:
                answer = replace_banned_ebook_phrases(
                    normalise_title_substitution_text(answer, slug=slug, title=title, topic=topic),
                    slug,
                )
            cleaned_faq.append({
                "@type": item.get("@type") or "Question",
                "name": question,
                "acceptedAnswer": {
                    "@type": (item.get("acceptedAnswer") or {}).get("@type") or "Answer",
                    "text": answer,
                },
            })
        record["faq"] = cleaned_faq
    return record

def read_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))



def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")



def load_workbook_normalisations() -> Dict[str, Dict[str, Dict[str, str]]]:
    payload = read_json(WORKBOOK_NORMALISATIONS_PATH, default={}) or {}
    if not isinstance(payload, dict):
        return {}
    approved: Dict[str, Dict[str, Dict[str, str]]] = {}
    for slug, fields in payload.items():
        if not isinstance(fields, dict):
            continue
        field_map: Dict[str, Dict[str, str]] = {}
        for field_name, entry in fields.items():
            if not isinstance(entry, dict):
                continue
            field_map[clean_paragraph(field_name)] = {
                "raw": clean_paragraph(entry.get("raw", "")),
                "approved": clean_paragraph(entry.get("approved", "")),
            }
        approved[clean_paragraph(slug)] = field_map
    return approved



def utc_now() -> str:
    return dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def infer_build_timestamp() -> str:
    if MASTER_PATH.exists():
        return dt.datetime.fromtimestamp(MASTER_PATH.stat().st_mtime, tz=dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return utc_now()


def parse_timestamp(value: Any) -> dt.datetime | None:
    cleaned = clean_paragraph(value)
    if not cleaned:
        return None
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", cleaned):
        return dt.datetime.fromisoformat(cleaned).replace(tzinfo=dt.timezone.utc)
    try:
        parsed = dt.datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
    except ValueError:
        match = re.search(r"(\d{4}-\d{2}-\d{2})", cleaned)
        if not match:
            return None
        return dt.datetime.fromisoformat(match.group(1)).replace(tzinfo=dt.timezone.utc)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed.astimezone(dt.timezone.utc).replace(microsecond=0)


def governed_generated_utc(books: List[Dict[str, Any]]) -> str:
    timestamps = [
        parse_timestamp(book.get("dateModified") or book.get("datePublished"))
        for book in books
    ]
    timestamps = [timestamp for timestamp in timestamps if timestamp is not None]
    if timestamps:
        return max(timestamps).isoformat().replace("+00:00", "Z")
    return infer_build_timestamp()


def normalise_lastmod(value: Any) -> str:
    cleaned = clean_paragraph(value)
    if not cleaned:
        return dt.date.today().isoformat()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", cleaned):
        return cleaned
    try:
        parsed = dt.datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
    except ValueError:
        match = re.search(r"(\d{4}-\d{2}-\d{2})", cleaned)
        return match.group(1) if match else dt.date.today().isoformat()
    return parsed.date().isoformat()


def file_lastmod(path: Path) -> str:
    return dt.datetime.fromtimestamp(path.stat().st_mtime, tz=dt.timezone.utc).date().isoformat()


def html_significant_lastmod(path: Path) -> str:
    """Return an explicit significant-content date, never a build/copy timestamp.

    Sitemap lastmod is omitted when a page does not carry a governed review or
    modification date. This avoids telling crawlers that every page changed
    merely because Cloudflare rebuilt the site.
    """
    try:
        source = path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""
    patterns = (
        r'<meta[^>]+name=["\']dateModified["\'][^>]+content=["\'](\d{4}-\d{2}-\d{2})',
        r'<meta[^>]+content=["\'](\d{4}-\d{2}-\d{2})["\'][^>]+name=["\']dateModified["\']',
        r'["\']dateModified["\']\s*:\s*["\'](\d{4}-\d{2}-\d{2})',
        r'Last reviewed\s+(\d{4}-\d{2}-\d{2})',
    )
    for pattern in patterns:
        match = re.search(pattern, source, flags=re.I)
        if match:
            return match.group(1)
    return ""


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()



def slugify(value: str) -> str:
    value = (value or "").strip().lower()
    value = value.replace("&", "and")
    value = re.sub(r"[’'`]", "", value)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")



def normalise_space(value: str) -> str:
    value = str(value or "").replace("\xa0", " ")
    return re.sub(r"\s+", " ", value).strip()



def clean_paragraph(value: str) -> str:
    value = normalise_space(value)
    value = value.replace(" .", ".")
    value = re.sub(r"\s+([,.;:!?])", r"\1", value)
    return value



def ensure_trailing_slash(url: str) -> str:
    url = clean_paragraph(url)
    if not url:
        return url
    return url if url.endswith("/") else url + "/"



def strip_pages_from_summary(text: str, pages: int | None = None) -> str:
    cleaned = clean_paragraph(text)
    cleaned = re.sub(r"\.?\s*Pages:\s*\d+\.?\s*$", ".", cleaned, flags=re.I)
    cleaned = re.sub(r"\s+\.$", ".", cleaned)
    if pages:
        cleaned = re.sub(rf"\b{pages}\s*-?\s*page guide\.?$", "", cleaned, flags=re.I).strip()
    return clean_paragraph(cleaned)


def canonical_text_for_role_check(value: Any, *, pages: int | None = None) -> str:
    text = clean_paragraph(value)
    if pages:
        text = strip_pages_from_summary(text, pages)
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def content_role_validation_errors(records: List[Dict[str, Any]]) -> List[str]:
    errors: List[str] = []
    audience_counter = Counter()
    who_for_counter = Counter()

    for record in records:
        pages = record.get("pages")
        roles = {
            "short": canonical_text_for_role_check(record.get("short"), pages=pages),
            "description": canonical_text_for_role_check(record.get("description"), pages=pages),
            "summary": canonical_text_for_role_check(record.get("summary"), pages=pages),
            "what_this_book_covers": canonical_text_for_role_check(record.get("what_this_book_covers"), pages=pages),
            "audience": canonical_text_for_role_check(record.get("audience")),
            "who_for": canonical_text_for_role_check(record.get("who_for")),
            "why_it_matters": canonical_text_for_role_check(record.get("why_it_matters")),
        }

        identical_pairs = [
            ("short", "description"),
            ("description", "summary"),
            ("summary", "what_this_book_covers"),
            ("audience", "who_for"),
            ("summary", "why_it_matters"),
        ]
        for left, right in identical_pairs:
            if roles[left] and roles[left] == roles[right]:
                errors.append(f"{record['slug']} collapses the distinct workbook roles for {left} and {right}.")

        if roles["audience"]:
            audience_counter[roles["audience"]] += 1
        if roles["who_for"]:
            who_for_counter[roles["who_for"]] += 1

    repeated_audience = [count for count in audience_counter.values() if count > max(6, len(records) // 4)]
    repeated_who_for = [count for count in who_for_counter.values() if count > max(6, len(records) // 4)]
    if repeated_audience:
        errors.append("Workbook audience field is still over-reused across the catalogue; vary it more before publishing.")
    if repeated_who_for:
        errors.append("Workbook 'Who this book is for' field is still over-reused across the catalogue; vary it more before publishing.")
    return errors



def split_field(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [clean_paragraph(v) for v in value if clean_paragraph(v)]
    text = str(value).replace("•", "\n")
    parts = re.split(r"\n|\||;", text)
    return [clean_paragraph(part) for part in parts if clean_paragraph(part)]



def unique_list(values: Iterable[str]) -> List[str]:
    seen = set()
    output: List[str] = []
    for value in values:
        cleaned = clean_paragraph(value)
        key = cleaned.lower()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        output.append(cleaned)
    return output



def choose_authoritative_text(workbook_content: Dict[str, Any], content_source: Dict[str, Any], field: str, *, allow_fallback: bool = False) -> Any:
    workbook_value = workbook_content.get(field)
    if isinstance(workbook_value, list):
        if any(clean_paragraph(item) for item in workbook_value):
            return workbook_value
    elif clean_paragraph(workbook_value):
        return workbook_value
    return content_source.get(field) if allow_fallback else None


def book_about_terms(book: Dict[str, Any]) -> List[str]:
    return unique_list([book.get("topic", ""), *(book.get("tags", [])[:5] or [])])


def template_contract_errors(page_text: str, book: Dict[str, Any]) -> List[str]:
    errors: List[str] = []
    for fragment in TEMPLATE_REQUIRED_FRAGMENTS:
        if fragment not in page_text:
            errors.append(f"{book['slug']} is missing a governed template fragment: {fragment}")

    positions: List[int] = []
    for marker in TEMPLATE_ORDERED_MARKERS:
        pos = page_text.find(marker)
        if pos == -1:
            continue
        positions.append(pos)
    if positions != sorted(positions):
        errors.append(f"{book['slug']} breaks the governed section order for canonical ebook pages.")

    required_ctas = [
        f'href="{book["buy_route"]}"',
        'href="/ebooks/"',
        'href="#deeper-overview"',
    ]
    for cta in required_ctas:
        if cta not in page_text:
            errors.append(f"{book['slug']} is missing a governed CTA target: {cta}")
    return errors


def text_tokens(value: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]+", clean_paragraph(value).lower())
        if len(token) > 2 and token not in STOPWORDS
    }


def stem_token(token: str) -> str:
    token = clean_paragraph(token).lower()
    for suffix in (
        "ization", "isation", "ational", "ability", "ibility", "ation", "ition",
        "fulness", "ousness", "iveness", "ingly", "ments", "ement", "ment",
        "able", "ible", "ness", "ship", "ance", "ence", "ally", "fully",
        "less", "iest", "ies", "ied", "ing", "ers", "er", "ed", "ly",
        "ity", "s",
    ):
        if len(token) > len(suffix) + 2 and token.endswith(suffix):
            return token[: -len(suffix)]
    return token


def significant_tokens(value: str) -> set[str]:
    return {stem_token(token) for token in text_tokens(value)}


def book_token_groups(book: Dict[str, Any]) -> Dict[str, set[str]]:
    return {
        "topic": significant_tokens(book.get("topic", "")),
        "title": significant_tokens(book.get("title", "")),
        "tags": {token for item in book.get("tags", []) for token in significant_tokens(item)},
        "keywords": {token for item in book.get("keywords", []) for token in significant_tokens(item)},
    }


def related_book_score(candidate: Dict[str, Any], current: Dict[str, Any], current_tokens: Dict[str, set[str]], candidate_tokens: Dict[str, set[str]]) -> tuple[int, tuple[int, int, int, int, str]]:
    same_topic = int(candidate.get("topic") == current.get("topic"))
    title_overlap = len(current_tokens["title"] & candidate_tokens["title"])
    topic_overlap = len(current_tokens["topic"] & candidate_tokens["topic"])
    tag_overlap = len(current_tokens["tags"] & candidate_tokens["tags"])
    keyword_overlap = len(current_tokens["keywords"] & candidate_tokens["keywords"])
    score = (same_topic * 100) + (title_overlap * 35) + (topic_overlap * 30) + (tag_overlap * 18) + (keyword_overlap * 8)
    tie_breaker = (same_topic, title_overlap, topic_overlap + tag_overlap, keyword_overlap, candidate["title"].lower())
    return score, tie_breaker


def broken_phrase_errors(books: List[Dict[str, Any]]) -> List[str]:
    errors: List[str] = []
    for book in books:
        title_core = clean_paragraph(book.get("title", "")).split(":", 1)[0].strip().lower()
        if not title_core:
            continue
        suspicious_fragments: List[str] = []
        if "ai" in title_core or "artificial intelligence" in title_core:
            suspicious_fragments = [
                f"ai in {title_core}",
                f"day-to-day reality of {title_core}",
                f"{title_core} in practice",
                f"fits inside {title_core}",
                f"arriving in {title_core}",
            ]
        for field in TITLE_SUBSTITUTION_FIELD_NAMES:
            value = clean_paragraph(book.get(field, ""))
            lower = value.lower()
            for fragment in BROKEN_PHRASE_FRAGMENTS:
                if fragment.lower() in lower:
                    errors.append(f"Broken phrase lint failed for {book['slug']} field {field}: {fragment}")
            if " in ai in " in f" {lower} ":
                errors.append(f"Broken phrase lint failed for {book['slug']} field {field}: in ai in")
            for fragment in suspicious_fragments:
                if fragment in lower:
                    errors.append(f"Broken title-substitution phrase detected for {book['slug']} field {field}: {fragment}")
    return errors


def title_contract_errors(relative_path: str, repo_title: str, workbook_title: str) -> List[str]:
    errors: List[str] = []
    if repo_title == workbook_title:
        return errors
    # Public catalogue totals are generated from the ebook master, so a stale
    # workbook Pages-sheet number must not overrule the governed BOOK_COUNT.
    def normalise_count(value: str) -> str:
        return re.sub(r"\b\d+(?=\s+Plain-English AI Books\b)", "{BOOK_COUNT}", value, flags=re.I)
    if relative_path in {"bio/index.html", "index.html"} and normalise_count(repo_title) == normalise_count(workbook_title):
        return errors
    errors.append(f"Workbook Pages title mismatch for {relative_path}: workbook '{workbook_title}' != repo '{repo_title}'.")
    return errors


def workbook_title_parity_audit(workbook_path: Path) -> Tuple[List[str], Dict[str, int]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Pages" not in wb.sheetnames:
        return ["Workbook is missing the Pages sheet."], {"checked": 0, "passed": 0, "mismatched": 0}

    ws = wb["Pages"]
    header_row = 5
    headers = {clean_paragraph(ws.cell(header_row, col).value).lower(): col for col in range(1, ws.max_column + 1) if clean_paragraph(ws.cell(header_row, col).value)}
    required = {"relative file path", "page title"}
    if not required.issubset(headers):
        return [], {"checked": 0, "passed": 0, "mismatched": 0}

    errors: List[str] = []
    checked = 0
    passed = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        relative_path = clean_paragraph(ws.cell(row_idx, headers["relative file path"]).value)
        workbook_title = clean_paragraph(ws.cell(row_idx, headers["page title"]).value)
        if not relative_path or not workbook_title or relative_path.startswith("/book/"):
            continue
        file_path = ROOT / relative_path
        if not file_path.exists() or file_path.suffix.lower() != ".html":
            continue
        checked += 1
        html_text = file_path.read_text(encoding="utf-8", errors="ignore")
        title_match = re.search(r"<title>(.*?)</title>", html_text, re.I | re.S)
        if not title_match:
            errors.append(f"Workbook Pages title mismatch for {relative_path}: workbook '{workbook_title}' != repo '<missing title tag>'.")
            continue
        repo_title = clean_paragraph(html.unescape(re.sub(r"\s+", " ", title_match.group(1))))
        route_errors = title_contract_errors(relative_path, repo_title, workbook_title)
        if route_errors:
            errors.extend(route_errors)
            continue
        passed += 1
    return errors, {"checked": checked, "passed": passed, "mismatched": len(errors)}

def workbook_static_route_contract_errors(workbook_path: Path) -> List[str]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Pages" not in wb.sheetnames:
        return ["Workbook is missing the Pages sheet."]

    ws = wb["Pages"]
    header_row = 5
    headers = {clean_paragraph(ws.cell(header_row, col).value).lower(): col for col in range(1, ws.max_column + 1) if clean_paragraph(ws.cell(header_row, col).value)}
    required = {"relative file path", "public url path", "page title"}
    if not required.issubset(headers):
        return []

    errors, _ = workbook_title_parity_audit(workbook_path)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        relative_path = clean_paragraph(ws.cell(row_idx, headers["relative file path"]).value)
        public_path = clean_paragraph(ws.cell(row_idx, headers["public url path"]).value)
        if not relative_path or relative_path.startswith("/book/"):
            continue
        file_path = ROOT / relative_path
        if not file_path.exists() or file_path.suffix.lower() != ".html":
            continue
        html_text = file_path.read_text(encoding="utf-8", errors="ignore")
        canonical_match = re.search(r'<link[^>]+rel="canonical"[^>]+href="([^"]+)"', html_text, re.I)
        if canonical_match and public_path:
            canonical_path = url_to_path(canonical_match.group(1))
            if canonical_path != public_path:
                errors.append(f"Workbook Pages canonical mismatch for {relative_path}: workbook path '{public_path}' != canonical '{canonical_path}'.")
    return errors


def related_book_contract_errors(books: List[Dict[str, Any]]) -> List[str]:
    errors: List[str] = []
    books_by_slug = {book["slug"]: book for book in books}
    for book in books:
        same_topic_slugs = [candidate["slug"] for candidate in books if candidate["slug"] != book["slug"] and candidate.get("topic") == book.get("topic")]
        governed_related = [slug for slug in book.get("related_slugs", []) if slug in books_by_slug]
        expected_same_topic = min(4, len(same_topic_slugs))
        if expected_same_topic and any(books_by_slug[slug].get("topic") != book.get("topic") for slug in governed_related[:expected_same_topic]):
            errors.append(f"Related book contract drift for {book['slug']}: same-topic titles are not ranked first.")
    return errors



def url_to_path(value: str) -> str:
    value = clean_paragraph(value)
    if not value:
        return ""
    if value.startswith("/"):
        return value
    if value.startswith("http://") or value.startswith("https://"):
        parsed = urlparse(value)
        path = parsed.path or "/"
        if parsed.query:
            path += "?" + parsed.query
        return path
    return value



def humanise_slug(slug: str) -> str:
    return clean_paragraph(slug.replace("-", " ").title())



def format_date(value: str) -> str:
    if not value:
        return ""
    try:
        return dt.date.fromisoformat(value).strftime("%d %B %Y")
    except ValueError:
        return value



def escape_paragraphs(text: str) -> str:
    paragraphs = [clean_paragraph(p) for p in re.split(r"\n{2,}", str(text or "")) if clean_paragraph(p)]
    return "".join(f"<p>{html.escape(p)}</p>" for p in paragraphs)



def json_script(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))



def render_header() -> str:
    return HEADER_PARTIAL.read_text(encoding="utf-8").strip()



def render_footer() -> str:
    return FOOTER_PARTIAL.read_text(encoding="utf-8").strip()



def render_tag_pills(tags: Iterable[str], class_name: str = "ebook-pill") -> str:
    return "".join(f'<span class="{class_name}">{html.escape(tag)}</span>' for tag in tags if clean_paragraph(tag))



def load_site_facts() -> Dict[str, Any]:
    try:
        payload = json.loads(SITE_FACTS_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def build_person_schema() -> Dict[str, Any]:
    facts = load_site_facts()
    podcast_name = clean_paragraph(facts.get("podcast_name", "")) or "Turing’s Torch: AI Weekly"
    return {
        "@context": "https://schema.org",
        "@type": "Person",
        "@id": f"{SITE_URL}/#person",
        "name": SITE_NAME,
        "url": f"{SITE_URL}/bio/",
        "jobTitle": ["AI Author", "Podcast Host"],
        "description": f"Jonathan Harris is an artificial intelligence author and host of {podcast_name}. He writes plain-English books explaining how AI works across industries including healthcare, finance, law, manufacturing, and education.",
        "knowsAbout": ["Artificial Intelligence", "Machine Learning", "Generative AI", "AI Ethics", "Applied AI", "LLMs"],
        "sameAs": [
            "https://youtube.com/@jonathanharris-r7i",
            "https://www.amazon.com/kindle-dbs/author?ref=dbs_G_A_C&asin=B0DNCHC337",
            "https://www.goodreads.com/author/show/54004095.Jonathan_Harris",
            "https://twitter.com/jonathan_harris_01",
            "https://www.instagram.com/jonathan.harris1975",
            "https://www.tiktok.com/@jonathan_harris_01",
            "https://www.facebook.com/share/1G7v4P69xa",
        ],
    }



def build_website_schema() -> Dict[str, Any]:
    return {
        "@context": "https://schema.org",
        "@type": "WebSite",
        "@id": f"{SITE_URL}/#website",
        "url": f"{SITE_URL}/",
        "name": "Jonathan Harris - AI Author & Podcast Host",
        "publisher": {"@id": f"{SITE_URL}/#person"},
        "about": {"@id": f"{SITE_URL}/#person"},
        "inLanguage": "en",
    }



_AMAZON_SIGNALS_CACHE: Dict[str, Any] | None = None


def load_amazon_book_signals() -> Dict[str, Any]:
    global _AMAZON_SIGNALS_CACHE
    if _AMAZON_SIGNALS_CACHE is not None:
        return _AMAZON_SIGNALS_CACHE
    try:
        payload = json.loads(AMAZON_BOOK_SIGNALS_PATH.read_text(encoding="utf-8"))
        books = payload.get("books", {}) if isinstance(payload, dict) else {}
        _AMAZON_SIGNALS_CACHE = books if isinstance(books, dict) else {}
    except Exception:
        _AMAZON_SIGNALS_CACHE = {}
    return _AMAZON_SIGNALS_CACHE


def amazon_signal_for(book: Dict[str, Any]) -> Dict[str, Any]:
    signals = load_amazon_book_signals()
    asin = clean_paragraph(book.get("asin", "")).upper()
    slug = clean_paragraph(book.get("slug", ""))
    candidate = signals.get(asin) or signals.get(slug) or {}
    if not isinstance(candidate, dict):
        return {}
    if not clean_paragraph(candidate.get("source_url", "")) or not clean_paragraph(candidate.get("checked_at", "")):
        return {}
    return candidate


def render_book_market_signal(book: Dict[str, Any]) -> str:
    """Amazon purchase CTAs remain available without volatile price/rating copy."""
    return ""


def render_inline_newsletter_form(
    source: str,
    next_path: str = "/downloads/ai-glossary-cheat-sheet/",
    cta: str = "Join AI Edge",
    heading: str = "Get the free AI glossary with AI Edge",
    description: str = "Practical AI analysis, plus the plain-English AI glossary. No duplicate form, no second signup route.",
) -> str:
    """Render a lightweight tracked route into the single governed Jotform signup.

    AI Edge collection is intentionally centralised on /newsletter/. Source and
    placement travel in the query string and are forwarded into Jotform there.
    ``next_path`` is retained for call-site compatibility but no longer creates
    competing post-submit behaviour on product/content pages.
    """
    query = urlencode({"source": source, "placement": "inline"})
    href = f"/newsletter/?{query}"
    return f'''<div class="inline-newsletter" data-newsletter-shell data-newsletter-source="{html.escape(source)}">
  <h3>{html.escape(heading)}</h3>
  <p>{html.escape(description)}</p>
  <a class="button" href="{html.escape(href, quote=True)}" data-newsletter-cta data-placement="{html.escape(source)}">{html.escape(cta)}</a>
  <p class="newsletter-fallback-copy"><a href="/downloads/ai-glossary-cheat-sheet/">Already subscribed? Get the AI glossary</a></p>
</div>'''


def newsletter_offer_for_book(book: Dict[str, Any]) -> Dict[str, str]:
    topic = clean_paragraph(book.get("topic_slug", ""))
    mapping = {
        "future-of-work": ("/resources/uk-workplace-ai-literacy-checklist/", "Get AI Edge + the workplace AI literacy checklist"),
        "business": ("/resources/ai-procurement-questions-for-small-businesses/", "Get AI Edge + the small-business AI procurement checklist"),
        "law": ("/resources/responsible-ai-checklist-for-managers/", "Get AI Edge + the responsible AI manager checklist"),
        "finance": ("/resources/ai-regulated-industries-evidence-map/", "Get AI Edge + the regulated-industries evidence map"),
        "cyber-security": ("/resources/deepfake-verification-checklist/", "Get AI Edge + the verification checklist"),
        "media": ("/resources/deepfake-verification-checklist/", "Get AI Edge + the verification checklist"),
    }
    next_path, heading = mapping.get(topic, ("/downloads/ai-glossary-cheat-sheet/", "Get AI Edge + the free AI glossary"))
    return {
        "next_path": next_path,
        "heading": heading,
        "description": "Practical AI analysis plus a useful resource matched to what you are reading now.",
    }


PRIORITY_BOOK_EVIDENCE_MAP = {
    "ai-literacy-for-the-modern-workplace": "workplace-ai-literacy",
    "the-artificial-intelligence-job-shift-navigating-the-future-of-work": "workplace-ai-literacy",
    "ai-agents-for-everyday-work": "ai-agents-for-ordinary-work",
    "artificial-intelligence-for-small-business": "ai-for-small-business",
    "artificial-intelligence-and-the-law-case-studies-and-future-trends": "ai-governance-and-law",
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information": "ai-governance-and-law",
    "artificial-intelligence-in-banking-revolutionizing-finance-and-data-security": "ai-in-finance",
    "digital-diagnosis-how-ai-is-revolutionizing-healthcare": "ai-in-healthcare",
    "artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention": "deepfake-detection-and-synthetic-media",
    "deepfakes-ai-scams-and-synthetic-reality": "deepfake-detection-and-synthetic-media",
}


def render_priority_evidence_module(book: Dict[str, Any]) -> str:
    evidence_slug = PRIORITY_BOOK_EVIDENCE_MAP.get(clean_paragraph(book.get("slug", "")))
    if not evidence_slug:
        return ""
    try:
        payload = json.loads((DATA_DIR / "evidence-content.json").read_text(encoding="utf-8"))
    except Exception:
        return ""
    items = payload.get("items", []) if isinstance(payload, dict) else []
    item = next(
        (x for x in items if isinstance(x, dict) and clean_paragraph(x.get("slug", "")) == evidence_slug),
        None,
    )
    if not item:
        return ""
    stats = item.get("stats", []) if isinstance(item.get("stats"), list) else []
    questions = item.get("questions", []) if isinstance(item.get("questions"), list) else []
    claim_html = ""
    if stats and isinstance(stats[0], dict):
        stat = stats[0]
        source = stat.get("source", {}) if isinstance(stat.get("source"), dict) else {}
        claim = clean_paragraph(stat.get("claim", ""))
        organisation = clean_paragraph(source.get("organisation", ""))
        publication_date = clean_paragraph(source.get("publication_date", ""))
        source_url = clean_paragraph(source.get("url", ""))
        if claim and source_url:
            source_label = organisation or "Primary source"
            date_suffix = f" · {html.escape(publication_date)}" if publication_date else ""
            claim_html = (
                f'<p><strong>Current evidence:</strong> {html.escape(claim)} '
                f'<a href="{html.escape(source_url, quote=True)}" rel="noopener">{html.escape(source_label)}</a>'
                f'{date_suffix}.</p>'
            )
    counterpoint = clean_paragraph(item.get("counterpoint", ""))
    question_items = "".join(
        f'<li>{html.escape(clean_paragraph(q.get("q", "")))}</li>'
        for q in questions[:4]
        if isinstance(q, dict) and clean_paragraph(q.get("q", ""))
    )
    return f'''<section class="card ebook-section ebook-current-evidence" aria-labelledby="current-evidence-{html.escape(evidence_slug, quote=True)}">
      <h2 id="current-evidence-{html.escape(evidence_slug, quote=True)}">Current evidence and decision questions</h2>
      {claim_html}
      <h3>What often gets oversimplified</h3>
      <p>{html.escape(counterpoint)}</p>
      <h3>Questions worth asking before acting</h3>
      <ul class="ebook-learn-list">{question_items}</ul>
      <p><a href="/evidence/{html.escape(evidence_slug, quote=True)}/">Read the full source-backed evidence guide</a>.</p>
    </section>'''


def podcast_link_for_book(book: Dict[str, Any]) -> Dict[str, str]:
    topic = clean_paragraph(book.get("topic_slug", ""))
    label = f"Turing’s Torch on {clean_paragraph(book.get('topic', 'AI'))}"
    return {"href": f"/podcast/?topic={topic}" if topic else "/podcast/", "label": label}


def render_book_bundle_links(book: Dict[str, Any]) -> str:
    try:
        payload = json.loads((DATA_DIR / "ebook-bundles.json").read_text(encoding="utf-8"))
        bundles = payload.get("bundles", []) if isinstance(payload, dict) else []
    except Exception:
        bundles = []
    links = []
    for bundle in bundles if isinstance(bundles, list) else []:
        if not isinstance(bundle, dict) or book.get("slug") not in bundle.get("books", []):
            continue
        slug = clean_paragraph(bundle.get("slug", ""))
        title = clean_paragraph(bundle.get("title", ""))
        if slug and title:
            links.append(f'<a class="button secondary" href="/bundles/{html.escape(slug)}/">Reading path: {html.escape(title)}</a>')
    return "".join(links)


def book_preview_path(book: Dict[str, Any]) -> str:
    return f"/ebooks/{book['slug']}/sample/"


def load_book_sample_chapters() -> Dict[str, Dict[str, Any]]:
    if not BOOK_SAMPLE_CHAPTERS_PATH.exists():
        return {}
    try:
        payload = json.loads(BOOK_SAMPLE_CHAPTERS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    books = payload.get("books", []) if isinstance(payload, dict) else []
    return {
        clean_paragraph(item.get("slug", "")): item
        for item in books
        if isinstance(item, dict) and clean_paragraph(item.get("slug", ""))
    }


def render_book_sample_page(book: Dict[str, Any]) -> str:
    canonical = f"{SITE_URL}{book_preview_path(book)}"
    title = html.escape(book["title"])
    sample = load_book_sample_chapters().get(book["slug"])
    header = render_header()
    footer = render_footer()

    if not sample or not sample.get("paragraphs"):
        # Production CI requires a genuine manuscript extraction before release.
        # This state exists only so local diagnostics fail honestly instead of
        # manufacturing a "preview" from marketing copy.
        schema = {
            "@context": "https://schema.org",
            "@type": "WebPage",
            "name": f"Sample chapter: {book['title']}",
            "url": canonical,
            "description": f"Sample chapter for {book['title']}.",
            "isPartOf": {"@type": "Book", "name": book["title"], "url": book["canonical_url"]},
            "author": {"@id": f"{SITE_URL}/#person"},
            "inLanguage": "en-GB",
        }
        return f'''<!doctype html>
<html lang="en-GB"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
<title>Sample chapter: {title} | Jonathan Harris</title><meta name="robots" content="noindex,follow"/><link rel="canonical" href="{canonical}"/>
<script type="application/ld+json">{json.dumps(schema, ensure_ascii=False)}</script>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link rel="stylesheet" href="/assets/css/site.css"/><link rel="stylesheet" href="/assets/css/ebook-template.css"/></head>
<body class="ebook-detail">{header}<main class="main" id="main"><div class="wrap ebook-shell"><section class="card ebook-section">
<p class="eyebrow">Sample chapter</p><h1>{title}</h1><p>The manuscript chapter has not been extracted for this build, so no substitute preview is being shown.</p>
<div class="ebook-actions"><a class="button" href="{html.escape(book['buy_route'])}">Buy on Amazon</a><a class="button secondary" href="/ebooks/{html.escape(book['slug'])}/">Back to book page</a></div>
</section></div></main>{footer}<script defer src="/assets/js/site-ui.min.js"></script></body></html>'''

    chapter_title = html.escape(clean_paragraph(sample.get("chapter_title", "Chapter sample")))
    paragraphs = sample.get("paragraphs", [])
    body = "".join(
        f"<p>{html.escape(clean_paragraph(paragraph))}</p>"
        for paragraph in paragraphs
        if clean_paragraph(paragraph)
    )
    schema = {
        "@context": "https://schema.org",
        "@type": "Article",
        "@id": f"{canonical}#article",
        "headline": f"{sample.get('chapter_title', 'Sample chapter')} — {book['title']}",
        "url": canonical,
        "description": f"Read a genuine sample chapter from {book['title']} by Jonathan Harris.",
        "datePublished": book["datePublished"],
        "dateModified": book.get("dateModified") or book["datePublished"],
        "mainEntityOfPage": {"@type": "WebPage", "@id": canonical},
        "isPartOf": {"@type": "Book", "name": book["title"], "url": book["canonical_url"]},
        "author": {"@id": f"{SITE_URL}/#person"},
        "inLanguage": "en-GB",
    }
    description = html.escape(
        f"Read a genuine sample chapter from {book['title']} by Jonathan Harris before buying the Kindle ebook.",
        quote=True,
    )
    return f'''<!doctype html>
<html lang="en-GB"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
<title>Free chapter: {title} | Jonathan Harris</title><meta name="description" content="{description}"/><meta name="robots" content="index,follow"/><link rel="canonical" href="{canonical}"/>
<script type="application/ld+json">{json.dumps(schema, ensure_ascii=False)}</script>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link rel="stylesheet" href="/assets/css/site.css"/><link rel="stylesheet" href="/assets/css/ebook-template.css"/></head>
<body class="ebook-detail ebook-sample">{header}<main class="main" id="main"><div class="wrap ebook-shell">
<nav aria-label="Breadcrumb" class="breadcrumbs"><a href="/">Home</a> / <a href="/ebooks/">eBooks</a> / <a href="/ebooks/{html.escape(book['slug'])}/">{title}</a> / Sample chapter</nav>
<section class="card ebook-section ebook-sample-intro"><p class="eyebrow">Read before you buy</p><h1>{title}</h1><h2>{chapter_title}</h2><p>This page contains text extracted from the actual manuscript, not a summary or generated substitute.</p><div class="ebook-actions"><a class="button" href="{html.escape(book['buy_route'])}" data-ebook-amazon data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="sample_top">Buy the full book on Amazon</a><a class="button secondary" href="/ebooks/{html.escape(book['slug'])}/">Book details</a></div></section>
<article class="card ebook-section ebook-sample-chapter" aria-labelledby="sample-chapter-heading"><h2 id="sample-chapter-heading">{chapter_title}</h2>{body}<p class="meta">Genuine manuscript extract.</p></article>
<section class="card ebook-section ebook-section--accent"><h2>Continue reading</h2><p>If this chapter is useful, the full Kindle ebook continues the argument with the rest of the book’s practical guidance.</p><div class="ebook-actions"><a class="button" href="{html.escape(book['buy_route'])}" data-ebook-amazon data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="sample_bottom">Buy on Amazon</a><a class="button secondary" href="/newsletter/">Get the free AI glossary</a></div></section>
</div></main>{footer}<script defer src="/assets/js/funnel-events.min.js"></script><script defer src="/assets/js/site-ui.min.js"></script></body></html>'''

def build_book_schema(book: Dict[str, Any]) -> Dict[str, Any]:
    about_terms = [{"@type": "Thing", "name": name} for name in book_about_terms(book)]
    schema = {
        "@context": "https://schema.org",
        "@type": "Book",
        "@id": f"{book['canonical_url']}#book",
        "name": book["title"],
        "url": book["canonical_url"],
        "description": book["description"],
        "image": [book["cover"]],
        "author": {"@id": f"{SITE_URL}/#person"},
        "bookFormat": "EBook",
        "datePublished": book["datePublished"],
        "dateModified": book.get("dateModified") or infer_build_timestamp(),
        "inLanguage": "en-GB",
        "numberOfPages": book["pages"],
        "sameAs": [book["buy_url"]] if book.get("buy_url") else [],
        "publisher": {"@id": f"{SITE_URL}/#person"},
        "identifier": [
            {"@type": "PropertyValue", "propertyID": "ASIN", "value": book["asin"]},
            {"@type": "PropertyValue", "propertyID": "Jonathan Harris internal identifier", "value": book["identifier"]},
        ],
        "about": about_terms,
    }
    signal = amazon_signal_for(book)
    if signal.get("rating") is not None and signal.get("rating_count") is not None:
        schema["aggregateRating"] = {
            "@type": "AggregateRating",
            "ratingValue": float(signal["rating"]),
            "ratingCount": int(signal["rating_count"]),
            "bestRating": 5,
            "worstRating": 1,
        }
    return schema


def build_breadcrumb_schema(book: Dict[str, Any]) -> Dict[str, Any]:
    items = [
        {"@type": "ListItem", "position": 1, "name": "Home", "item": f"{SITE_URL}/"},
        {"@type": "ListItem", "position": 2, "name": "eBooks", "item": f"{SITE_URL}/ebooks/"},
    ]
    position = 3
    if book.get("topic_url"):
        items.append({"@type": "ListItem", "position": position, "name": book["topic"], "item": f"{SITE_URL}{book['topic_url']}"})
        position += 1
    items.append({"@type": "ListItem", "position": position, "name": book["title"], "item": book["canonical_url"]})
    return {"@context": "https://schema.org", "@type": "BreadcrumbList", "itemListElement": items}



def build_topic_breadcrumb_schema(topic: str) -> Dict[str, Any]:
    topic_slug = slugify(topic)
    return {
        "@context": "https://schema.org",
        "@type": "BreadcrumbList",
        "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "Home", "item": f"{SITE_URL}/"},
            {"@type": "ListItem", "position": 2, "name": "Topics", "item": f"{SITE_URL}/topics/"},
            {"@type": "ListItem", "position": 3, "name": topic, "item": f"{SITE_URL}/catalogue/{topic_slug}/"},
        ],
    }



def render_topic_breadcrumbs(topic: str) -> str:
    return "".join([
        '<a href="/">Home</a>',
        '<span aria-hidden="true">›</span>',
        '<a href="/topics/">Topics</a>',
        '<span aria-hidden="true">›</span>',
        f'<span>{html.escape(topic)}</span>',
    ])



def parse_master_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    sanitise_content: bool = True,
) -> Dict[str, Dict[str, Any]]:
    headers = [clean_paragraph(cell.value).lower() for cell in ws[1]]
    index = {header: idx for idx, header in enumerate(headers) if header}
    required = {"slug", "title", "asin", "amazon ebook page count", "publication date", "book url", "buy now url", "redirect url", "cover art url", "legacy alias url"}
    missing = required - set(index)
    if missing:
        raise ValueError(f"Ebooks Master sheet is missing required columns: {', '.join(sorted(missing))}")

    field_map = {
        "slug": "slug",
        "title": "title",
        "short description": "short",
        "short": "short",
        "description": "description",
        "summary": "summary",
        "topic": "topic",
        "category": "topic",
        "tags": "tags",
        "keywords": "keywords",
        "audience": "audience",
        "who this book is for": "who_for",
        "who_for": "who_for",
        "what this book covers": "what_this_book_covers",
        "what_youll_learn": "what_youll_learn",
        "what you'll learn": "what_youll_learn",
        "why it matters": "why_it_matters",
        "tone": "tone",
        "author": "author",
        "identifier": "identifier",
        "showcase heading": "showcase_heading",
        "distinct angle": "distinct_angle",
        "asin": "asin",
        "amazon ebook page count": "pages",
        "publication date": "datePublished",
        "book url": "book_url",
        "buy now url": "buy_route_full",
        "redirect url": "buy_url",
        "cover art url": "cover",
        "legacy alias url": "legacy_alias_url",
    }
    list_fields = {"tags", "keywords", "what_youll_learn"}

    results: Dict[str, Dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        slug = clean_paragraph(row[index["slug"]] if index.get("slug") is not None else "")
        if not slug:
            continue
        record: Dict[str, Any] = {}
        for header, idx in index.items():
            target = field_map.get(header)
            if not target:
                continue
            value = row[idx]
            if target in list_fields:
                record[target] = split_field(value)
            elif target == "pages":
                record[target] = int(value) if value not in (None, "") else None
            elif target == "datePublished":
                if isinstance(value, dt.datetime):
                    record[target] = value.date().isoformat()
                elif hasattr(value, "isoformat") and value is not None:
                    record[target] = value.isoformat()
                else:
                    record[target] = clean_paragraph(value)
            else:
                record[target] = clean_paragraph(value)
        record["book_url"] = ensure_trailing_slash(record.get("book_url") or f"{SITE_URL}/ebooks/{slug}/")
        buy_route_full = record.get("buy_route_full") or f"{SITE_URL}/ebooks/{slug}/buy-now"
        record["buy_route_full"] = buy_route_full
        record["buy_route"] = url_to_path(buy_route_full)
        record["legacy_alias_url"] = clean_paragraph(record.get("legacy_alias_url") or f"{SITE_URL}/book/{slug}/buy-now")
        record["cover"] = clean_paragraph(record.get("cover", "")).replace("https://images.Jonathan-harris.online", "https://images.jonathan-harris.online")
        if sanitise_content:
            record = sanitise_record_copy(record)
        results[slug] = record
    return results


def parse_workbook(
    workbook_path: Path,
    *,
    sanitise_content: bool = True,
) -> Tuple[List[str], Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if "Ebooks Master" not in wb.sheetnames:
        raise ValueError("Workbook is missing the Ebooks Master sheet")

    master_sheet = parse_master_sheet(wb["Ebooks Master"], sanitise_content=sanitise_content)
    if not master_sheet:
        raise ValueError("No ebook rows found in Ebooks Master")

    notes_by_slug: Dict[str, str] = {}
    if "BuyNow Redirects" in wb.sheetnames:
        ws = wb["BuyNow Redirects"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            slug = clean_paragraph(row[0] or "")
            if not slug:
                continue
            notes_by_slug[slug] = clean_paragraph(row[5] or "")

    order = list(master_sheet.keys())
    records: Dict[str, Dict[str, Any]] = {}
    for slug in order:
        row = dict(master_sheet[slug])
        row["slug"] = slug
        row["notes"] = notes_by_slug.get(slug, "")
        records[slug] = row
    return order, records, master_sheet


def validate_pages_sheet_operational_view(workbook_path: Path) -> List[str]:
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True)
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False)
    if "Pages" not in wb_values.sheetnames:
        return ["Workbook is missing the Pages sheet."]

    ws = wb_values["Pages"]
    ws_formula = wb_formulas["Pages"]
    header_row = 5
    headers = {clean_paragraph(ws.cell(header_row, col).value).lower(): col for col in range(1, ws.max_column + 1) if clean_paragraph(ws.cell(header_row, col).value)}
    required_headers = ["relative file path", "public url path", "full url", "url type", "buy now url", "redirect url", "asin", "amazon ebook page count", "publication date", "cover art url"]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        return [f"Pages sheet is missing required columns: {', '.join(missing_headers)}"]

    base_domain = clean_paragraph(ws["B1"].value)
    errors: List[str] = []
    malformed_paths: List[str] = []
    missing_full_url: List[str] = []
    mismatched_full_url: List[str] = []
    missing_ebook_metadata: List[str] = []
    master_by_slug = {row["slug"]: row for row in load_master()}

    def has_formula(row_idx: int, header_name: str) -> bool:
        cell = ws_formula.cell(row_idx, headers[header_name])
        return isinstance(cell.value, str) and cell.value.startswith("=")

    def slug_from_relative_path(relative_path: str) -> str:
        if relative_path.startswith("ebooks/") and relative_path.endswith("/index.html"):
            return relative_path[len("ebooks/"):-len("/index.html")]
        return ""

    for row_idx in range(header_row + 1, ws.max_row + 1):
        relative_path = clean_paragraph(ws.cell(row_idx, headers["relative file path"]).value)
        if not relative_path:
            continue

        public_path = clean_paragraph(ws.cell(row_idx, headers["public url path"]).value)
        full_url = clean_paragraph(ws.cell(row_idx, headers["full url"]).value)
        url_type = clean_paragraph(ws.cell(row_idx, headers["url type"]).value).lower()
        buy_now_url = clean_paragraph(ws.cell(row_idx, headers["buy now url"]).value)
        redirect_url = clean_paragraph(ws.cell(row_idx, headers["redirect url"]).value)
        asin = clean_paragraph(ws.cell(row_idx, headers["asin"]).value)
        page_count = ws.cell(row_idx, headers["amazon ebook page count"]).value
        publication_date = ws.cell(row_idx, headers["publication date"]).value
        cover_art_url = clean_paragraph(ws.cell(row_idx, headers["cover art url"]).value)

        if public_path:
            path_without_leading_slash = public_path[1:] if public_path.startswith("/") else public_path
            if "//" in path_without_leading_slash:
                malformed_paths.append(relative_path)
            if not public_path.startswith("/"):
                malformed_paths.append(relative_path)
        else:
            malformed_paths.append(relative_path)

        if public_path:
            expected_full_url = f"{base_domain}{public_path}" if base_domain else ""
            if not full_url and not has_formula(row_idx, "full url"):
                missing_full_url.append(relative_path)
            elif full_url and expected_full_url and full_url != expected_full_url:
                mismatched_full_url.append(relative_path)

        is_canonical_ebook_row = url_type == "canonical ebook route"
        if is_canonical_ebook_row:
            slug = slug_from_relative_path(relative_path)
            master_row = master_by_slug.get(slug, {})
            metadata_checks = [
                (buy_now_url, "buy now url", clean_paragraph(master_row.get("buy_route", ""))),
                (redirect_url, "redirect url", clean_paragraph(master_row.get("buy_url", ""))),
                (asin, "asin", clean_paragraph(master_row.get("asin", ""))),
                (page_count, "amazon ebook page count", master_row.get("pages")),
                (publication_date, "publication date", clean_paragraph(master_row.get("datePublished", ""))),
                (cover_art_url, "cover art url", clean_paragraph(master_row.get("cover", ""))),
            ]
            for value, header_name, derived_value in metadata_checks:
                if value not in (None, ""):
                    continue
                if has_formula(row_idx, header_name) and derived_value not in (None, ""):
                    continue
                missing_ebook_metadata.append(relative_path)
                break

    if malformed_paths:
        sample = ", ".join(malformed_paths[:5])
        errors.append(f"Workbook Pages sheet has {len(malformed_paths)} malformed public URL path value(s) with double-slash or missing leading slash drift. Sample rows: {sample}")
    if missing_full_url:
        sample = ", ".join(missing_full_url[:5])
        errors.append(f"Workbook Pages sheet has {len(missing_full_url)} row(s) with missing cached Full URL values. Sample rows: {sample}")
    if mismatched_full_url:
        sample = ", ".join(mismatched_full_url[:5])
        errors.append(f"Workbook Pages sheet has {len(mismatched_full_url)} row(s) where Full URL does not match Base domain + Public URL path. Sample rows: {sample}")
    if missing_ebook_metadata:
        sample = ", ".join(missing_ebook_metadata[:5])
        errors.append(f"Workbook Pages sheet has {len(missing_ebook_metadata)} canonical ebook row(s) missing cached metadata lookup values. Sample rows: {sample}")
    return errors


def topic_intro(topic: str) -> str:
    return helper_topic_intro(topic)



def default_learning_points(topic: str) -> List[str]:
    topic_lc = topic.lower()
    return [
        f"How AI is already being used in {topic_lc} and where the claims run ahead of the evidence.",
        f"The workflows, trade-offs, and decision points that matter in {topic_lc}.",
        f"The awkward questions around risk, adoption, governance, and long-term impact in {topic_lc}.",
    ]


BOOK_SPECIFIC_LEARN_TAILS: Dict[str, str] = {'ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology': 'where milliseconds beat marketing '
                                                                               'myths',
 'ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future': 'where weather, margins, and soil matter',
 'ai-in-aviation-transforming-safety-and-sustainability': 'where safety systems meet hard limits',
 'ai-in-education-reimagining-learning-for-every-student': 'where classroom goals beat shiny demos',
 'ai-in-maritime-revolutionizing-shipping-for-sustainability': 'where ports, fleets, and costs collide',
 'ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation': 'where grid maths meets field '
                                                                                  'reality',
 'ai-revolution-in-railways-modernizing-travel-for-a-smarter-future': 'where signalling meets service reality',
 'artificial-intelligence-and-the-law-case-studies-and-future-trends': 'where precedent meets messy deployment',
 'artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention': 'where attacks outpace '
                                                                                           'slide decks',
 'artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology': 'where '
                                                                                                                 'field '
                                                                                                                 'data '
                                                                                                                 'meets '
                                                                                                                 'habitat '
                                                                                                                 'limits',
 'artificial-intelligence-in-banking-revolutionizing-finance-and-data-security': 'where fraud models meet compliance '
                                                                                 'desks',
 'artificial-intelligence-in-construction-building-a-sustainable-future': 'where sites, crews, and delays decide',
 'artificial-intelligence-in-industry-a-comprehensive-guide': 'where plant floors test every promise',
 'artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability': 'where routing models hit warehouse '
                                                                                  'floors',
 'artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare': 'where lab promise meets regulation',
 'artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement': 'where performance data meets '
                                                                                     'human nerves',
 'artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation': 'where '
                                                                                                     'diagnostics meet '
                                                                                                     'day-to-day care',
 'artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future': 'where basket '
                                                                                                        'data meets '
                                                                                                        'buyer '
                                                                                                        'patience',
 'artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery': 'where '
                                                                                                                'uptime, '
                                                                                                                'scrap, '
                                                                                                                'and '
                                                                                                                'margins '
                                                                                                                'bite',
 'beyond-earth-how-ai-is-transforming-space-exploration': 'where mission risk beats sci-fi gloss',
 'climate-intelligence-harnessing-ai-for-a-greener-future': 'where emissions maths meets policy friction',
 'digital-defense-the-role-of-ai-in-modern-warfare': 'where autonomy meets command risk',
 'digital-diagnosis-how-ai-is-revolutionizing-healthcare': 'where clinical use meets human judgement',
 'from-reporters-to-robots-how-ai-is-reshaping-journalism': 'where deadlines meet verification pressure',
 'game-ai-unleashed-from-finite-state-machines-to-machine-learning': 'where design tricks meet player intent',
 'lights-camera-algorithm-ai-s-role-in-modern-filmmaking': 'where production budgets meet creative calls',
 'smart-buildings-ai-powered-efficiency-and-sustainability': 'where sensors meet maintenance budgets',
 'the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media': 'where feeds meet moderation '
                                                                                      'blowback',
 'the-ai-music-revolution-creativity-controversy-and-collaboration': 'where new tools meet old rights fights',
 'the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead': 'where breakthroughs meet the human cost',
 'the-artificial-intelligence-job-shift-navigating-the-future-of-work': 'where job redesign beats slogan fog',
 'the-artificial-intelligence-revolution-from-algorithms-to-consciousness': 'where theory collides with deployment',
 'the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry': 'where sensors meet '
                                                                                                'traffic and liability',
 'the-dumbening-how-ai-is-reshaping-our-minds': 'where convenience starts taxing attention',
 'the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information': 'where services meet scrutiny '
                                                                                         'and budget',
 'the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming': 'where optimisation meets ethical lines'}

BOOK_SPECIFIC_ADVANCED_TAILS: Dict[str, str] = {'ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology': 'strategy calls, telemetry, and pace',
 'ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future': 'yield data, labour, and seasons',
 'ai-in-aviation-transforming-safety-and-sustainability': 'flight safety, failure modes, and ops',
 'ai-in-education-reimagining-learning-for-every-student': 'learning design, bias, and outcomes',
 'ai-in-maritime-revolutionizing-shipping-for-sustainability': 'port systems, delays, and fuel use',
 'ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation': 'grid load, faults, and resilience',
 'ai-revolution-in-railways-modernizing-travel-for-a-smarter-future': 'signalling, safety, and uptime',
 'artificial-intelligence-and-the-law-case-studies-and-future-trends': 'case law, compliance, and grey areas',
 'artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention': 'attack paths, defence, and '
                                                                                           'risk',
 'artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology': 'field '
                                                                                                                 'sensing, '
                                                                                                                 'habitat, '
                                                                                                                 'and '
                                                                                                                 'risk',
 'artificial-intelligence-in-banking-revolutionizing-finance-and-data-security': 'fraud controls, models, and trust',
 'artificial-intelligence-in-construction-building-a-sustainable-future': 'site ops, safety, and delays',
 'artificial-intelligence-in-industry-a-comprehensive-guide': 'industrial systems, risk, and return',
 'artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability': 'routing logic, stock flow, and '
                                                                                  'timing',
 'artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare': 'trials, regulation, and lab friction',
 'artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement': 'performance data, tactics, and '
                                                                                     'nerves',
 'artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation': 'case triage, '
                                                                                                     'trust, and '
                                                                                                     'treatment',
 'artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future': 'customer '
                                                                                                        'data, stock, '
                                                                                                        'and margin',
 'artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery': 'uptime, '
                                                                                                                'quality, '
                                                                                                                'and '
                                                                                                                'plant '
                                                                                                                'reality',
 'beyond-earth-how-ai-is-transforming-space-exploration': 'mission risk, autonomy, and delay',
 'climate-intelligence-harnessing-ai-for-a-greener-future': 'carbon maths, policy, and scale',
 'digital-defense-the-role-of-ai-in-modern-warfare': 'command risk, autonomy, and doctrine',
 'digital-diagnosis-how-ai-is-revolutionizing-healthcare': 'clinical judgement, data, and harm',
 'from-reporters-to-robots-how-ai-is-reshaping-journalism': 'verification, deadlines, and pressure',
 'game-ai-unleashed-from-finite-state-machines-to-machine-learning': 'npc logic, difficulty, and design',
 'lights-camera-algorithm-ai-s-role-in-modern-filmmaking': 'workflows, budgets, and authorship',
 'smart-buildings-ai-powered-efficiency-and-sustainability': 'building ops, costs, and comfort',
 'the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media': 'feeds, moderation, and '
                                                                                      'influence',
 'the-ai-music-revolution-creativity-controversy-and-collaboration': 'rights fights, tools, and taste',
 'the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead': 'history, motives, and fallout',
 'the-artificial-intelligence-job-shift-navigating-the-future-of-work': 'job design, skills, and leverage',
 'the-artificial-intelligence-revolution-from-algorithms-to-consciousness': 'theory, limits, and real deployment',
 'the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry': 'road risk, autonomy, '
                                                                                                'and liability',
 'the-dumbening-how-ai-is-reshaping-our-minds': 'attention, habits, and drift',
 'the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information': 'public trust, systems, and '
                                                                                         'budgets',
 'the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming': 'risk scoring, consent, and harm'}

CATALOGUE_INTRO_VARIANTS: Dict[str, str] = {'Agriculture': 'These books cut through ag-tech waffle and focus on yields, soil, labour, and what works in the '
                'field.',
 'Artificial Intelligence': 'These titles zoom out from slogans and look at how AI behaves once theory meets products, '
                            'decisions, and people.',
 'Construction': 'These books stay grounded in sites, schedules, safety, and the stubborn reality of getting '
                 'technology to work.',
 'Creativity': 'These titles treat creative work as craft and business, not pixie dust, from studios and sets to the '
               'rights mess behind them.',
 'Cyber Security': 'These books focus on defensive reality: attack surfaces, breach prevention, response pressure, and '
                   'what actually improves resilience.',
 'Defence': 'These titles keep one eye on capability and the other on doctrine, escalation, reliability, and the risks '
            'nobody can spin away.',
 'Education': 'These books look at learning, teaching, and assessment with more classroom realism and less '
              'shiny-edtech theatre.',
 'Energy': 'These titles dig into grids, demand, resilience, and the engineering headaches hiding underneath clean '
           'strategic slogans.',
 'Environment': 'These books stay practical about biodiversity, climate, monitoring, and the gap between good intent '
                'and field conditions.',
 'Ethics': 'These titles lean into the awkward bits: incentives, consent, addiction, cognition, and the costs of '
           'pretending AI is neutral.',
 'Finance': 'These books deal in fraud, risk, security, and trust, not fintech smoke and mirrors.',
 'Future of Work': 'These titles examine what changes in jobs, skills, power, and management once AI leaves the '
                   'keynote stage.',
 'Gaming': 'These books look past gamer jargon and get into design logic, player behaviour, and the machinery behind '
           'believable systems.',
 'Government': 'These titles focus on service delivery, security, accountability, and the bureaucratic plumbing that '
               'makes or breaks adoption.',
 'Healthcare': 'These books keep the focus on diagnosis, care, evidence, and the human judgement that software still '
               'cannot replace.',
 'History': 'These titles trace how AI got here, who shaped it, and why the old breakthroughs still cast a long '
            'shadow.',
 'Industry': 'These books centre on operations, reliability, process change, and the difference between automation '
             'talk and plant-floor reality.',
 'Law': 'These titles deal with precedent, liability, evidence, and the legal friction that appears when code meets '
        'institutions.',
 'Manufacturing': 'These books concentrate on uptime, quality, maintenance, and the grimly practical details that '
                  'decide whether rollout pays off.',
 'Media': 'These titles unpack how AI collides with journalism, feeds, moderation, and the attention economy that '
          'warps the whole lot.',
 'Retail': 'These books stay close to customer behaviour, stock flow, pricing, and the commercial trade-offs behind '
           'personalisation.',
 'Science': 'These titles look at exploration, discovery, uncertainty, and what happens when ambitious models meet '
            'real-world constraints.',
 'Sports': 'These books cut through performance hype and get into coaching decisions, analytics, fan dynamics, and '
           'pressure.',
 'Transportation': 'These titles track how AI changes movement systems in the wild: roads, rails, fleets, ports, and '
                   'split-second decisions.'}

CATALOGUE_CTA_VARIANTS: Dict[str, str] = {'Agriculture': 'Open the full breakdown, then decide whether it earns a place on your Kindle.',
 'Artificial Intelligence': 'Read the detailed page, then see the current Amazon listing when you are ready.',
 'Construction': 'Start with the full overview, then check the live Amazon page for the latest details.',
 'Creativity': 'Open the full description first, then head to Amazon if the angle fits what you need.',
 'Cyber Security': 'Read the deeper summary, then inspect the Amazon page for the current edition and price.',
 'Defence': 'Check the complete book page, then use Amazon for the latest buying details.',
 'Education': 'Open the full breakdown, then decide whether it earns a place on your Kindle.',
 'Energy': 'Read the detailed page, then see the current Amazon listing when you are ready.',
 'Environment': 'Start with the full overview, then check the live Amazon page for the latest details.',
 'Ethics': 'Open the full description first, then head to Amazon if the angle fits what you need.',
 'Finance': 'Read the deeper summary, then inspect the Amazon page for the current edition and price.',
 'Future of Work': 'Check the complete book page, then use Amazon for the latest buying details.',
 'Gaming': 'Open the full breakdown, then decide whether it earns a place on your Kindle.',
 'Government': 'Read the detailed page, then see the current Amazon listing when you are ready.',
 'Healthcare': 'Start with the full overview, then check the live Amazon page for the latest details.',
 'History': 'Open the full description first, then head to Amazon if the angle fits what you need.',
 'Industry': 'Read the deeper summary, then inspect the Amazon page for the current edition and price.',
 'Law': 'Check the complete book page, then use Amazon for the latest buying details.',
 'Manufacturing': 'Open the full breakdown, then decide whether it earns a place on your Kindle.',
 'Media': 'Read the detailed page, then see the current Amazon listing when you are ready.',
 'Retail': 'Start with the full overview, then check the live Amazon page for the latest details.',
 'Science': 'Open the full description first, then head to Amazon if the angle fits what you need.',
 'Sports': 'Read the deeper summary, then inspect the Amazon page for the current edition and price.',
 'Transportation': 'Check the complete book page, then use Amazon for the latest buying details.'}

BOOK_SPECIFIC_SHORT_OVERRIDES: Dict[str, str] = {'ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology': 'A practical look '
                                                                               'at AI in race '
                                                                               'strategy, '
                                                                               'simulation, '
                                                                               'telemetry, and '
                                                                               'split-second '
                                                                               'decisions, without '
                                                                               'pretending the '
                                                                               'software drives '
                                                                               'the car.',
 'artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation': 'A '
                                                                                                     'grounded '
                                                                                                     'guide '
                                                                                                     'to '
                                                                                                     'diagnostic '
                                                                                                     'tools, '
                                                                                                     'clinic '
                                                                                                     'workflow, '
                                                                                                     'animal '
                                                                                                     'health '
                                                                                                     'data, '
                                                                                                     'and '
                                                                                                     'the '
                                                                                                     'judgement '
                                                                                                     'vets '
                                                                                                     'still '
                                                                                                     'bring '
                                                                                                     'to '
                                                                                                     'the '
                                                                                                     'table.',
 'ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation': 'A clear guide '
                                                                                  'to demand '
                                                                                  'forecasting, '
                                                                                  'grid balancing, '
                                                                                  'renewables, and '
                                                                                  'the awkward bit '
                                                                                  'where energy '
                                                                                  'theory meets '
                                                                                  'field reality.',
 'artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability': 'A practical '
                                                                                  'look at '
                                                                                  'routing, demand '
                                                                                  'forecasting, '
                                                                                  'warehouse '
                                                                                  'automation, and '
                                                                                  'the points '
                                                                                  'where '
                                                                                  'supply-chain '
                                                                                  'theory meets '
                                                                                  'loading-bay '
                                                                                  'reality.',
 'artificial-intelligence-and-the-law-case-studies-and-future-trends': 'A plain-English look at '
                                                                       'legal research, contracts, '
                                                                       'evidence, compliance, and '
                                                                       'the expensive consequences '
                                                                       'of getting automated law '
                                                                       'wrong.',
 'artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention': 'A '
                                                                                           'no-hype '
                                                                                           'guide '
                                                                                           'to '
                                                                                           'threat '
                                                                                           'detection, '
                                                                                           'breach '
                                                                                           'prevention, '
                                                                                           'alert '
                                                                                           'noise, '
                                                                                           'and '
                                                                                           'the '
                                                                                           'security '
                                                                                           'decisions '
                                                                                           'humans '
                                                                                           'still '
                                                                                           'have '
                                                                                           'to '
                                                                                           'own.',
 'the-artificial-intelligence-revolution-from-algorithms-to-consciousness': 'A grounded tour of AI '
                                                                            'history, core ideas, '
                                                                            'big claims, and the '
                                                                            'awkward gap between '
                                                                            'impressive demos and '
                                                                            'actual intelligence.',
 'ai-in-aviation-transforming-safety-and-sustainability': 'A practical guide to predictive '
                                                          'maintenance, air traffic tools, fuel '
                                                          'efficiency, and the safety checks '
                                                          'aviation cannot afford to treat '
                                                          'casually.',
 'ai-in-maritime-revolutionizing-shipping-for-sustainability': 'A clear look at autonomous '
                                                               'vessels, route planning, port '
                                                               'logistics, emissions tracking, and '
                                                               'the rough water beneath the '
                                                               'sustainability pitch.',
 'artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare': 'A plain-English guide '
                                                                          'to drug discovery, '
                                                                          'clinical trials, '
                                                                          'treatment '
                                                                          'personalisation, and '
                                                                          'the regulatory grind '
                                                                          'behind the lab promise.',
 'the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry': 'A '
                                                                                                'grounded '
                                                                                                'look '
                                                                                                'at '
                                                                                                'autonomous '
                                                                                                'vehicles, '
                                                                                                'predictive '
                                                                                                'maintenance, '
                                                                                                'smart '
                                                                                                'manufacturing, '
                                                                                                'and '
                                                                                                'the '
                                                                                                'liability '
                                                                                                'questions '
                                                                                                'parked '
                                                                                                'behind '
                                                                                                'the '
                                                                                                'showroom '
                                                                                                'shine.',
 'artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future': 'A '
                                                                                                        'practical '
                                                                                                        'guide '
                                                                                                        'to '
                                                                                                        'recommendations, '
                                                                                                        'inventory, '
                                                                                                        'pricing, '
                                                                                                        'supply '
                                                                                                        'chains, '
                                                                                                        'and '
                                                                                                        'the '
                                                                                                        'point '
                                                                                                        'where '
                                                                                                        'personalisation '
                                                                                                        'starts '
                                                                                                        'feeling '
                                                                                                        'too '
                                                                                                        'personal.',
 'artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery': 'A '
                                                                                                                'clear '
                                                                                                                'guide '
                                                                                                                'to '
                                                                                                                'predictive '
                                                                                                                'maintenance, '
                                                                                                                'production '
                                                                                                                'planning, '
                                                                                                                'quality '
                                                                                                                'control, '
                                                                                                                'and '
                                                                                                                'what '
                                                                                                                'survives '
                                                                                                                'contact '
                                                                                                                'with '
                                                                                                                'the '
                                                                                                                'plant '
                                                                                                                'floor.',
 'artificial-intelligence-in-industry-a-comprehensive-guide': 'A broad but grounded guide to '
                                                              'automation, analytics, risk, and '
                                                              'the industrial decisions that still '
                                                              'need more than a dashboard.',
 'the-dumbening-how-ai-is-reshaping-our-minds': 'A sceptical look at automation, attention, '
                                                'memory, creativity, and the cost of outsourcing '
                                                'too much thinking to the machine.',
 'ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future': 'A practical guide to crop '
                                                                       'monitoring, precision '
                                                                       'farming, weather risk, '
                                                                       'labour pressure, and the '
                                                                       'field conditions that ruin '
                                                                       'tidy demos.',
 'ai-in-education-reimagining-learning-for-every-student': 'A clear look at personalised learning, '
                                                           'assessment, classroom support, and '
                                                           'where education technology helps '
                                                           'rather than just generating paperwork.',
 'artificial-intelligence-in-banking-revolutionizing-finance-and-data-security': 'A grounded guide '
                                                                                 'to fraud '
                                                                                 'detection, '
                                                                                 'credit scoring, '
                                                                                 'compliance, '
                                                                                 'customer data, '
                                                                                 'and the risks '
                                                                                 'that keep '
                                                                                 'banking AI on a '
                                                                                 'short lead.',
 'digital-diagnosis-how-ai-is-revolutionizing-healthcare': 'A practical look at diagnostics, '
                                                           'triage, imaging, clinical workflow, '
                                                           'and the point where healthcare AI '
                                                           'still needs human judgement.',
 'artificial-intelligence-in-construction-building-a-sustainable-future': 'A plain-English guide '
                                                                          'to planning, safety, '
                                                                          'budgets, materials, and '
                                                                          'why construction AI has '
                                                                          'to work on sites, not '
                                                                          'just slides.',
 'the-artificial-intelligence-job-shift-navigating-the-future-of-work': 'A no-hype look at '
                                                                        'automation, reskilling, '
                                                                        'job redesign, management '
                                                                        'pressure, and what '
                                                                        'changes when AI moves '
                                                                        'into ordinary work.',
 'the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information': 'A '
                                                                                         'plain-English '
                                                                                         'look at '
                                                                                         'AI in '
                                                                                         'public '
                                                                                         'services: '
                                                                                         'what it '
                                                                                         'can '
                                                                                         'improve, '
                                                                                         'where '
                                                                                         'the '
                                                                                         'risks '
                                                                                         'sit, and '
                                                                                         'which '
                                                                                         'questions '
                                                                                         'teams '
                                                                                         'should '
                                                                                         'ask '
                                                                                         'before '
                                                                                         'buying '
                                                                                         'the '
                                                                                         'pitch.',
 'smart-buildings-ai-powered-efficiency-and-sustainability': 'A grounded guide to energy '
                                                             'management, sensors, maintenance, '
                                                             'occupant comfort, and the building '
                                                             'systems that still need human '
                                                             'oversight.',
 'digital-defense-the-role-of-ai-in-modern-warfare': 'A sober guide to drones, intelligence, cyber '
                                                     'operations, autonomy, and the command risks '
                                                     'that military AI cannot wish away.',
 'artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology': 'A '
                                                                                                                 'practical '
                                                                                                                 'guide '
                                                                                                                 'to '
                                                                                                                 'habitat '
                                                                                                                 'monitoring, '
                                                                                                                 'species '
                                                                                                                 'identification, '
                                                                                                                 'anti-poaching '
                                                                                                                 'tools, '
                                                                                                                 'and '
                                                                                                                 'the '
                                                                                                                 'field '
                                                                                                                 'data '
                                                                                                                 'problems '
                                                                                                                 'conservation '
                                                                                                                 'teams '
                                                                                                                 'face.',
 'climate-intelligence-harnessing-ai-for-a-greener-future': 'A clear guide to emissions tracking, '
                                                            'climate modelling, renewable '
                                                            'planning, and the difference between '
                                                            'useful measurement and green gloss.',
 'ai-revolution-in-railways-modernizing-travel-for-a-smarter-future': 'A grounded look at '
                                                                      'predictive maintenance, '
                                                                      'signalling, scheduling, '
                                                                      'safety, and why railway AI '
                                                                      'has to survive real service '
                                                                      'disruption.',
 'artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement': 'A practical '
                                                                                     'look at '
                                                                                     'performance '
                                                                                     'data, injury '
                                                                                     'prevention, '
                                                                                     'scouting, '
                                                                                     'fan '
                                                                                     'engagement, '
                                                                                     'and the '
                                                                                     'judgement '
                                                                                     'that still '
                                                                                     'decides what '
                                                                                     'happens on '
                                                                                     'the pitch.',
 'lights-camera-algorithm-ai-s-role-in-modern-filmmaking': 'A plain-English guide to scripts, '
                                                           'editing, visual effects, production '
                                                           'budgets, and where AI helps filmmaking '
                                                           'without replacing taste.',
 'the-ai-music-revolution-creativity-controversy-and-collaboration': 'A grounded look at '
                                                                     'composition, production, '
                                                                     'rights, collaboration, and '
                                                                     'the row that starts when '
                                                                     'software learns the tune.',
 'the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead': 'A clear guide to the people, '
                                                                   'breakthroughs, false starts, '
                                                                   'and stubborn questions that '
                                                                   'shaped artificial '
                                                                   'intelligence.',
 'the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming': 'A sceptical guide to '
                                                                             'personalisation, '
                                                                             'gambling design, '
                                                                             'addiction risk, '
                                                                             'regulation, and why '
                                                                             'the house rarely '
                                                                             'needs extra help.',
 'beyond-earth-how-ai-is-transforming-space-exploration': 'A practical look at rovers, mission '
                                                          'planning, data analysis, autonomy, and '
                                                          'why space AI has no room for cheerful '
                                                          'guesswork.',
 'from-reporters-to-robots-how-ai-is-reshaping-journalism': 'A grounded guide to automated '
                                                            'reporting, fact-checking, newsroom '
                                                            'pressure, platform incentives, and '
                                                            'the trust problem journalism cannot '
                                                            'dodge.',
 'the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media': 'A clear '
                                                                                      'look at '
                                                                                      'recommendations, '
                                                                                      'moderation, '
                                                                                      'privacy, '
                                                                                      'platform '
                                                                                      'incentives, '
                                                                                      'and why '
                                                                                      'your feed '
                                                                                      'knows more '
                                                                                      'than feels '
                                                                                      'comfortable.',
 'game-ai-unleashed-from-finite-state-machines-to-machine-learning': 'A practical guide to NPC '
                                                                     'behaviour, procedural '
                                                                     'systems, player experience, '
                                                                     'and the design choices '
                                                                     'hiding behind clever game '
                                                                     'AI.'}

BOOK_SPECIFIC_DESCRIPTION_OVERRIDES: Dict[str, str] = {'ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology': 'AI is used in '
                                                                               'Formula 1 for race '
                                                                               'strategy, '
                                                                               'simulation, '
                                                                               'telemetry, '
                                                                               'predictive '
                                                                               'analytics, and '
                                                                               'decisions made '
                                                                               'under ridiculous '
                                                                               'time pressure.',
 'artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation': 'AI '
                                                                                                     'is '
                                                                                                     'being '
                                                                                                     'used '
                                                                                                     'in '
                                                                                                     'veterinary '
                                                                                                     'medicine '
                                                                                                     'for '
                                                                                                     'diagnostic '
                                                                                                     'support, '
                                                                                                     'predictive '
                                                                                                     'health '
                                                                                                     'monitoring, '
                                                                                                     'clinic '
                                                                                                     'workflow, '
                                                                                                     'and '
                                                                                                     'treatment '
                                                                                                     'planning.',
 'ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation': 'Artificial '
                                                                                  'intelligence is '
                                                                                  'already being '
                                                                                  'used in smart '
                                                                                  'grids for '
                                                                                  'demand '
                                                                                  'forecasting, '
                                                                                  'grid balancing, '
                                                                                  'fault '
                                                                                  'detection, and '
                                                                                  'renewable '
                                                                                  'integration.',
 'artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability': 'AI is used in '
                                                                                  'logistics for '
                                                                                  'route planning, '
                                                                                  'demand '
                                                                                  'forecasting, '
                                                                                  'warehouse '
                                                                                  'automation, '
                                                                                  'stock movement, '
                                                                                  'and the daily '
                                                                                  'grind of '
                                                                                  'getting things '
                                                                                  'where they need '
                                                                                  'to be.',
 'artificial-intelligence-and-the-law-case-studies-and-future-trends': 'AI is changing legal work '
                                                                       'through research tools, '
                                                                       'contract analysis, case '
                                                                       'review, compliance checks, '
                                                                       'and uncomfortable '
                                                                       'questions about '
                                                                       'accountability.',
 'artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention': 'AI is '
                                                                                           'used '
                                                                                           'in '
                                                                                           'cyber '
                                                                                           'security '
                                                                                           'for '
                                                                                           'threat '
                                                                                           'detection, '
                                                                                           'breach '
                                                                                           'prevention, '
                                                                                           'anomaly '
                                                                                           'spotting, '
                                                                                           'response '
                                                                                           'support, '
                                                                                           'and '
                                                                                           'filtering '
                                                                                           'signal '
                                                                                           'from '
                                                                                           'noise.',
 'the-artificial-intelligence-revolution-from-algorithms-to-consciousness': 'This book traces '
                                                                            'artificial '
                                                                            'intelligence from '
                                                                            'early algorithms to '
                                                                            'modern systems, with '
                                                                            'the big claims tested '
                                                                            'against what the '
                                                                            'technology can '
                                                                            'actually justify.',
 'ai-in-aviation-transforming-safety-and-sustainability': 'AI is used in aviation for predictive '
                                                          'maintenance, air traffic support, fuel '
                                                          'planning, safety monitoring, and the '
                                                          'decisions where failure is not a '
                                                          'charming option.',
 'ai-in-maritime-revolutionizing-shipping-for-sustainability': 'AI is used in maritime shipping '
                                                               'for route planning, autonomous '
                                                               'vessels, port operations, '
                                                               'emissions tracking, and fleet '
                                                               'decisions under real-world '
                                                               'pressure.',
 'artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare': 'AI is used in '
                                                                          'pharmaceuticals for '
                                                                          'drug discovery, trial '
                                                                          'design, patient '
                                                                          'matching, treatment '
                                                                          'personalisation, and '
                                                                          'the regulatory work '
                                                                          'behind medical '
                                                                          'progress.',
 'the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry': 'AI '
                                                                                                'is '
                                                                                                'reshaping '
                                                                                                'the '
                                                                                                'automotive '
                                                                                                'industry '
                                                                                                'through '
                                                                                                'autonomous '
                                                                                                'driving, '
                                                                                                'predictive '
                                                                                                'maintenance, '
                                                                                                'smart '
                                                                                                'manufacturing, '
                                                                                                'mobility '
                                                                                                'services, '
                                                                                                'and '
                                                                                                'liability '
                                                                                                'headaches.',
 'artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future': 'AI '
                                                                                                        'is '
                                                                                                        'used '
                                                                                                        'in '
                                                                                                        'retail '
                                                                                                        'for '
                                                                                                        'recommendations, '
                                                                                                        'stock '
                                                                                                        'planning, '
                                                                                                        'pricing, '
                                                                                                        'supply '
                                                                                                        'chains, '
                                                                                                        'customer '
                                                                                                        'service, '
                                                                                                        'and '
                                                                                                        'the '
                                                                                                        'line '
                                                                                                        'between '
                                                                                                        'useful '
                                                                                                        'and '
                                                                                                        'creepy.',
 'artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery': 'AI '
                                                                                                                'is '
                                                                                                                'used '
                                                                                                                'in '
                                                                                                                'manufacturing '
                                                                                                                'for '
                                                                                                                'predictive '
                                                                                                                'maintenance, '
                                                                                                                'production '
                                                                                                                'planning, '
                                                                                                                'quality '
                                                                                                                'control, '
                                                                                                                'service '
                                                                                                                'delivery, '
                                                                                                                'and '
                                                                                                                'the '
                                                                                                                'operational '
                                                                                                                'details '
                                                                                                                'that '
                                                                                                                'decide '
                                                                                                                'whether '
                                                                                                                'it '
                                                                                                                'pays '
                                                                                                                'off.',
 'artificial-intelligence-in-industry-a-comprehensive-guide': 'AI is used across industry for '
                                                              'automation, analytics, safety, '
                                                              'process control, and decisions that '
                                                              'still need evidence rather than '
                                                              'glossy confidence.',
 'the-dumbening-how-ai-is-reshaping-our-minds': 'This book examines how AI and automation affect '
                                                'attention, memory, creativity, judgement, and the '
                                                'parts of thinking we should be wary of '
                                                'outsourcing.',
 'ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future': 'AI is used in agriculture '
                                                                       'for crop monitoring, '
                                                                       'precision farming, weather '
                                                                       'risk, yield forecasting, '
                                                                       'and decisions made with '
                                                                       'mud on the boots.',
 'ai-in-education-reimagining-learning-for-every-student': 'AI is used in education for '
                                                           'personalised learning, assessment '
                                                           'support, classroom tools, admin '
                                                           'relief, and the question of what '
                                                           'actually helps students learn.',
 'artificial-intelligence-in-banking-revolutionizing-finance-and-data-security': 'AI is used in '
                                                                                 'banking for '
                                                                                 'fraud detection, '
                                                                                 'credit '
                                                                                 'decisions, risk '
                                                                                 'monitoring, '
                                                                                 'compliance, '
                                                                                 'customer '
                                                                                 'service, and the '
                                                                                 'data problems '
                                                                                 'finance cannot '
                                                                                 'shrug off.',
 'digital-diagnosis-how-ai-is-revolutionizing-healthcare': 'AI is used in healthcare for '
                                                           'diagnostics, triage, medical imaging, '
                                                           'workflow support, and clinical '
                                                           'decisions that still demand human '
                                                           'judgement.',
 'artificial-intelligence-in-construction-building-a-sustainable-future': 'AI is used in '
                                                                          'construction for '
                                                                          'planning, safety '
                                                                          'monitoring, materials, '
                                                                          'scheduling, '
                                                                          'sustainability, and the '
                                                                          'messy realities of '
                                                                          'actual building sites.',
 'the-artificial-intelligence-job-shift-navigating-the-future-of-work': 'AI is changing work '
                                                                        'through automation, '
                                                                        'reskilling pressure, job '
                                                                        'redesign, management '
                                                                        'decisions, and the '
                                                                        'awkward question of who '
                                                                        'benefits.',
 'the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information': 'AI is '
                                                                                         'used in '
                                                                                         'government '
                                                                                         'for '
                                                                                         'public '
                                                                                         'services, '
                                                                                         'admin '
                                                                                         'support, '
                                                                                         'security, '
                                                                                         'forecasting, '
                                                                                         'and '
                                                                                         'decisions '
                                                                                         'where '
                                                                                         'accountability '
                                                                                         'cannot '
                                                                                         'be '
                                                                                         'treated '
                                                                                         'as '
                                                                                         'optional.',
 'smart-buildings-ai-powered-efficiency-and-sustainability': 'AI is used in smart buildings for '
                                                             'energy management, maintenance, '
                                                             'sensors, comfort systems, and '
                                                             'infrastructure decisions that still '
                                                             'need oversight.',
 'digital-defense-the-role-of-ai-in-modern-warfare': 'AI is used in modern warfare for drones, '
                                                     'intelligence, cyber operations, targeting '
                                                     'support, and command decisions with '
                                                     'consequences attached.',
 'artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology': 'AI '
                                                                                                                 'is '
                                                                                                                 'used '
                                                                                                                 'in '
                                                                                                                 'wildlife '
                                                                                                                 'conservation '
                                                                                                                 'for '
                                                                                                                 'habitat '
                                                                                                                 'monitoring, '
                                                                                                                 'species '
                                                                                                                 'identification, '
                                                                                                                 'anti-poaching '
                                                                                                                 'work, '
                                                                                                                 'field '
                                                                                                                 'data, '
                                                                                                                 'and '
                                                                                                                 'biodiversity '
                                                                                                                 'protection '
                                                                                                                 'under '
                                                                                                                 'pressure.',
 'climate-intelligence-harnessing-ai-for-a-greener-future': 'AI is used in climate work for '
                                                            'emissions tracking, climate '
                                                            'modelling, renewable planning, '
                                                            'forecasting, and measurement that has '
                                                            'to survive scrutiny.',
 'ai-revolution-in-railways-modernizing-travel-for-a-smarter-future': 'AI is used in railways for '
                                                                      'predictive maintenance, '
                                                                      'signalling, scheduling, '
                                                                      'safety monitoring, and the '
                                                                      'service realities '
                                                                      'passengers notice first.',
 'artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement': 'AI is used '
                                                                                     'in sport for '
                                                                                     'performance '
                                                                                     'analysis, '
                                                                                     'injury '
                                                                                     'prevention, '
                                                                                     'scouting, '
                                                                                     'fan '
                                                                                     'engagement, '
                                                                                     'and '
                                                                                     'decisions '
                                                                                     'where '
                                                                                     'numbers '
                                                                                     'still meet '
                                                                                     'nerves.',
 'lights-camera-algorithm-ai-s-role-in-modern-filmmaking': 'AI is used in filmmaking for script '
                                                           'analysis, editing, visual effects, '
                                                           'production planning, and the craft '
                                                           'decisions algorithms cannot make for '
                                                           'you.',
 'the-ai-music-revolution-creativity-controversy-and-collaboration': 'AI is used in music for '
                                                                     'composition, production, '
                                                                     'mastering, collaboration, '
                                                                     'and rights questions that '
                                                                     'refuse to stay politely in '
                                                                     'the background.',
 'the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead': 'This book follows the people, '
                                                                   'breakthroughs, false starts, '
                                                                   'and hard questions behind '
                                                                   'artificial intelligence.',
 'the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming': 'AI is used in '
                                                                             'gambling for '
                                                                             'personalisation, '
                                                                             'prediction, '
                                                                             'behavioural nudges, '
                                                                             'risk detection, and '
                                                                             'ethical problems '
                                                                             'that deserve more '
                                                                             'than a shrug.',
 'beyond-earth-how-ai-is-transforming-space-exploration': 'AI is used in space exploration for '
                                                          'rovers, mission planning, data '
                                                          'analysis, autonomy, and decisions made '
                                                          'far beyond easy rescue.',
 'from-reporters-to-robots-how-ai-is-reshaping-journalism': 'AI is used in journalism for '
                                                            'automated reporting, fact-checking, '
                                                            'newsroom support, personalisation, '
                                                            'and the trust problem no tool can '
                                                            'hand-wave away.',
 'the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media': 'AI shapes '
                                                                                      'social '
                                                                                      'media '
                                                                                      'through '
                                                                                      'recommendations, '
                                                                                      'moderation, '
                                                                                      'ad '
                                                                                      'targeting, '
                                                                                      'privacy '
                                                                                      'trade-offs, '
                                                                                      'and '
                                                                                      'platform '
                                                                                      'incentives '
                                                                                      'most users '
                                                                                      'never see.',
 'game-ai-unleashed-from-finite-state-machines-to-machine-learning': 'AI is used in games for NPC '
                                                                     'behaviour, procedural '
                                                                     'design, player experience, '
                                                                     'adaptive systems, and the '
                                                                     'craft behind believable '
                                                                     'play.'}

SUMMARY_VARIANT_OPENERS: List[str] = [
    "Instead, it stays with the real sticking points in",
    "It keeps its eye on the pressure points inside",
    "Rather than puffery, it follows the live fault lines in",
    "It sticks to the decisions and bottlenecks shaping",
    "This one stays close to the hard realities inside",
    "It tracks the moments that actually decide whether AI works in",
    "It keeps the spotlight on the practical pressure points across",
    "It follows the places where deployment gets real in",
    "It stays with the awkward but useful details inside",
    "It maps the parts of adoption that actually bite in",
    "It keeps to the practical judgement calls running through",
    "It follows the working pressures that define",
    "It keeps the focus on the stubborn realities inside",
    "It tracks the places where theory collides with operating reality in",
    "It stays with the decisions, constraints, and side-effects shaping",
    "It follows the practical friction points running through",
    "It keeps the argument tied to what really happens inside",
    "It tracks the real-world pressure points shaping",
]

WORKFLOW_VARIANT_OPENERS: List[str] = [
    "The operational detail that matters:",
    "The practical mechanics worth watching:",
    "The decision points that actually count:",
    "The grounded view of deployment comes through:",
    "The useful detail sits here:",
    "The sharper operational lens is on",
    "The reality check comes from",
    "The practical argument turns on",
    "The nuts-and-bolts angle covers",
    "The working detail is in",
    "The less glamorous but more useful part is",
    "The serious implementation view covers",
    "The day-to-day pressure points are",
    "The operational story really sits in",
    "The meaningful detail runs through",
    "The practical reading starts with",
    "The thing worth paying attention to is",
    "The grounded takeaway centres on",
]

SUMMARY_VARIANT_OVERRIDES: Dict[str, int] = {
    "ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology": 4,
    "artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability": 4,
    "the-artificial-intelligence-revolution-from-algorithms-to-consciousness": 12,
    "artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare": 14,
    "artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery": 12,
    "ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future": 4,
    "digital-diagnosis-how-ai-is-revolutionizing-healthcare": 10,
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information": 11,
    "artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology": 13,
    "artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement": 11,
    "the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead": 13,
    "from-reporters-to-robots-how-ai-is-reshaping-journalism": 1,
}

WORKFLOW_VARIANT_OVERRIDES: Dict[str, int] = {
    "ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology": 4,
    "artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation": 7,
    "ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation": 0,
    "artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability": 4,
    "artificial-intelligence-and-the-law-case-studies-and-future-trends": 0,
    "artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention": 9,
    "the-artificial-intelligence-revolution-from-algorithms-to-consciousness": 12,
    "ai-in-aviation-transforming-safety-and-sustainability": 16,
    "ai-in-maritime-revolutionizing-shipping-for-sustainability": 15,
    "artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare": 14,
    "the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry": 11,
    "artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future": 3,
    "artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery": 12,
    "artificial-intelligence-in-industry-a-comprehensive-guide": 12,
    "the-dumbening-how-ai-is-reshaping-our-minds": 6,
    "ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future": 4,
    "ai-in-education-reimagining-learning-for-every-student": 2,
    "artificial-intelligence-in-banking-revolutionizing-finance-and-data-security": 1,
    "digital-diagnosis-how-ai-is-revolutionizing-healthcare": 10,
    "artificial-intelligence-in-construction-building-a-sustainable-future": 10,
    "the-artificial-intelligence-job-shift-navigating-the-future-of-work": 17,
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information": 11,
    "smart-buildings-ai-powered-efficiency-and-sustainability": 10,
    "digital-defense-the-role-of-ai-in-modern-warfare": 6,
    "artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology": 13,
    "climate-intelligence-harnessing-ai-for-a-greener-future": 8,
    "ai-revolution-in-railways-modernizing-travel-for-a-smarter-future": 6,
    "artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement": 11,
    "lights-camera-algorithm-ai-s-role-in-modern-filmmaking": 4,
    "the-ai-music-revolution-creativity-controversy-and-collaboration": 4,
    "the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead": 13,
    "the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming": 5,
    "beyond-earth-how-ai-is-transforming-space-exploration": 11,
    "from-reporters-to-robots-how-ai-is-reshaping-journalism": 1,
    "the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media": 2,
    "game-ai-unleashed-from-finite-state-machines-to-machine-learning": 0,
}


def stable_variant_index(key: str, size: int) -> int:
    if size <= 0:
        raise ValueError("Variant collections must not be empty")
    digest = hashlib.sha256(clean_paragraph(key).encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") % size



def variant_index_for_slug(slug: str, variants: List[str], overrides: Dict[str, int]) -> int:
    override = overrides.get(clean_paragraph(slug))
    if override is not None and 0 <= override < len(variants):
        return override
    return stable_variant_index(slug, len(variants))



def replace_banned_ebook_phrases(value: str, slug: str) -> str:
    cleaned = clean_paragraph(value)
    if not cleaned:
        return cleaned
    learn_tail = BOOK_SPECIFIC_LEARN_TAILS.get(slug)
    advanced_tail = BOOK_SPECIFIC_ADVANCED_TAILS.get(slug)
    if learn_tail:
        cleaned = cleaned.replace("where the claims are running ahead of reality", learn_tail)
        summary_prefix = "This book follows the workflows, trade-offs, and decision points shaping "
        summary_suffix = ", so readers can separate useful systems from glossy nonsense."
        if summary_prefix in cleaned and summary_suffix in cleaned:
            start = cleaned.index(summary_prefix)
            end = cleaned.index(summary_suffix, start) + len(summary_suffix)
            topic_fragment = cleaned[start + len(summary_prefix): cleaned.index(summary_suffix, start)].strip().rstrip(',')
            opener = SUMMARY_VARIANT_OPENERS[variant_index_for_slug(slug, SUMMARY_VARIANT_OPENERS, SUMMARY_VARIANT_OVERRIDES)]
            summary_sentence = f"{opener} {topic_fragment}, especially {learn_tail}."
            cleaned = cleaned[:start] + summary_sentence + cleaned[end:]
    if advanced_tail:
        cleaned = cleaned.replace("trade-offs, and implementation angles", advanced_tail)
        workflow_prefix = "The workflows, systems, and trade-offs behind practical "
        workflow_suffix = " use cases, explained in plain English."
        if cleaned.startswith(workflow_prefix) and cleaned.endswith(workflow_suffix):
            opener = WORKFLOW_VARIANT_OPENERS[variant_index_for_slug(slug, WORKFLOW_VARIANT_OPENERS, WORKFLOW_VARIANT_OVERRIDES)]
            if opener.endswith(":"):
                cleaned = f"{opener} {advanced_tail}."
            else:
                cleaned = f"{opener} {advanced_tail}."
    return cleaned

def catalogue_intro_copy(topic: str) -> str:
    return CATALOGUE_INTRO_VARIANTS.get(topic, f"These books stay practical about {topic.lower()}, with less hype and more on how the work actually gets done.")

def catalogue_cta_copy(topic: str) -> str:
    return CATALOGUE_CTA_VARIANTS.get(topic, "Read the detailed page, then see the current Amazon listing when you are ready.")



def default_why_it_matters(topic: str) -> str:
    topic_lc = topic.lower()
    return f"Because AI in {topic_lc} changes decisions, workflows, and risk. Getting the basics right matters before the hype machine starts throwing confetti."


TOPIC_FAMILY_GROUPS = {
    "regulated": {"Healthcare", "Law", "Finance", "Government", "Education", "Defence"},
    "operations": {"Manufacturing", "Industry", "Transportation", "Construction", "Energy", "Agriculture", "Retail"},
    "security": {"Cyber Security"},
    "creative": {"Creativity", "Media", "Gaming"},
    "foundation": {"Artificial Intelligence", "Ethics", "History", "Science", "Future of Work"},
    "environment": {"Environment"},
    "sports": {"Sports"},
}


def topic_family(topic: str) -> str:
    topic_name = clean_paragraph(topic)
    for family, topics in TOPIC_FAMILY_GROUPS.items():
        if topic_name in topics:
            return family
    return "general"


TOPIC_GUIDE_DIRECTORY = {
    "ai-for-beginners": {
        "title": "AI for Beginners",
        "summary": "Plain-English foundations covering how AI works, where it is used, and what to pay attention to before the hype swallows the basics.",
    },
    "ai-in-business": {
        "title": "AI in Business",
        "summary": "Strategy, adoption, risk, and practical decisions for organisations implementing AI in the real world rather than in a slide deck.",
    },
    "ai-in-healthcare": {
        "title": "AI in Healthcare",
        "summary": "Clinical AI, diagnostics, NHS deployment pressures, and the regulatory questions that decide whether the tools help or hinder care.",
    },
    "ai-ethics": {
        "title": "AI Ethics",
        "summary": "Bias, accountability, transparency, safety, and the governance questions that matter when AI decisions affect real people.",
    },
    "ai-in-education": {
        "title": "AI in Education",
        "summary": "Personalised learning, classroom tools, assessment, and what the evidence actually says once the marketing smoke clears.",
    },
    "ai-in-finance": {
        "title": "AI in Finance",
        "summary": "Fraud detection, credit, trading, regulation, and the awkward trade-off between speed, risk, and accountability.",
    },
    "deep-learning": {
        "title": "Deep Learning",
        "summary": "Neural networks, training, pattern recognition, and where the models earn their reputation versus where they still fail noisily.",
    },
    "generative-ai": {
        "title": "Generative AI",
        "summary": "LLMs, image generation, content automation, and the difference between a useful tool and an expensive hallucination machine.",
    },
    "machine-learning": {
        "title": "Machine Learning",
        "summary": "Core machine-learning ideas, real-world use cases, and the data-quality problems that decide whether a model is useful or decorative.",
    },
    "robotics-automation": {
        "title": "Robotics & Automation",
        "summary": "Industrial robotics, autonomous systems, and the point where AI leaves the screen and starts affecting physical operations.",
    },
}

CATEGORY_TO_TOPIC_GUIDE = {
    "Agriculture": "machine-learning",
    "Artificial Intelligence": "ai-for-beginners",
    "Construction": "robotics-automation",
    "Creativity": "generative-ai",
    "Cyber Security": "machine-learning",
    "Defence": "ai-ethics",
    "Education": "ai-in-education",
    "Energy": "machine-learning",
    "Environment": "machine-learning",
    "Ethics": "ai-ethics",
    "Finance": "ai-in-finance",
    "Future of Work": "ai-in-business",
    "Gaming": "generative-ai",
    "Government": "ai-in-business",
    "Healthcare": "ai-in-healthcare",
    "History": "ai-for-beginners",
    "Industry": "ai-in-business",
    "Law": "ai-ethics",
    "Manufacturing": "robotics-automation",
    "Media": "generative-ai",
    "Retail": "ai-in-business",
    "Science": "deep-learning",
    "Sports": "machine-learning",
    "Transportation": "robotics-automation",
}


def topic_guide_record(slug: str | None) -> Dict[str, str] | None:
    if not slug:
        return None
    return TOPIC_GUIDE_DIRECTORY.get(slug)


def category_question_heading(topic: str) -> str:
    topic_name = clean_paragraph(topic)
    if topic_name.lower() == "artificial intelligence":
        return "What does this category actually cover?"
    return f"What is AI in {topic_name}?"


def category_answer_first_copy(topic: str, books: List[Dict[str, Any]]) -> str:
    topic_name = clean_paragraph(topic)
    intro = topic_intro(topic_name)
    featured = books[0]
    if len(books) == 1:
        return f"{intro} Start with {featured['title']} if you want one grounded route into the main use cases, trade-offs, and implementation questions."
    return f"{intro} This category brings together {len(books)} books, so you can move from the broad question to the title that best matches your use case."


def category_scope_copy(topic: str, books: List[Dict[str, Any]]) -> str:
    topic_name = clean_paragraph(topic)
    family = topic_family(topic_name)
    if len(books) == 1:
        return {
            "regulated": f"The {topic_name.lower()} category focuses on evidence, adoption pressure, oversight, and the point where AI convenience collides with accountability. It is useful when you need more than a glossy vendor promise.",
            "operations": f"The {topic_name.lower()} category focuses on workflow, reliability, cost, and what happens once AI has to survive contact with real operations instead of a keynote stage.",
            "security": f"The {topic_name.lower()} category focuses on signal, attacker behaviour, operational noise, and where automation genuinely improves defence rather than just moving the mess around.",
            "creative": f"The {topic_name.lower()} category focuses on speed, craft, control, and the rights-shaped complications that appear the moment AI output starts looking commercially useful.",
            "foundation": f"The {topic_name.lower()} category focuses on first principles, practical claims, sharper judgement, and the context readers need before treating broad AI arguments as settled fact.",
            "environment": f"The {topic_name.lower()} category focuses on measurable environmental outcomes, data quality, and the difference between genuine optimisation and eco-flavoured marketing copy.",
            "sports": f"The {topic_name.lower()} category focuses on performance analysis, decision-making, and the stubborn fact that data still has to coexist with human judgement.",
        }.get(family, f"The {topic_name.lower()} category focuses on practical use cases, trade-offs, and the questions worth asking before anyone treats AI as a magic trick.")
    titles = ", ".join(book["title"] for book in books[:2])
    if len(books) > 2:
        titles += ", and more"
    return f"This category spans {len(books)} titles, starting with {titles}. Together they cover the main use cases, trade-offs, and adjacent decisions readers normally need before choosing a narrower book."


def category_best_start_copy(topic: str, books: List[Dict[str, Any]]) -> str:
    featured = books[0]
    if len(books) == 1:
        return f"If you want one grounded entry point, start with <a href='/ebooks/{html.escape(featured['slug'])}/'>{html.escape(featured['title'])}</a>. It is the clearest way into this topic without having to untangle three tabs, two buzzword decks, and someone's suspiciously cheerful vendor PDF."
    return f"Start with <a href='/ebooks/{html.escape(featured['slug'])}/'>{html.escape(featured['title'])}</a>, then use the related titles below to go narrower once you know which part of the subject matters most to you."


def category_faq_markup(topic: str, books: List[Dict[str, Any]]) -> str:
    topic_name = clean_paragraph(topic)
    family = topic_family(topic_name)
    questions = [
        (
            f"What does the {topic_name} category help me understand?",
            category_scope_copy(topic_name, books),
        ),
        (
            "Who should start here?",
            f"Readers who want a grounded overview of {topic_name.lower()} before picking a specific title, plus professionals who need a fast way to identify the book most relevant to their role.",
        ),
        (
            "Where should I go next after the featured title?",
            {
                "regulated": "Move into the glossary for key terms, then use the comparison page to pressure-test claims, risks, and implementation trade-offs across sectors.",
                "operations": "Move into the glossary for core terms, then use the comparison page and related topic guide to compare workflow gains against real-world constraints.",
                "security": "Move into the glossary for technical language, then use the comparison page and topic guide to keep the tooling anchored to risk rather than hype.",
                "creative": "Move into the glossary for the core concepts, then use the comparison page and topic guide to separate useful creative acceleration from rights and trust problems.",
                "foundation": "Move into the glossary for definitions, then use the comparison page and topic guide to connect the broad theory with specific use cases and trade-offs.",
                "environment": "Move into the glossary for the key concepts, then use the comparison page and topic guide to separate measurable gains from greenwashed claims.",
                "sports": "Move into the glossary for the core terms, then use the comparison page and topic guide to connect performance data with real-world judgement.",
            }.get(family, "Move into the glossary, comparison page, and related topic guide to connect this category with the wider AI estate."),
        ),
    ]
    parts = []
    for idx, (question, answer) in enumerate(questions, start=1):
        open_attr = " open" if idx == 1 else ""
        parts.append(
            f'<details class="ebook-faq-item"{open_attr}><summary>{html.escape(question)}</summary><div><p>{html.escape(answer)}</p></div></details>'
        )
    return "\n".join(parts)


def category_reading_path(books: List[Dict[str, Any]]) -> tuple[str, str] | None:
    """Return the first governed reading path that genuinely contains a category title."""
    try:
        payload = read_json(DATA_DIR / "ebook-bundles.json", default={}) or {}
        bundles = payload.get("bundles", []) if isinstance(payload, dict) else []
    except Exception:
        bundles = []
    category_slugs = {clean_paragraph(book.get("slug")) for book in books}
    for bundle in bundles if isinstance(bundles, list) else []:
        if not isinstance(bundle, dict):
            continue
        members = {clean_paragraph(slug) for slug in bundle.get("books", [])}
        if not category_slugs.intersection(members):
            continue
        slug = clean_paragraph(bundle.get("slug"))
        title = clean_paragraph(bundle.get("title"))
        if slug and title:
            return slug, title
    return None


def category_internal_links(topic: str, books: List[Dict[str, Any]]) -> str:
    topic_name = clean_paragraph(topic)
    family = topic_family(topic_name)
    featured = books[0]
    guide_slug = CATEGORY_TO_TOPIC_GUIDE.get(topic_name)
    guide = topic_guide_record(guide_slug)
    compare_anchor = {
        "regulated": "Compare AI risks, evidence, and trust trade-offs",
        "operations": "Compare AI use cases across operational sectors",
        "security": "Compare AI applications, risks, and adoption patterns",
        "creative": "Compare generative AI use cases and trade-offs",
        "foundation": "Compare core AI concepts, claims, and use cases",
        "environment": "Compare AI use cases, risks, and measurable outcomes",
        "sports": "Compare AI use cases, judgement calls, and trade-offs",
    }.get(family, "Compare AI use cases and trade-offs")
    links = [
        (
            f"/ebooks/{featured['slug']}/",
            f"Start with {featured['title']}",
            "Featured book page with the full summary, FAQ, and buy route.",
        ),
        (
            "/glossary/",
            f"Use the AI glossary to decode {topic_name.lower()} terms",
            "Definitions and plain-English explanations for the jargon this category keeps bumping into.",
        ),
        (
            "/compare/",
            compare_anchor,
            "Side-by-side context for how different AI areas solve different problems and create different headaches.",
        ),
    ]
    if guide:
        links.append(
            (
                f"/topics/{guide_slug}/",
                f"Read the {guide['title']} guide",
                guide["summary"],
            )
        )
    reading_path = category_reading_path(books)
    if reading_path:
        path_slug, path_title = reading_path
        links.append(
            (
                f"/bundles/{path_slug}/",
                f"Reading path: {path_title}",
                "A curated multi-book route for this topic; titles are purchased separately on Amazon.",
            )
        )
    links.append(
        (
            f"/podcast/?topic={slugify(topic_name)}",
            f"Listen next: Turing’s Torch on {topic_name.lower()}",
            "Editorial context and broader AI commentary beyond the catalogue pages.",
        )
    )
    links.append(
        (
            "/newsletter/",
            "Get the newsletter for ongoing AI developments",
            "Useful if you want the moving story, not just the evergreen version.",
        )
    )
    return "\n".join(
        f'<li><a href="{html.escape(href)}">{html.escape(label)}</a><span>{html.escape(description)}</span></li>'
        for href, label, description in links[:6]
    )


def topic_guide_cards_markup() -> str:
    cards = []
    for slug, payload in TOPIC_GUIDE_DIRECTORY.items():
        cards.append(
            f'<article class="card topic-card topic-card--guide"><h2><a href="/topics/{slug}/">{html.escape(payload["title"])}</a></h2><p>{html.escape(payload["summary"])}</p></article>'
        )
    return "\n".join(cards)


def topics_index_support_links() -> str:
    links = [
        ("/glossary/", "Use the AI glossary for key terms", "Start here if the language gets dense or suspiciously overconfident."),
        ("/compare/", "Compare AI use cases and trade-offs", "Useful for seeing how different AI categories solve different problems."),
        ("/podcast/", "Listen to the podcast for wider AI context", "Weekly editorial context for the bigger shifts around the catalogue."),
        ("/newsletter/", "Join the newsletter for ongoing AI developments", "The moving story, minus the usual noise and confetti."),
    ]
    return "\n".join(
        f'<li><a href="{html.escape(href)}">{html.escape(label)}</a><span>{html.escape(description)}</span></li>'
        for href, label, description in links
    )


def strip_terminal_stop(value: str) -> str:
    return clean_paragraph(value).rstrip(".!?")


def audience_core_for_practice(value: str) -> str:
    cleaned = strip_terminal_stop(value)
    patterns = [
        r"^A practical overview of .+? for ",
        r"^A clear, no-hype briefing on .+? for ",
        r"^Readers who want ",
    ]
    if re.match(patterns[0], cleaned, flags=re.I):
        return re.sub(patterns[0], "", cleaned, flags=re.I)
    if re.match(patterns[1], cleaned, flags=re.I):
        return re.sub(patterns[1], "", cleaned, flags=re.I)
    if re.match(patterns[2], cleaned, flags=re.I):
        return "readers who want " + re.sub(patterns[2], "", cleaned, flags=re.I)
    return cleaned[:1].lower() + cleaned[1:] if cleaned else "readers who want a grounded overview"


def book_unique_evidence_passage(book: Dict[str, Any]) -> str:
    topic_lc = clean_paragraph(book.get("topic", "")).lower() or "the field"
    learn_line = strip_terminal_stop(book["what_youll_learn"][0] if book.get("what_youll_learn") else "")
    family = topic_family(book.get("topic", ""))
    tradeoff = {
        "regulated": f"the awkward point where speed, evidence, and accountability stop pretending to be friends in {topic_lc}",
        "operations": "promised efficiency has to survive contact with maintenance logs, handovers, outages, and real operating constraints",
        "security": f"stronger detection in {topic_lc} has to avoid burying teams under a fresh layer of operational noise",
        "creative": f"convenience in {topic_lc} starts dragging ownership, trust, and control questions in behind it",
        "foundation": f"broad claims in {topic_lc} meet what the systems can actually justify",
        "environment": f"environmental promise in {topic_lc} has to be measured without flattering the numbers",
        "sports": f"data-led gains in {topic_lc} meet the human judgement that still decides outcomes",
    }.get(family, f"impressive claims in {topic_lc} meet what the work actually demands")
    audience = audience_core_for_practice(book.get("audience", ""))
    if learn_line:
        learning_sentence = f"It tackles {learn_line[:1].lower() + learn_line[1:]}"
    else:
        learning_sentence = "It keeps the focus on the decisions that matter once AI leaves the demo stage"
    return (
        f"This book is useful when {tradeoff}. "
        f"It is written for {audience}. "
        f"{learning_sentence}."
    )


def book_semantic_journey_links(book: Dict[str, Any]) -> str:
    guide_slug = CATEGORY_TO_TOPIC_GUIDE.get(clean_paragraph(book.get("topic", "")))
    guide = topic_guide_record(guide_slug)
    links = [
        ("/ebooks/", "Browse all books"),
        (book.get("topic_url") or "/topics/", f'Browse {clean_paragraph(book.get("topic", "AI"))} books'),
        ("/glossary/", "Glossary"),
        ("/compare/", "Comparisons"),
    ]
    if guide:
        links.append((f"/topics/{guide_slug}/", guide["title"] + " guide"))
    links.extend([
        ("/podcast/", "Podcast"),
        ("/newsletter/", "Newsletter"),
    ])
    return "\n".join(f'<a href="{html.escape(href)}">{html.escape(label)}</a>' for href, label in links[:6])


def showcase_subhead(book: Dict[str, Any]) -> str:
    family = topic_family(book.get("topic", ""))
    topic_lc = clean_paragraph(book.get("topic", "")).lower() or "the field"
    return {
        "regulated": f"From real deployment in {topic_lc} to evidence, oversight, and real-world consequence.",
        "operations": f"From day-to-day work in {topic_lc} to gains, failure modes, and trade-offs.",
        "security": f"From active defence in {topic_lc} to attacker adaptation, blind spots, and operational noise.",
        "creative": f"From creative speed in {topic_lc} to ownership, control, and the compromises buried inside convenience.",
        "foundation": f"From first principles in {topic_lc} to practical claims, limitations, and sharper judgement.",
        "environment": f"From environmental promise in {topic_lc} to measurable outcomes, constraints, and trade-offs.",
        "sports": f"From performance decisions in {topic_lc} to human judgement, edge cases, and competitive trade-offs.",
    }.get(family, f"From practical deployment in {topic_lc} to trade-offs, judgement, and real-world constraints.")


def showcase_note(book: Dict[str, Any]) -> str:
    family = topic_family(book.get("topic", ""))
    topic_lc = clean_paragraph(book.get("topic", "")).lower() or "the field"
    return {
        "regulated": f"Built for readers who need {topic_lc} explained as a real operating environment, not a compliance-free demo.",
        "operations": f"Built for people who care whether AI in {topic_lc} survives contact with the workflow rather than just the keynote.",
        "security": f"Built for readers who know the tooling only matters if it improves the signal without burying the team in noise.",
        "creative": f"Built for people who want the upside in {topic_lc} without politely ignoring the rights, trust, and control problems.",
        "foundation": f"Built for readers who want the claims around {topic_lc} translated, tested, and relieved of their marketing costume.",
        "environment": f"Built for readers who want environmental realism in {topic_lc}, not green-tinted dashboards and applause.",
        "sports": f"Built for people who know the numbers can help in {topic_lc}, but not more than the humans making the call.",
    }.get(family, f"Built for readers who want practical judgement in {topic_lc} rather than brochure copy.")


def practical_outcomes_intro(book: Dict[str, Any]) -> str:
    family = topic_family(book.get("topic", ""))
    topic_lc = clean_paragraph(book.get("topic", "")).lower() or "the field"
    return {
        "regulated": f"You should finish it better able to separate usable AI in {topic_lc} from risky shortcuts, loose governance, and expensive confidence.",
        "operations": f"You should finish it with a clearer feel for where AI in {topic_lc} improves the workflow, where it adds fragility, and what to pilot before anyone starts chest-thumping.",
        "security": f"You should finish it better at separating useful automation in {topic_lc} from noisy promises and more alert to where attackers or blind spots creep in.",
        "creative": f"You should finish it with a sharper sense of where AI in {topic_lc} genuinely helps the work and where it starts borrowing tomorrow's headache.",
        "foundation": f"You should finish it with the jargon around {topic_lc} translated, the stronger claims stress-tested, and a better map of where to dig deeper.",
        "environment": f"You should finish it better able to tell the difference between measurable gains in {topic_lc}, modelling optimism, and plain old green lipstick on a dashboard.",
        "sports": f"You should finish it able to judge which parts of {topic_lc} belong to data, which still belong to people, and where the line keeps moving.",
    }.get(family, f"You should leave this one with clearer judgement about {topic_lc}, fewer lazy assumptions, and a better sense of where to press further or walk away.")


def default_distinct_angle(title: str, topic: str) -> str:
    family = topic_family(topic)
    topic_lc = topic.lower()
    return {
        "regulated": f"{title} keeps its eye on evidence, accountability, and the point where a slick demo meets real-world responsibility in {topic_lc}.",
        "operations": f"{title} keeps its boots on the ground, looking at workflow, failure modes, and whether the gains survive contact with real operations in {topic_lc}.",
        "security": f"{title} keeps the focus on signal, defence, and the cost of getting the call wrong in {topic_lc}.",
        "creative": f"{title} keeps one eye on craft and the other on ownership, control, and the compromises hiding inside convenience in {topic_lc}.",
        "foundation": f"{title} keeps the applause to a minimum and asks what the systems actually do, what they break, and what they are being oversold to solve in {topic_lc}.",
        "environment": f"{title} keeps the focus on measurable environmental value in {topic_lc} rather than eco-flavoured marketing copy.",
        "sports": f"{title} keeps the focus on performance, judgement, and how data changes decisions in {topic_lc} without pretending sport becomes an equation.",
    }.get(family, f"{title} keeps the focus on practical judgement in {topic_lc} rather than drifting into brochure-speak.")


WORKBOOK_GOVERNED_COPY_FIELDS = (
    ("title", "title"),
    ("short", "short"),
    ("short", "short_description"),
    ("description", "description"),
    ("summary", "summary"),
    ("topic", "topic"),
    ("audience", "audience"),
    ("who_for", "who_for"),
    ("what_this_book_covers", "what_this_book_covers"),
    ("why_it_matters", "why_it_matters"),
)



def default_short(topic: str, pages: int | None) -> str:
    return helper_default_short(topic, pages)


SEO_TITLE_OVERRIDES = {
    "ai-and-formula-1-redefining-speed-and-strategy-with-intelligent-technology": "AI and Formula 1",
    "ai-powered-smart-grid-revolutionizing-electricity-distribution-and-generation": "AI-Powered Smart Grid",
    "artificial-intelligence-and-the-law-case-studies-and-future-trends": "AI and the Law",
    "artificial-intelligence-for-cyber-security-a-practical-guide-to-data-breach-prevention": "AI for Cyber Security",
    "artificial-intelligence-for-wildlife-conservation-revolutionizing-biodiversity-protection-through-technology": "AI for Wildlife Conservation",
    "artificial-intelligence-in-banking-revolutionizing-finance-and-data-security": "AI in Banking",
    "artificial-intelligence-in-logistics-optimizing-efficiency-and-sustainability": "AI in Logistics",
    "artificial-intelligence-in-sports-revolutionizing-performance-and-fan-engagement": "AI in Sports",
    "artificial-intelligence-in-veterinary-medicine-transforming-animal-healthcare-through-innovation": "AI in Veterinary Medicine",
    "artificial-intelligence-powered-retail-revolutionizing-customer-experience-for-a-sustainable-future": "AI-Powered Retail",
    "artificial-intelligence-revolution-in-manufacturing-modernizing-operations-maintenance-and-service-delivery": "AI Revolution in Manufacturing",
    "the-ai-behind-your-feed-personalization-moderation-and-the-future-of-social-media": "The AI Behind Your Feed",
    "the-autonomous-revolution-artificial-intelligence-and-the-future-of-the-automotive-industry": "The Autonomous Revolution",
    "the-future-of-government-leveraging-ai-to-enhance-services-and-safeguard-information": "The Future of Government and AI",
    "the-house-always-knows-ai-gambling-and-the-ethics-of-personalized-gaming": "The House Always Knows",
    "ai-in-aviation-transforming-safety-and-sustainability": "AI in Aviation",
    "ai-in-maritime-revolutionizing-shipping-for-sustainability": "AI in Maritime",
    "ai-in-agriculture-revolutionizing-farming-for-a-sustainable-future": "AI in Agriculture",
    "ai-in-education-reimagining-learning-for-every-student": "AI in Education: Reimagining Learning for Every Student",
    "ai-revolution-in-railways-modernizing-travel-for-a-smarter-future": "AI Revolution in Railways",
    "artificial-intelligence-in-construction-building-a-sustainable-future": "AI in Construction",
    "artificial-intelligence-in-industry-a-comprehensive-guide": "AI in Industry",
    "artificial-intelligence-in-pharmaceuticals-revolutionizing-healthcare": "AI in Pharmaceuticals",
    "beyond-earth-how-ai-is-transforming-space-exploration": "Beyond Earth: AI in Space",
    "climate-intelligence-harnessing-ai-for-a-greener-future": "Climate Intelligence",
    "digital-defense-the-role-of-ai-in-modern-warfare": "Digital Defense and AI",
    "digital-diagnosis-how-ai-is-revolutionizing-healthcare": "Digital Diagnosis",
    "from-reporters-to-robots-how-ai-is-reshaping-journalism": "From Reporters to Robots",
    "game-ai-unleashed-from-finite-state-machines-to-machine-learning": "Game AI Unleashed",
    "lights-camera-algorithm-ai-s-role-in-modern-filmmaking": "Lights, Camera, Algorithm",
    "smart-buildings-ai-powered-efficiency-and-sustainability": "Smart Buildings and AI",
    "the-ai-music-revolution-creativity-controversy-and-collaboration": "The AI Music Revolution",
    "the-architects-of-ai-pioneers-breakthroughs-and-the-road-ahead": "The Architects of AI",
    "the-artificial-intelligence-job-shift-navigating-the-future-of-work": "The AI Job Shift",
    "the-artificial-intelligence-revolution-from-algorithms-to-consciousness": "The AI Revolution",
}


def book_meta_title(book: Dict[str, Any]) -> str:
    slug = clean_paragraph(book.get("slug", ""))
    if slug in SEO_TITLE_OVERRIDES:
        return SEO_TITLE_OVERRIDES[slug]
    return clean_paragraph(book.get("title", ""))



def book_meta_description(book: Dict[str, Any]) -> str:
    primary = clean_paragraph(book.get("description", ""))
    if primary and len(primary) <= 155:
        return primary
    fallback = clean_paragraph(book.get("short") or book.get("short_description") or book.get("summary") or primary)
    return fallback or primary



def catalogue_meta_description(topic: str) -> str:
    topic_name = clean_paragraph(topic).lower() or "artificial intelligence"
    return f"Browse Jonathan Harris AI books on {topic_name} with plain-English guides, practical analysis, clear summaries, and direct Amazon routes."



def topics_index_meta_description() -> str:
    return "Explore AI topics across Jonathan Harris's ebook library, with practical guides, grounded summaries, and direct routes into each catalogue page."



def extract_html_title(text: str) -> str:
    match = re.search(r"<title>(.*?)</title>", text, flags=re.S | re.I)
    return html.unescape(clean_paragraph(match.group(1))) if match else ""



def extract_meta_description(text: str) -> str:
    for meta_tag in re.findall(r"<meta\b[^>]*>", text, flags=re.I | re.S):
        name_match = re.search(r'\bname=(["\'])(.*?)\1', meta_tag, flags=re.I | re.S)
        if not name_match or name_match.group(2).strip().lower() != "description":
            continue
        content_match = re.search(r'\bcontent=(["\'])(.*?)\1', meta_tag, flags=re.I | re.S)
        if content_match:
            return html.unescape(clean_paragraph(content_match.group(2)))
    return ""

def extract_img_src(tag: str) -> str:
    match = re.search(r'\bsrc="([^"]+)"', tag, re.I)
    return clean_paragraph(match.group(1)) if match else ""


def is_remote_image_src(src: str) -> bool:
    cleaned = clean_paragraph(src)
    return cleaned.startswith(("http://", "https://"))


def metadata_budget_errors(
    label: str,
    text: str,
    *,
    max_title: int | None = None,
    min_description: int | None = None,
    max_description: int | None = None,
) -> List[str]:
    errors: List[str] = []
    title = extract_html_title(text)
    description = extract_meta_description(text)
    if not title:
        errors.append(f"{label} is missing a title tag.")
    elif max_title is not None and len(title) > max_title:
        errors.append(f"{label} title tag exceeds {max_title} characters ({len(title)}).")
    if not description:
        errors.append(f"{label} is missing a meta description.")
    else:
        if min_description is not None and len(description) < min_description:
            errors.append(f"{label} meta description is shorter than {min_description} characters ({len(description)}).")
        if max_description is not None and len(description) > max_description:
            errors.append(f"{label} meta description exceeds {max_description} characters ({len(description)}).")
    return errors



def build_default_faq(book: Dict[str, Any]) -> List[Dict[str, Any]]:
    learn_line = book["what_youll_learn"][0] if book.get("what_youll_learn") else topic_intro(book["topic"])
    topic_name = clean_paragraph(book.get("topic", "Artificial Intelligence"))
    topic_lc = topic_name.lower()
    topic_phrase = "artificial intelligence" if topic_name.lower() == "artificial intelligence" else f"AI in {topic_lc}"
    return [
        {
            "@type": "Question",
            "name": f"What does this book explain about {topic_phrase}?",
            "acceptedAnswer": {"@type": "Answer", "text": learn_line},
        },
        {
            "@type": "Question",
            "name": f"Who gets the most value from this {topic_lc} guide?",
            "acceptedAnswer": {"@type": "Answer", "text": audience_faq_answer(book["audience"], book["topic"])},
        },
        {
            "@type": "Question",
            "name": "How detailed is the coverage?",
            "acceptedAnswer": {
                "@type": "Answer",
                "text": f"It runs to {book['pages']} pages and focuses on {clean_paragraph(book.get('what_this_book_covers', '')).rstrip('.')}.",
            },
        },
        {
            "@type": "Question",
            "name": "Where can I get the eBook?",
            "acceptedAnswer": {"@type": "Answer", "text": "Available as an eBook via Amazon using the buy link on this page."},
        },
    ]


def build_master_from_workbook(workbook_path: Path) -> List[Dict[str, Any]]:
    order, workbook_map, workbook_content = parse_workbook(workbook_path, sanitise_content=False)
    if not order:
        raise ValueError("No ebook rows found in workbook")

    required_master_fields = [
        "title", "short", "description", "summary", "topic", "tags", "keywords", "audience",
        "who_for", "what_this_book_covers", "what_youll_learn", "why_it_matters", "tone", "author",
        "identifier", "asin", "pages", "datePublished", "book_url", "buy_route_full", "buy_url",
        "cover", "legacy_alias_url",
    ]

    build_timestamp = utc_now()
    records: List[Dict[str, Any]] = []
    missing_fields: List[str] = []
    for idx, slug in enumerate(order, start=1):
        workbook = workbook_map[slug]
        content = workbook_content.get(slug, {})

        for field in required_master_fields:
            value = workbook.get(field) if field in workbook else content.get(field)
            if isinstance(value, list):
                has_value = any(clean_paragraph(item) for item in value)
            else:
                has_value = bool(clean_paragraph(value))
            if not has_value:
                missing_fields.append(f"{slug}: missing canonical workbook field '{field}'")

        title = clean_paragraph(content.get("title")) or humanise_slug(slug)
        topic = clean_paragraph(content.get("topic")) or "Artificial Intelligence"
        tags = unique_list(content.get("tags") or [topic, "Artificial Intelligence", "AI Trends"])
        keywords = unique_list(content.get("keywords") or [topic, title, *tags])
        pages = workbook.get("pages")
        summary_seed = clean_paragraph(content.get("summary")) or clean_paragraph(content.get("description")) or topic_intro(topic)
        description_seed = clean_paragraph(content.get("description")) or summary_seed
        summary = strip_pages_from_summary(summary_seed, pages)
        description = clean_paragraph(description_seed)
        what_this_book_covers = clean_paragraph(content.get("what_this_book_covers")) or summary
        audience = clean_paragraph(content.get("audience")) or clean_paragraph(content.get("who_for")) or DEFAULT_AUDIENCE
        who_for = clean_paragraph(content.get("who_for")) or audience
        what_youll_learn = unique_list(content.get("what_youll_learn") or default_learning_points(topic))
        why_it_matters = clean_paragraph(content.get("why_it_matters")) or default_why_it_matters(topic)
        short = strip_pages_from_summary(content.get("short"), pages) or default_short(topic, pages)
        short = BOOK_SPECIFIC_SHORT_OVERRIDES.get(slug, short)
        description_override = BOOK_SPECIFIC_DESCRIPTION_OVERRIDES.get(slug)
        if description_override:
            previous_description = description
            description = description_override
            if previous_description and summary.startswith(previous_description):
                summary = description + summary[len(previous_description):]
            elif description not in summary:
                summary = description
        canonical_url = ensure_trailing_slash(workbook.get("book_url") or f"{SITE_URL}/ebooks/{slug}/")
        topic_slug = slugify(topic)
        topic_url = f"/catalogue/{topic_slug}/"
        identifier = clean_paragraph(content.get("identifier")) or f"JH-AI-EBOOK-{idx:02d}"
        author = clean_paragraph(content.get("author")) or SITE_NAME
        tone = clean_paragraph(content.get("tone")) or DEFAULT_TONE
        cover = workbook.get("cover") or clean_paragraph(content.get("cover") or content.get("image") or "")

        book: Dict[str, Any] = {
            "id": idx,
            "key": f"{idx}-ebook",
            "slug": slug,
            "title": title,
            "short": short,
            "short_description": short,
            "description": description,
            "summary": summary,
            "topic": topic,
            "topic_slug": topic_slug,
            "topic_url": topic_url,
            "filter": topic,
            "tags": tags,
            "keywords": keywords,
            "cover": cover,
            "main_image": cover,
            "image": cover,
            "buy_url": workbook.get("buy_url", ""),
            "buy_route": workbook.get("buy_route") or f"/ebooks/{slug}/buy-now",
            "buy_route_full": workbook.get("buy_route_full") or f"{SITE_URL}/ebooks/{slug}/buy-now",
            "canonical_url": canonical_url,
            "book_url": canonical_url,
            "legacy_alias_url": workbook.get("legacy_alias_url", ""),
            "pages": pages,
            "asin": workbook.get("asin", ""),
            "datePublished": workbook.get("datePublished", ""),
            "dateModified": "",
            "author": author,
            "tone": tone,
            "audience": audience,
            "identifier": identifier,
            "what_this_book_covers": what_this_book_covers,
            "who_for": who_for,
            "what_youll_learn": what_youll_learn,
            "why_it_matters": why_it_matters,
            "showcase_heading": clean_paragraph(content.get("showcase_heading")) or (f"How AI is reshaping {topic.lower()}" if topic.lower() != "artificial intelligence" else "AI without the carnival barker routine"),
            "distinct_angle": clean_paragraph(content.get("distinct_angle")) or default_distinct_angle(title, topic),
            "notes": workbook.get("notes", ""),
        }
        book = sanitise_record_copy(book)
        faq_payload = content.get("faq")
        book["faq"] = faq_payload if isinstance(faq_payload, list) else build_default_faq(book)
        book = sanitise_record_copy(book)
        records.append(book)

    if missing_fields:
        raise ValueError("Canonical workbook is missing required live fields:\n- " + "\n- ".join(missing_fields))

    role_errors = content_role_validation_errors(records)
    if role_errors:
        raise ValueError("Canonical workbook content roles are collapsing:\n- " + "\n- ".join(role_errors))

    add_related_books(records)
    existing_master = {
        clean_paragraph(item.get("slug")): item
        for item in (read_json(MASTER_PATH, default=[]) or [])
        if isinstance(item, dict) and clean_paragraph(item.get("slug"))
    }
    for record in records:
        existing = existing_master.get(record["slug"])
        if existing and {k: v for k, v in existing.items() if k != "dateModified"} == {k: v for k, v in record.items() if k != "dateModified"}:
            record["dateModified"] = clean_paragraph(existing.get("dateModified")) or build_timestamp
        else:
            record["dateModified"] = build_timestamp
        published = clean_paragraph(record.get("datePublished"))
        modified = clean_paragraph(record.get("dateModified"))
        # Structured-data chronology is a hard invariant. If the governed source has
        # no trustworthy later modification date, publication date is the minimum
        # valid fallback rather than an invented timestamp.
        if published and modified and modified[:10] < published[:10]:
            record["dateModified"] = published
    return records


def add_related_books(records: List[Dict[str, Any]]) -> None:
    token_map = {record["slug"]: book_token_groups(record) for record in records}

    def sort_key(current: Dict[str, Any], candidate: Dict[str, Any]) -> tuple[int, int, int, int, int, str]:
        score, tie_breaker = related_book_score(candidate, current, token_map[current["slug"]], token_map[candidate["slug"]])
        same_topic, title_overlap, grouped_overlap, keyword_overlap, candidate_title = tie_breaker
        return (-same_topic, -score, -title_overlap, -grouped_overlap, -keyword_overlap, candidate_title)

    for current in records:
        ranked = sorted(
            [candidate for candidate in records if candidate["slug"] != current["slug"]],
            key=lambda candidate: sort_key(current, candidate),
        )
        current["related_slugs"] = [book["slug"] for book in ranked[:4]]
        current["title_tokens"] = sorted(token_map[current["slug"]]["title"])
        current["topic_tokens"] = sorted(token_map[current["slug"]]["topic"])



def load_master() -> List[Dict[str, Any]]:
    master = read_json(MASTER_PATH, default=[])
    if not master:
        raise ValueError(f"Master file not found: {MASTER_PATH}")
    fallback_timestamp = infer_build_timestamp()
    for book in master:
        book.setdefault("dateModified", fallback_timestamp)
        book.setdefault("short_description", book.get("short", ""))
        sanitise_record_copy(book)
    add_related_books(master)
    return master



def save_master(records: List[Dict[str, Any]]) -> None:
    build_timestamp = utc_now()
    for book in records:
        book.setdefault("dateModified", build_timestamp)
    write_json(MASTER_PATH, records)



def book_to_public_record(book: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": book["id"],
        "key": book["key"],
        "title": book["title"],
        "short": book["short"],
        "short_description": book["short"],
        "cover": book["cover"],
        "main_image": book["main_image"],
        "tags": book["tags"],
        "filter": book["filter"],
        "keywords": book["keywords"],
        "buy_url": book["buy_url"],
        "buy_target_url": book["buy_url"],
        "slug": book["slug"],
        "asin": book["asin"],
        "pages": book["pages"],
        "datePublished": book["datePublished"],
        "dateModified": book.get("dateModified") or infer_build_timestamp(),
        "canonical_url": book["canonical_url"],
        "buy_route": book["buy_route"],
        "buy_route_full": book["buy_route_full"],
        "topic_url": book["topic_url"],
        "related_slugs": book.get("related_slugs", []),
        "title_tokens": book.get("title_tokens", []),
        "topic_tokens": book.get("topic_tokens", []),
    }



def featured_rotation_selection(now: dt.datetime | None = None) -> Dict[str, int | str]:
    current = now or dt.datetime.now(dt.timezone.utc)
    if current.tzinfo is None:
        current = current.replace(tzinfo=dt.timezone.utc)
    current_utc = current.astimezone(dt.timezone.utc)
    iso_year, iso_week, _ = current_utc.date().isocalendar()
    return {
        "method": "iso_week_rotation",
        "iso_week": iso_week,
        "year": iso_year,
    }



def select_featured_book_record(public_records: List[Dict[str, Any]], now: dt.datetime | None = None) -> tuple[Dict[str, Any] | None, Dict[str, int | str]]:
    selection = featured_rotation_selection(now=now)
    if not public_records:
        return None, selection
    return public_records[selection["iso_week"] % len(public_records)], selection



def build_podcast_sponsor_payload(book: Dict[str, Any]) -> Dict[str, str]:
    title = clean_paragraph(book.get("title"))
    short = clean_paragraph(book.get("short"))
    topic = clean_paragraph(book.get("filter"))
    canonical_url = clean_paragraph(book.get("canonical_url"))
    buy_route_full = clean_paragraph(book.get("buy_route_full"))
    pages = book.get("pages")
    tags = [clean_paragraph(tag) for tag in (book.get("tags") or []) if clean_paragraph(tag)]
    tags_text = ", ".join(tags[:3])

    page_sentence = f" It runs {pages} pages." if isinstance(pages, int) and pages > 0 else ""
    topic_sentence = f" It focuses on {topic}." if topic else ""
    tags_sentence = f" Topics include {tags_text}." if tags_text else ""
    short_sentence = short if short.endswith((".", "!", "?")) else (f"{short}." if short else "")

    return {
        "label": "This week's sponsor",
        "headline": f"This week's sponsor is {title}",
        "cta": f"See the book at {canonical_url} or buy on Amazon at {buy_route_full}",
        "midroll_15": f"This week's sponsor is {title}. {short_sentence} Read more at {canonical_url}".strip(),
        "midroll_30": (
            f"This week's sponsor is {title}."
            f" {short_sentence}"
            f"{topic_sentence}"
            f"{page_sentence}"
            f"{tags_sentence}"
            f" See the full book page at {canonical_url} or go straight to Amazon at {buy_route_full}."
        ).replace("  ", " ").strip(),
    }



def build_featured_book_payload(public_records: List[Dict[str, Any]], now: dt.datetime | None = None) -> Dict[str, Any]:
    book, selection = select_featured_book_record(public_records, now=now)
    return {
        "version": "v1",
        "selection": selection,
        "book": book or {},
        "podcast_sponsor": build_podcast_sponsor_payload(book or {}),
    }



def load_related_book_curation() -> Dict[str, List[Dict[str, str]]]:
    try:
        payload = json.loads(RELATED_BOOK_CURATION_PATH.read_text(encoding="utf-8"))
        items = payload.get("books", {}) if isinstance(payload, dict) else {}
        return items if isinstance(items, dict) else {}
    except Exception:
        return {}


def render_related_links(book: Dict[str, Any], all_books: List[Dict[str, Any]]) -> str:
    by_slug = {item["slug"]: item for item in all_books}
    curation = load_related_book_curation().get(book.get("slug", ""), [])
    related_specs: List[Tuple[str, str]] = []
    if isinstance(curation, list):
        for spec in curation[:4]:
            if isinstance(spec, dict):
                related_specs.append((clean_paragraph(spec.get("slug", "")), clean_paragraph(spec.get("reason", ""))))
    if not related_specs:
        related_specs = [(slug, "") for slug in book.get("related_slugs", [])[:4]]

    items = []
    for slug, reason in related_specs:
        related = by_slug.get(slug)
        if not related:
            continue
        reason_html = f'<span class="related-book-reason">{html.escape(reason)}</span>' if reason else f'<span>{html.escape(related["topic"])} · {related["pages"]} pages</span>'
        items.append(
            '<li><a href="/ebooks/{slug}/">{title}</a>{reason}</li>'.format(
                slug=html.escape(related["slug"]),
                title=html.escape(related["title"]),
                reason=reason_html,
            )
        )
    return "\n".join(items)

def render_faq_markup(book: Dict[str, Any]) -> str:
    parts = []
    for idx, item in enumerate(book.get("faq", []), start=1):
        question = clean_paragraph(item.get("name", ""))
        answer = clean_paragraph(item.get("acceptedAnswer", {}).get("text", ""))
        open_attr = " open" if idx == 1 else ""
        parts.append(
            "<details class=\"ebook-faq-item\"%s><summary>%s</summary><div><p>%s</p></div></details>"
            % (open_attr, html.escape(question), html.escape(answer))
        )
    return "\n".join(parts)



def render_breadcrumbs(book: Dict[str, Any]) -> str:
    crumbs = [
        '<a href="/">Home</a>',
        '<span aria-hidden="true">›</span>',
        '<a href="/ebooks/">eBooks</a>',
    ]
    if book.get("topic_url"):
        crumbs.extend([
            '<span aria-hidden="true">›</span>',
            '<a href="{url}">{topic}</a>'.format(url=html.escape(book["topic_url"]), topic=html.escape(book["topic"])),
        ])
    crumbs.extend([
        '<span aria-hidden="true">›</span>',
        '<span>{title}</span>'.format(title=html.escape(book["title"])),
    ])
    return "".join(crumbs)

def render_image_tag(*, src: str, alt: str, class_name: str, loading: str = "lazy", width: int | None = None, height: int | None = None, srcset: str | None = None, sizes: str | None = None, fetchpriority: str | None = None) -> str:
    attrs = [
        f'alt="{html.escape(alt)}"',
        f'class="{html.escape(class_name)}"',
        f'decoding="async"',
        f'loading="{html.escape(loading)}"',
        f'src="{html.escape(src)}"',
    ]
    if srcset:
        attrs.append(f'srcset="{html.escape(srcset)}"')
        if sizes:
            attrs.append(f'sizes="{html.escape(sizes)}"')
    if fetchpriority:
        attrs.append(f'fetchpriority="{html.escape(fetchpriority)}"')
    if width is not None:
        attrs.append(f'width="{width}"')
    if height is not None:
        attrs.append(f'height="{height}"')
    return "<img " + " ".join(attrs) + "/>"


def render_cover_image(book: Dict[str, Any], class_name: str, loading: str = "lazy") -> str:
    return render_image_tag(
        src=book["cover"],
        alt=f"{book['title']} cover",
        class_name=class_name,
        loading=loading,
        width=BOOK_COVER_WIDTH,
        height=BOOK_COVER_HEIGHT,
        srcset=build_same_source_srcset(book["cover"], BOOK_COVER_WIDTH),
        sizes=cover_sizes(class_name),
        fetchpriority="high" if loading == "eager" else None,
    )


def audience_bullets(book: Dict[str, Any]) -> List[str]:
    topic_name = clean_paragraph(book.get("topic", "Artificial Intelligence"))
    topic_lc = topic_name.lower()
    title_stem = clean_paragraph(book.get("title", "")).split(":", 1)[0]
    family = topic_family(topic_name)
    learn_items = book.get("what_youll_learn", [])
    first_signal = clean_paragraph(learn_items[0] if learn_items else "").rstrip(".")
    second_signal = clean_paragraph(learn_items[1] if len(learn_items) > 1 else "").rstrip(".")
    family_line = {
        "regulated": f"Readers working in or around {topic_lc} who need the practical trade-offs explained before policy, procurement, or implementation decisions harden.",
        "operations": f"Operators, managers, and curious readers who want to know whether AI in {topic_lc} improves the workflow or just adds another dashboard to ignore.",
        "security": f"Security-minded readers who need a clearer feel for where AI helps defenders in {topic_lc} and where it simply reshuffles the noise.",
        "creative": f"People working around {topic_lc} who want the upside explained without pretending the ownership and control questions vanished overnight.",
        "foundation": f"Readers who want the bigger arguments around {topic_lc} translated into plain English and tested against how the systems behave in practice.",
        "environment": f"Readers who want environmental claims in {topic_lc} measured against evidence, costs, and real deployment constraints.",
        "sports": f"Readers interested in how AI changes judgement, preparation, and performance in {topic_lc} without turning humans into footnotes.",
    }.get(family, f"Readers who want a practical, grounded route into {topic_lc} rather than another brochure about inevitable disruption.")
    slug = clean_paragraph(book.get("slug", ""))
    signal_line = first_signal or f"How AI is being used in {topic_lc} today"
    follow_on = second_signal or clean_paragraph(book.get("what_this_book_covers", "")).split(".")[0]
    signal_line = signal_line[0].lower() + signal_line[1:] if signal_line else topic_lc
    follow_on = follow_on[0].lower() + follow_on[1:] if follow_on else topic_lc
    bullets = [
        replace_banned_ebook_phrases(f"Curious readers who want a grounded view of {title_stem} without the applause soundtrack.", slug),
        replace_banned_ebook_phrases(family_line, slug),
        replace_banned_ebook_phrases(f"Anyone who wants clear context on {signal_line} before they trust the louder claims.", slug),
        replace_banned_ebook_phrases(f"Readers looking for sharper judgement on {follow_on} rather than recycled buzzwords.", slug),
    ]
    return bullets


def chapter_signal_cards(book: Dict[str, Any]) -> str:
    """Use genuinely distinct source fields, and chapter data when supplied."""
    chapters = book.get("chapters")
    cards: List[str] = []
    if isinstance(chapters, list) and chapters:
        for chapter in chapters[:3]:
            if isinstance(chapter, dict):
                heading = clean_paragraph(chapter.get("title", "Chapter signal"))
                text = clean_paragraph(chapter.get("fact") or chapter.get("summary") or "")
            else:
                heading, text = "Chapter signal", clean_paragraph(chapter)
            if text:
                cards.append(f'<article class="ebook-signal-card"><h3>{html.escape(heading)}</h3><p>{html.escape(text)}</p></article>')
    if not cards:
        distinct = [
            ("Use cases and workflow", clean_paragraph(book.get("what_this_book_covers", ""))),
            ("Why the stakes matter", clean_paragraph(book.get("why_it_matters", ""))),
            ("The book's distinct angle", clean_paragraph(book.get("distinct_angle", ""))),
        ]
        cards = [
            f'<article class="ebook-signal-card"><h3>{html.escape(heading)}</h3><p>{html.escape(text)}</p></article>'
            for heading, text in distinct if text
        ]
    return "\n".join(cards[:3])


def problem_framing(book: Dict[str, Any]) -> str:
    topic_lc = clean_paragraph(book.get("topic", "")).lower() or "the field"
    family = topic_family(book.get("topic", ""))
    learn_line = clean_paragraph(book["what_youll_learn"][0] if book.get("what_youll_learn") else "").rstrip(".")
    framing_tail = f" It keeps coming back to {learn_line[0].lower() + learn_line[1:] if learn_line else 'the practical decisions that expose weak claims fastest'}."
    frames = {
        "regulated": f"{book['topic']} is where speed, evidence, compliance, and accountability all start elbowing each other for room. This title keeps the focus on what AI is genuinely doing in {topic_lc}, where oversight has to tighten, and where the expensive mistakes tend to hide.{framing_tail}",
        "operations": f"{book['topic']} is where efficiency claims meet maintenance logs, handovers, failure modes, and people who still have to run the place. This title looks at what AI is actually changing in {topic_lc}, which gains are solid, and where the shiny promise falls apart under operational pressure.{framing_tail}",
        "security": f"{book['topic']} is one of those domains where signal, false positives, attacker behaviour, and tool sprawl all collide at speed. This title looks at what AI is really doing in {topic_lc}, where it strengthens the work, and where automation simply changes the shape of the problem.{framing_tail}",
        "creative": f"{book['topic']} gets messy fast because speed, originality, ownership, and platform incentives do not naturally get along. This title looks at what AI is really doing in {topic_lc}, what it improves, and what it muddies the moment the convenience starts looking irresistible.{framing_tail}",
        "foundation": f"{book['topic']} is one of those areas where the argument gets noisy very quickly: claims versus evidence, fluency versus substance, novelty versus context. This title cuts through that din and looks at what AI is actually doing in {topic_lc}, where it helps, and where it starts creating fresh headaches.{framing_tail}",
        "environment": f"{book['topic']} attracts hopeful claims because everyone likes a cleaner future and a clever dashboard. This title looks at what AI is actually doing in {topic_lc}, where the measurable gains are, and where the story outruns the evidence.{framing_tail}",
        "sports": f"{book['topic']} sits in that awkward space where numbers can sharpen judgement or flatten it into false certainty. This title looks at what AI is actually doing in {topic_lc}, where the edge is real, and where the human part still refuses to disappear.{framing_tail}",
    }
    return frames.get(
        family,
        f"{book['topic']} is one of those areas where the argument gets noisy: efficiency versus judgement, convenience versus control, automation versus accountability. This title cuts through that din and looks at what AI is actually doing in {topic_lc}, where it helps, and where it starts to create fresh headaches.{framing_tail}",
    )


def practical_outcomes(book: Dict[str, Any]) -> List[str]:
    """Return practical, decision-oriented outcomes distinct from what_youll_learn bullets.
    
    These are action-framed: what can the reader DO or DECIDE differently after reading?
    Derived from why_it_matters, audience, and topic — NOT from what_youll_learn items.
    """
    topic = book.get("topic", "AI")
    topic_lc = topic.lower()
    audience = book.get("audience", "")
    why = book.get("why_it_matters", "")
    
    # Build decision-framed outcomes from distinct source fields
    outcomes = []
    
    # Outcome 1: from why_it_matters (if available and not empty)
    if why and len(why) > 20:
        # Extract first sentence as an action-frame
        first_sentence = why.split(".")[0].strip()
        if first_sentence and len(first_sentence) > 15:
            outcomes.append(f"Understand why {topic_lc} matters now and what the evidence actually says.")
    
    # Outcome 2: from audience/use-case context
    if audience and len(audience) > 10:
        outcomes.append(f"Assess whether {topic_lc} is applicable to your context before committing resources.")
    else:
        outcomes.append(f"Identify the specific use cases where {topic_lc} delivers measurable value.")
    
    # Outcome 3: always a governance/decision frame
    outcomes.append(f"Ask the right governance and implementation questions before adoption decisions become expensive.")
    
    return outcomes[:3] or default_learning_points(topic)[:3]



def render_book_page(book: Dict[str, Any], all_books: List[Dict[str, Any]]) -> str:
    header = render_header()
    footer = render_footer()
    hero_summary = book["summary"] or strip_pages_from_summary(book["description"], book.get("pages"))
    title = html.escape(book["title"])
    meta_title = html.escape(book_meta_title(book))
    description = html.escape(book_meta_description(book))
    canonical = html.escape(book["canonical_url"])
    cover = html.escape(book["cover"])
    faq_schema = {
        "@context": "https://schema.org",
        "@type": "FAQPage",
        "@id": f"{book['canonical_url']}#faq",
        "mainEntity": book.get("faq", []),
    }
    learn_items = "\n".join(f"<li>{html.escape(item)}</li>" for item in book.get("what_youll_learn", []))
    audience_items = "\n".join(f"<li>{html.escape(item)}</li>" for item in audience_bullets(book))
    key_theme_items = "\n".join(f"<li>{html.escape(item)}</li>" for item in book.get("tags", []))
    signal_items = "\n".join(f"<li>{html.escape(item)}</li>" for item in practical_outcomes(book))
    sample = load_book_sample_chapters().get(book["slug"])
    sample_available = bool(sample and sample.get("paragraphs") and int(sample.get("word_count") or 0) >= 350)
    newsletter_offer = newsletter_offer_for_book(book)
    podcast_link = podcast_link_for_book(book)
    sample_primary_cta = (
        f'<a class="button secondary" href="{book_preview_path(book)}" data-ebook-preview data-book-slug="{html.escape(book["slug"])}" data-topic="{html.escape(book["topic_slug"])}" data-placement="ebook_primary">Read a free chapter</a>'
        if sample_available else ""
    )
    sample_confidence = " · Free chapter available" if sample_available else ""
    sample_section = (
        f"""<section class="card ebook-section ebook-preview-capture" id="free-preview">
      <h2>Read a real chapter before you buy</h2>
      <p>Open a genuine chapter from the manuscript now. The AI Edge signup is optional, not a gate between you and the sample.</p>
      <div class="ebook-actions"><a class="button secondary" href="{book_preview_path(book)}" data-ebook-preview data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="ebook_sample_section">Read the sample chapter</a></div>
      {render_inline_newsletter_form(f"ebook:{book['slug']}", next_path=newsletter_offer["next_path"], cta="Join AI Edge", heading=newsletter_offer["heading"], description=newsletter_offer["description"])}
    </section>"""
        if sample_available
        else f"""<section class="card ebook-section ebook-preview-capture" id="free-preview">
      <h2>Get the free AI glossary</h2>
      <p>The manuscript sample is not available in this build, so this page does not promise a chapter it cannot deliver.</p>
      {render_inline_newsletter_form(f"ebook:{book['slug']}", next_path=newsletter_offer["next_path"], cta="Join AI Edge", heading=newsletter_offer["heading"], description=newsletter_offer["description"])}
    </section>"""
    )

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<!-- Google Tag Manager -->
<script>(function(w,d,s,l,i){{w[l]=w[l]||[];w[l].push({{'gtm.start':
new Date().getTime(),event:'gtm.js'}});var f=d.getElementsByTagName(s)[0],
j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
}})(window,document,'script','dataLayer','GTM-PC4K9KRK');</script>
<!-- End Google Tag Manager -->
<meta charset="utf-8"/>
<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>
<link href="https://images.jonathan-harris.online" rel="preconnect"/>
<link href="https://assets.jonathan-harris.online" rel="preconnect"/>
<meta content="width=device-width, initial-scale=1, viewport-fit=cover" name="viewport"/>
<title>{meta_title} | Jonathan Harris</title>
<meta content="@jonathan_harris_01" name="twitter:site"/>
<meta content="@jonathan_harris_01" name="twitter:creator"/>
<meta content="#0D1420" name="theme-color"/>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link as="style" href="/assets/css/site.css" rel="preload"/>
<link href="/assets/css/site.css" rel="stylesheet"/>


<link href="/assets/css/ebook-template.css" rel="stylesheet"/>
<meta content="GB" name="geo.region"/>
<link href="{canonical}" rel="canonical"/>
<link href="{canonical}" hreflang="en" rel="alternate"/>
<link href="{canonical}" hreflang="x-default" rel="alternate"/>
<meta content="{description}" name="description"/>
<meta content="index,follow" name="robots"/>
<meta content="{html.escape(book['asin'])}" name="book:asin"/>
<meta content="{html.escape(book['datePublished'])}" name="datePublished"/>
<meta content="{html.escape(book['datePublished'])}" property="book:release_date"/>
<meta content="books.book" property="og:type"/>
<meta content="{canonical}" property="og:url"/>
<meta content="{meta_title} | Jonathan Harris" property="og:title"/>
<meta content="{description}" property="og:description"/>
<meta content="{cover}" property="og:image"/>
<meta content="{title} cover" property="og:image:alt"/>
<meta content="summary_large_image" name="twitter:card"/>
<meta content="{cover}" name="twitter:image"/>
<meta content="{meta_title} | Jonathan Harris" name="twitter:title"/>
<meta content="{description}" name="twitter:description"/>
<meta content="{title}" name="ai:topic"/>
<meta content="Artificial Intelligence" name="ai:primary"/>
<meta content="Book" name="ai:entity"/>
<meta content="{html.escape(book['identifier'])}" name="ai:identifier"/>
<meta content="book" name="ai:content_type"/>
<meta content="{html.escape(', '.join(book['keywords']))}" name="ai:keywords"/>
<meta content="{html.escape(book['tone'])}" name="ai-style"/>
<meta content="{html.escape(book['audience'])}" name="ai-target-audience"/>
<meta content="search=y, train-ai=y, citation-preferred=y" name="content-usage"/>
<script type="application/ld+json">{json_script(build_breadcrumb_schema(book))}</script>
<script data-jh-ai-pack="person" type="application/ld+json">{json_script(build_person_schema())}</script>
<script data-jh-ai-pack="website" type="application/ld+json">{json_script(build_website_schema())}</script>
<script type="application/ld+json">{json_script(build_book_schema(book))}</script>
<script type="application/ld+json">{json_script(faq_schema)}</script>
<link href="https://tracker.metricool.com" rel="dns-prefetch"/>
<link href="https://botsailor.com" rel="dns-prefetch"/>
<script defer="" data-cookieyes="ignore" data-cookieconsent="ignore" src="/assets/js/script-governance.min.js"></script>
</head>
<body class="ebook-detail" data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-page-type="ebook">
<!-- Google Tag Manager (noscript) -->
<noscript><iframe src="https://www.googletagmanager.com/ns.html?id=GTM-PC4K9KRK"
height="0" width="0" style="display:none;visibility:hidden"></iframe></noscript>
<!-- End Google Tag Manager (noscript) -->
{header}
<div aria-hidden="false" class="page-loader" id="pageLoader">
  <div aria-label="Preparing page" aria-live="polite" class="loader-card" role="status">
    <div aria-hidden="true" class="spinner"></div>
  </div>
</div>
<header aria-label="Book header" class="hero ebook-hero" role="region">
  <div class="wrap">
    <img alt="Jonathan Harris site logo" class="logo-plain" height="120" src="https://images.jonathan-harris.online/site-logo" width="120"/>
    <h1>{title}</h1>
    <p>{html.escape(hero_summary)}</p>
  </div>
</header>
<main aria-label="Book content" class="main" id="main" role="main">
  <div class="wrap ebook-shell">
    <nav aria-label="Breadcrumb" class="breadcrumbs">{render_breadcrumbs(book)}</nav>

    <section class="card ebook-section quick-facts">
      <h2>Quick facts</h2>
      <ul class="ebook-facts">
        <li><strong>Topic:</strong> {html.escape(book['topic'])}</li>
        <li><strong>Tags:</strong> {html.escape(', '.join(book['tags']))}</li>
        <li><strong>Length:</strong> {book['pages']} pages</li>
        <li><strong>Best for:</strong> {html.escape(book['audience'])}</li>
      </ul>
    </section>

    <section class="ebook-showcase">
      <article class="card ebook-showcase__media">
        {render_cover_image(book, class_name="cover ebook-showcase__cover", loading="eager")}
        <div class="ebook-actions">
          <a class="button" href="{html.escape(book['buy_route'])}" data-ebook-amazon data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="ebook_primary">Buy on Amazon</a>
          {sample_primary_cta}
          <a class="button secondary" href="/ebooks/">Browse related books</a>
        </div>
        <aside class="book-confidence" aria-label="Buying information"><strong>Before you buy</strong> — {book['pages']} pages · Kindle ebook · Published {html.escape(format_date(book['datePublished']))}{sample_confidence} · {html.escape(book['audience'])}<br/><a href="#deeper-overview">See exactly what this book covers ↓</a></aside>
        {render_book_market_signal(book)}
        <p class="meta">ASIN: {html.escape(book['asin'])} · Published {html.escape(format_date(book['datePublished']))}</p>
      </article>
      <article class="card ebook-showcase__content">
        <h2>{html.escape(book['showcase_heading'])}</h2>
        <p class="ebook-showcase__lead">{html.escape(book['what_this_book_covers'])}</p>
        <p class="ebook-showcase__subhead">{html.escape(showcase_subhead(book))}</p>
        <ul class="ebook-signal-list">
          {''.join(f'<li>► {html.escape(item)}</li>' for item in book.get('what_youll_learn', [])[:3])}
        </ul>
        <p class="ebook-showcase__note">{html.escape(showcase_note(book))}</p>
        <div class="ebook-inline-actions">
          <a href="{html.escape(book['buy_route'])}">Buy on Amazon</a>
          <a href="#deeper-overview">Read full overview</a>
          <a href="/ebooks/">Browse related books</a>
        </div>
      </article>
    </section>

    <section class="card ebook-section">
      <h2>Who is this book for?</h2>
      <ul class="ebook-audience-list">
        {audience_items}
      </ul>
    </section>

    <section class="card ebook-section">
      <h2>Key themes</h2>
      <ul class="ebook-key-themes">
        {key_theme_items}
      </ul>
      <div class="ebook-theme-pills">{render_tag_pills(book['tags'])}</div>
    </section>

    <section class="card ebook-section">
      <h2>What will you learn?</h2>
      <ul class="ebook-learn-list">
        {learn_items}
      </ul>
    </section>

    <section class="card ebook-section">
      <h2>Audience fit</h2>
      {escape_paragraphs(book['who_for'])}
    </section>

    <section class="card ebook-section" id="deeper-overview">
      <h2>Deeper overview</h2>
      {escape_paragraphs(book['summary'])}
    </section>


    <section class="card ebook-section">
      <h2>Why this title is useful in practice</h2>
      <p>{html.escape(book_unique_evidence_passage(book))}</p>
    </section>

    {render_priority_evidence_module(book)}

    <section class="card ebook-section ebook-section--accent">
      <h2>Why does this topic get messy?</h2>
      {escape_paragraphs(problem_framing(book))}
    </section>

    <section class="card ebook-section">
      <h2>What practical decisions will this help with?</h2>
      <p>{html.escape(practical_outcomes_intro(book))}</p>
      <ul class="ebook-learn-list">
        {signal_items}
      </ul>
    </section>

    <section class="card ebook-section">
      <h2>What evidence lenses does the book use?</h2>
      <div class="ebook-signal-grid">
        {chapter_signal_cards(book)}
      </div>
    </section>

    <section class="card ebook-section">
      <h2>What makes this title distinct</h2>
      {escape_paragraphs(book['distinct_angle'])}
      {escape_paragraphs(book['why_it_matters'])}
    </section>

    {sample_section}

    <section class="related-books card">
      <h2>Related books</h2>
      <ul>
        {render_related_links(book, all_books)}
      </ul>
      <p class="jh-related-callout">Related titles are chosen from the catalogue based on topic and tag overlap, so the next step stays relevant instead of wandering off into the weeds.</p>
      <div class="ebook-inline-actions">{render_book_bundle_links(book)}</div>
    </section>

    <section class="card ebook-section ebook-listen-next" aria-labelledby="listen-next-heading">
      <h2 id="listen-next-heading">Listen next</h2>
      <p>Continue with current audio analysis related to {html.escape(book['topic'].lower())}. Episode metadata stays governed by the podcast feed rather than being copied into this book page.</p>
      <a class="button secondary" href="{html.escape(podcast_link['href'])}" data-podcast-contextual data-placement="ebook_listen_next">{html.escape(podcast_link['label'])}</a>
    </section>

    <section class="faq card" aria-label="Frequently asked questions">
      <h2>FAQ</h2>
      <div class="ebook-faq-list">
        {render_faq_markup(book)}
      </div>
    </section>

    <section class="jh-journey-panel">
      <h2>Keep exploring the Jonathan Harris AI library</h2>
      <p>Use the links below to carry on browsing the wider catalogue, the glossary, comparisons, podcast coverage, or a related guide.</p>
      <div class="jh-journey-actions">
        {book_semantic_journey_links(book)}
      </div>
      {render_inline_newsletter_form(f"ebook-footer:{book['slug']}", next_path=newsletter_offer["next_path"], cta="Join AI Edge", heading=newsletter_offer["heading"], description=newsletter_offer["description"])}
    </section>
  </div>
</main>
<script defer="" src="/assets/js/related-books.min.js"></script>
{footer}
<script defer="" src="/assets/js/site-ui.min.js"></script>
</body>
</html>
'''



def render_catalogue_card(book: Dict[str, Any], cta_copy: str) -> str:
    tags = "".join(f'<span class="tag">{html.escape(tag)}</span>' for tag in book.get("tags", [])[:4])
    return f'''
<article class="card ebook-card" aria-label="{html.escape(book['title'])}" data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}">
  {render_cover_image(book, class_name="cover")}
  <h2>{html.escape(book['title'])}</h2>
  <div class="topic-chip-wrap"><span class="topic-chip">{html.escape(book['filter'])}</span></div>
  <p>{html.escape(book['short'])}</p>
  {render_book_market_signal(book)}
  <div class="tags">{tags}</div>
  <div class="book-avail"><span class="book-avail__badge">🛍️ Available on Amazon Kindle</span></div>
  <div class="actions ebook-card-actions">
    <a class="button secondary" href="/ebooks/{html.escape(book['slug'])}/" data-ebook-action="view" data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="catalogue_card">View book</a>
    <a class="button" href="{html.escape(book['buy_route'])}" data-ebook-amazon data-book-slug="{html.escape(book['slug'])}" data-topic="{html.escape(book['topic_slug'])}" data-placement="catalogue_card">Buy on Amazon</a>
  </div>
  <details class="more">
    <summary aria-expanded="false">More details</summary>
    <div class="meta">{html.escape(cta_copy)}</div>
  </details>
</article>'''.strip()



def render_topic_hub_links(books: List[Dict[str, Any]]) -> str:
    topics = sorted({(book["topic"], book["topic_slug"]) for book in books}, key=lambda item: item[0].lower())
    return "\n".join(f'<a href="/catalogue/{slug}/">{html.escape(name)}</a>' for name, slug in topics[:12])



def render_ebooks_index(books: List[Dict[str, Any]]) -> str:
    header = render_header()
    footer = render_footer()
    canonical = f"{SITE_URL}/ebooks/"
    description = f"Ebook catalogue: {len(books)} AI titles by Jonathan Harris covering industries, ethics, safety, and practical adoption."
    item_list = {
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": "Jonathan Harris eBooks library",
        "itemListElement": [
            {"@type": "ListItem", "position": idx, "url": book["canonical_url"], "name": book["title"]}
            for idx, book in enumerate(books, start=1)
        ],
    }
    static_cards = "\n".join(render_catalogue_card(book, catalogue_cta_copy(book["topic"])) for book in books)
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<!-- Google Tag Manager -->
<script>(function(w,d,s,l,i){{w[l]=w[l]||[];w[l].push({{'gtm.start':
new Date().getTime(),event:'gtm.js'}});var f=d.getElementsByTagName(s)[0],
j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
}})(window,document,'script','dataLayer','GTM-PC4K9KRK');</script>
<!-- End Google Tag Manager -->
<meta charset="utf-8"/>
<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>
<link href="https://images.jonathan-harris.online" rel="preconnect"/>
<link href="https://assets.jonathan-harris.online" rel="preconnect"/>
<meta content="width=device-width, initial-scale=1, viewport-fit=cover" name="viewport"/>
<title>AI eBooks Catalogue | Jonathan Harris</title>
<meta content="{description}" name="description"/>
<meta content="index,follow" name="robots"/>
<meta content="{description}" name="ai:summary"/>
<meta content="#0D1420" name="theme-color"/>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link as="style" href="/assets/css/site.css" rel="preload"/>
<link href="/assets/css/site.css" rel="stylesheet"/>


<link href="/assets/css/ebook-template.css" rel="stylesheet"/>
<meta content="GB" name="geo.region"/>
<meta content="website" property="og:type"/>
<meta content="Jonathan Harris eBooks" property="og:site_name"/>
<meta content="{canonical}" property="og:url"/>
<meta content="AI eBooks Catalogue | Jonathan Harris" property="og:title"/>
<meta content="{description}" property="og:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" property="og:image"/>
<meta content="Jonathan Harris site logo" property="og:image:alt"/>
<meta content="1200" property="og:image:width"/>
<meta content="630" property="og:image:height"/>
<meta content="@jonathan_harris_01" name="twitter:site"/>
<meta content="@jonathan_harris_01" name="twitter:creator"/>
<meta content="summary_large_image" name="twitter:card"/>
<meta content="AI eBooks Catalogue | Jonathan Harris" name="twitter:title"/>
<meta content="{description}" name="twitter:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" name="twitter:image"/>
<meta content="AI ebook catalogue and practical AI guide" name="ai-role"/>
<meta content="Curious professionals, entrepreneurs, and non-technical readers who want practical AI insight" name="ai-target-audience"/>
<meta content="Plain-English, practical, sceptical, no-hype" name="ai-style"/>
<meta content="search=y, train-ai=y" name="content-usage"/>
<link href="{canonical}" rel="canonical"/>
<link href="{canonical}" hreflang="en" rel="alternate"/>
<link href="{canonical}" hreflang="x-default" rel="alternate"/>
<script type="application/ld+json">{json_script(item_list)}</script>
<script data-jh-ai-pack="person" type="application/ld+json">{json_script(build_person_schema())}</script>
<script data-jh-ai-pack="website" type="application/ld+json">{json_script(build_website_schema())}</script>
<link href="https://tracker.metricool.com" rel="dns-prefetch"/>
<link href="https://botsailor.com" rel="dns-prefetch"/>
<script defer="" data-cookieyes="ignore" data-cookieconsent="ignore" src="/assets/js/script-governance.min.js"></script>
</head>
<body class="ebooks-catalogue">
<!-- Google Tag Manager (noscript) -->
<noscript><iframe src="https://www.googletagmanager.com/ns.html?id=GTM-PC4K9KRK"
height="0" width="0" style="display:none;visibility:hidden"></iframe></noscript>
<!-- End Google Tag Manager (noscript) -->
{header}
<div aria-hidden="false" class="page-loader is-active" id="pageLoader">
  <div aria-label="Preparing page" aria-live="polite" class="loader-card" role="status">
    <div aria-hidden="true" class="spinner"></div>
  </div>
</div>
<header class="hero ebook-hero ebook-catalogue-hero" role="region" aria-label="eBook catalogue intro">
  <div class="wrap">
    <h1>AI eBooks Catalogue</h1>
    <p>Browse {len(books)} practical, plain-English AI titles covering industries, ethics, safety, and real-world adoption.</p>
  </div>
</header>
<main class="main" id="main" role="main" aria-label="eBook catalogue">
  <div class="wrap">
    <section class="ebook-catalogue-controls" aria-labelledby="ebook-finder-heading">
      <div class="ebook-catalogue-controls__intro">
        <div>
          <h2 id="ebook-finder-heading">Find the right AI eBook</h2>
          <p>Search by title or keyword, then narrow the catalogue with one topic filter.</p>
        </div>
        <p class="meta ebook-count" id="count">{len(books)} of {len(books)} books</p>
      </div>

      <div class="toolbar ebook-catalogue-toolbar" aria-label="Catalogue controls">
        <label class="ebook-search-label" for="search">Search the catalogue</label>
        <input aria-label="Search books" class="search" id="search" placeholder="Search by title, topic, or keyword" type="search"/>
        <div aria-label="Filter books by topic" class="chips" id="chips"></div>
      </div>
    </section>

    <section class="card book-finder-bridge"><h2>Prefer a guided starting point?</h2><p>Use the rule-based finder when a 40-book grid is a bit too much buffet.</p><a class="button secondary" href="/book-finder/?source=ebooks-index">Find the right AI book</a></section>

    <section aria-label="eBook grid" class="grid" id="booksGrid">
      {static_cards}
    </section>

    <nav aria-label="Pagination" class="pager u-s24" id="pager">
      <button class="button secondary" id="prevPage" type="button">Previous</button>
      <span class="meta" id="pageInfo"></span>
      <button class="button" id="nextPage" type="button">Next</button>
    </nav>

    <section class="jh-journey-panel">
      <h2>Keep exploring the wider library</h2>
      <p>Every ebook page links back into the catalogue, but you can also hop out to the podcast, newsletter, or topic guides when you want the broader view.</p>
      <div class="jh-journey-actions">
        <a href="/podcast/">Listen to the podcast</a>
        <a href="/newsletter/">Join the newsletter</a>
        <a href="/topics/">Explore topic guides</a>
      </div>
      <details class="ebook-topic-directory">
        <summary>Browse ebook collections by topic</summary>
        <nav aria-label="eBook topic collections" class="jh-topic-links">
          {render_topic_hub_links(books)}
        </nav>
      </details>
      {render_inline_newsletter_form("ebooks-index")}
    </section>
  </div>
</main>
{footer}
<script defer="" src="/assets/js/books.min.js"></script>
<script defer="" src="/assets/js/site-ui.min.js"></script>
</body>
</html>
'''



def render_topic_page(topic: str, books: List[Dict[str, Any]]) -> str:
    header = render_header()
    footer = render_footer()
    cards = "\n".join(render_catalogue_card(book, catalogue_cta_copy(topic)) for book in books)
    topic_slug = slugify(topic)
    canonical = f"{SITE_URL}/catalogue/{topic_slug}/"
    description = catalogue_meta_description(topic)
    item_list = {
        "@context": "https://schema.org",
        "@type": "ItemList",
        "name": f"{topic} books by Jonathan Harris",
        "itemListElement": [
            {"@type": "ListItem", "position": idx, "url": book["canonical_url"], "name": book["title"]}
            for idx, book in enumerate(books, start=1)
        ],
    }
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<!-- Google Tag Manager -->
<script>(function(w,d,s,l,i){{w[l]=w[l]||[];w[l].push({{'gtm.start':
new Date().getTime(),event:'gtm.js'}});var f=d.getElementsByTagName(s)[0],
j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
}})(window,document,'script','dataLayer','GTM-PC4K9KRK');</script>
<!-- End Google Tag Manager -->
<meta charset="utf-8"/>
<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>
<link href="https://images.jonathan-harris.online" rel="preconnect"/>
<link href="https://assets.jonathan-harris.online" rel="preconnect"/>
<meta content="width=device-width, initial-scale=1, viewport-fit=cover" name="viewport"/>
<title>{html.escape(topic)} AI Books | Jonathan Harris</title>
<meta content="{html.escape(description)}" name="description"/>
<meta content="index,follow" name="robots"/>
<meta content="#0D1420" name="theme-color"/>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link as="style" href="/assets/css/site.css" rel="preload"/>
<link href="/assets/css/site.css" rel="stylesheet"/>


<link href="/assets/css/ebook-template.css" rel="stylesheet"/>
<meta content="GB" name="geo.region"/>
<meta content="website" property="og:type"/>
<meta content="{canonical}" property="og:url"/>
<meta content="{html.escape(topic)} AI Books | Jonathan Harris" property="og:title"/>
<meta content="{html.escape(description)}" property="og:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" property="og:image"/>
<meta content="Jonathan Harris site logo" property="og:image:alt"/>
<meta content="@jonathan_harris_01" name="twitter:site"/>
<meta content="@jonathan_harris_01" name="twitter:creator"/>
<meta content="summary_large_image" name="twitter:card"/>
<meta content="{html.escape(topic)} AI Books | Jonathan Harris" name="twitter:title"/>
<meta content="{html.escape(description)}" name="twitter:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" name="twitter:image"/>
<link href="{canonical}" rel="canonical"/>
<link href="{canonical}" hreflang="en" rel="alternate"/>
<link href="{canonical}" hreflang="x-default" rel="alternate"/>
<script type="application/ld+json">{json_script(item_list)}</script>
<script type="application/ld+json">{json_script(build_topic_breadcrumb_schema(topic))}</script>
<script data-jh-ai-pack="person" type="application/ld+json">{json_script(build_person_schema())}</script>
<script data-jh-ai-pack="website" type="application/ld+json">{json_script(build_website_schema())}</script>
</head>
<body class="ebooks-catalogue topic-catalogue">
{header}
<header class="hero ebook-hero" role="region" aria-label="Topic catalogue intro">
  <div class="wrap">
    <img alt="Jonathan Harris site logo" class="logo-plain" height="120" src="https://images.jonathan-harris.online/site-logo" width="120"/>
    <h1>{html.escape(topic)} AI Books</h1>
    <p>{html.escape(topic_intro(topic))}</p>
  </div>
</header>
<main class="main" id="main" role="main">
  <div class="wrap ebook-shell">
    <nav aria-label="Breadcrumb" class="breadcrumbs">{render_topic_breadcrumbs(topic)}</nav>
    <section class="card ebook-index-intro">
      <h2>{category_question_heading(topic)}</h2>
      <p>{html.escape(category_answer_first_copy(topic, books))}</p>
    </section>
    <section class="card ebook-index-intro">
      <h2>What this category covers</h2>
      <p>{html.escape(category_scope_copy(topic, books))}</p>
    </section>
    <section class="card ebook-index-intro">
      {render_inline_newsletter_form(f"topic:{topic_slug}")}
    </section>
    <section class="card ebook-index-intro">
      <h2>Best place to start</h2>
      <p>{category_best_start_copy(topic, books)}</p>
    </section>
    <section class="card book-finder-bridge">
      <h2>Not sure which {html.escape(topic)} book fits?</h2>
      <p>Use the rule-based finder to narrow the 40-book catalogue by the problem you are trying to solve.</p>
      <a class="button secondary" href="/book-finder/?source=catalogue-{html.escape(topic_slug, quote=True)}" data-book-finder-bridge data-placement="catalogue:{html.escape(topic_slug, quote=True)}">Find the right AI book</a>
    </section>
    <section class="faq card" aria-label="Category questions">
      <h2>Common questions</h2>
      <div class="ebook-faq-list">
        {category_faq_markup(topic, books)}
      </div>
    </section>
    <section class="related-books card">
      <h2>Keep exploring this topic</h2>
      <ul>
        {category_internal_links(topic, books)}
      </ul>
    </section>
    <section class="card ebook-index-intro">
      <h2>{len(books)} title{'s' if len(books) != 1 else ''} in this topic</h2>
      <p>{html.escape(catalogue_intro_copy(topic))}</p>
    </section>
    <section aria-label="Topic book grid" class="grid" id="booksGrid">
      {cards}
    </section>
  </div>
</main>
{footer}
<script defer="" src="/assets/js/site-ui.min.js"></script>
</body>
</html>
'''


def render_topics_index(topic_map: Dict[str, List[Dict[str, Any]]]) -> str:
    header = render_header()
    footer = render_footer()
    canonical = f"{SITE_URL}/topics/"
    description = topics_index_meta_description()
    cards = []
    for topic in sorted(topic_map, key=str.lower):
        slug = slugify(topic)
        title_count = len(topic_map[topic])
        title_plural = "s" if title_count != 1 else ""
        cards.append(
            f'<article class="card topic-card"><h2><a href="/catalogue/{slug}/">{html.escape(topic)}</a></h2>'
            f'<p>{title_count} title{title_plural}</p></article>'
        )
    cards_html = "\n".join(cards)
    guide_cards_html = topic_guide_cards_markup()
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<!-- Google Tag Manager -->
<script>(function(w,d,s,l,i){{w[l]=w[l]||[];w[l].push({{'gtm.start':
new Date().getTime(),event:'gtm.js'}});var f=d.getElementsByTagName(s)[0],
j=d.createElement(s),dl=l!='dataLayer'?'&l='+l:'';j.async=true;j.src=
'https://www.googletagmanager.com/gtm.js?id='+i+dl;f.parentNode.insertBefore(j,f);
}})(window,document,'script','dataLayer','GTM-PC4K9KRK');</script>
<!-- End Google Tag Manager -->
<meta charset="utf-8"/>
<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>
<link href="https://images.jonathan-harris.online" rel="preconnect"/>
<link href="https://assets.jonathan-harris.online" rel="preconnect"/>
<meta content="width=device-width, initial-scale=1, viewport-fit=cover" name="viewport"/>
<title>AI Topics | Jonathan Harris</title>
<meta content="{description}" name="description"/>
<meta content="index,follow" name="robots"/>
<meta content="#0D1420" name="theme-color"/>
{SHARED_INTER_FONT_HEAD_BLOCK}
<link as="style" href="/assets/css/site.css" rel="preload"/>
<link href="/assets/css/site.css" rel="stylesheet"/>


<link href="/assets/css/ebook-template.css" rel="stylesheet"/>
<meta content="GB" name="geo.region"/>
<meta content="website" property="og:type"/>
<meta content="{canonical}" property="og:url"/>
<meta content="AI Topics | Jonathan Harris" property="og:title"/>
<meta content="{description}" property="og:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" property="og:image"/>
<meta content="Jonathan Harris site logo" property="og:image:alt"/>
<meta content="@jonathan_harris_01" name="twitter:site"/>
<meta content="@jonathan_harris_01" name="twitter:creator"/>
<meta content="summary_large_image" name="twitter:card"/>
<meta content="AI Topics | Jonathan Harris" name="twitter:title"/>
<meta content="{description}" name="twitter:description"/>
<meta content="https://images.jonathan-harris.online/site-logo" name="twitter:image"/>
<link href="{canonical}" rel="canonical"/>
<link href="{canonical}" hreflang="en" rel="alternate"/>
<link href="{canonical}" hreflang="x-default" rel="alternate"/>
<script data-jh-ai-pack="person" type="application/ld+json">{json_script(build_person_schema())}</script>
<script data-jh-ai-pack="website" type="application/ld+json">{json_script(build_website_schema())}</script>
</head>
<body class="ebooks-catalogue topics-index">
{header}
<header class="hero ebook-hero" role="region">
  <div class="wrap">
    <img alt="Jonathan Harris site logo" class="logo-plain" height="120" src="https://images.jonathan-harris.online/site-logo" width="120"/>
    <h1>Explore AI topics</h1>
    <p>Use this page to find the right route into the site: topic guides when you want the plain-English explanation, catalogue pages when you want the books, and glossary or comparison pages when the language starts getting slippery.</p>
  </div>
</header>
<main class="main" id="main" role="main">
  <div class="wrap ebook-shell">
    <section class="card ebook-index-intro">
      <h2>How to use this page</h2>
      <p>Start with a topic guide if you want the plain-English explanation first. Use the catalogue grid if you already know the subject you care about and want the relevant books without playing hide-and-seek.</p>
    </section>
    <section class="card ebook-index-intro">
      <h2>Topic guides</h2>
      <p>These pages explain the subject before you choose a book, so you are not buying your way through a fog bank.</p>
    </section>
    <section class="grid topic-grid" aria-label="Topic guides">{guide_cards_html}</section>
    <section class="card ebook-index-intro u-mt40">
      <h2>Browse by catalogue</h2>
      <p>These category pages group the books by subject, so you can move from a broad theme to the title that actually fits.</p>
    </section>
    <section class="grid topic-grid" aria-label="Catalogue pages">{cards_html}</section>
    <section class="related-books card">
      <h2>Keep exploring the wider AI estate</h2>
      <ul>
        {topics_index_support_links()}
      </ul>
    </section>
  </div>
</main>
{footer}
<script defer="" src="/assets/js/site-ui.min.js"></script>
</body>
</html>
'''


def path_to_public_url(relative_path: Path) -> str:
    if relative_path == Path("index.html"):
        return f"{SITE_URL}/"
    if relative_path.name == "index.html":
        return f"{SITE_URL}/{relative_path.parent.as_posix()}/"
    return f"{SITE_URL}/{relative_path.as_posix()}"



def html_declares_noindex(file_path: Path) -> bool:
    if file_path.suffix.lower() != ".html" or not file_path.exists():
        return False
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    return bool(re.search(r"<meta[^>]+(?:name=[\"']robots[\"'][^>]*content=[\"'][^\"']*noindex|content=[\"'][^\"']*noindex[^>]*name=[\"']robots[\"'])", text, re.I))



def is_r2_hosted_podcast_episode_path(relative_path: Path) -> bool:
    """Return True for legacy podcast compatibility redirects only.

    Canonical generated podcast leaves under podcast/episodes/ are now governed
    by the dynamic route manifest, workbook registration, sitemap, and audit
    coverage. The only exempt family is the /podcast/TT-* compatibility shim.
    """
    parts = relative_path.parts
    return len(parts) >= 2 and parts[0] == "podcast" and parts[1].startswith("TT-")


def _first_text(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = clean_paragraph(value)
        if text:
            return text
    return ""




def brand_safe_discovery_text(value: Any) -> str:
    """Normalise discovery copy for British, no-hype GEO/AEO surfaces."""
    text = clean_paragraph(value)
    if not text:
        return ""
    replacements = [
        (r"\bdelve into\b", "examine"),
        (r"\bdelves into\b", "examines"),
        (r"\bdelving into\b", "examining"),
        (r"\blandscape\b", "field"),
        (r"\blandscapes\b", "fields"),
        (r"\bgroundbreaking\b", "notable"),
        (r"\brevolutionize\b", "change"),
        (r"\brevolutionizes\b", "changes"),
        (r"\brevolutionized\b", "changed"),
        (r"\brevolutionizing\b", "changing"),
        (r"\bpersonalized\b", "personalised"),
        (r"\bpersonalization\b", "personalisation"),
        (r"\boptimized\b", "optimised"),
        (r"\boptimizing\b", "optimising"),
        (r"\boptimization\b", "optimisation"),
        (r"\bcenter\b", "centre"),
        (r"\bcentered\b", "centred"),
        (r"\bbehavior\b", "behaviour"),
    ]
    for pattern, repl in replacements:
        text = re.sub(pattern, repl, text, flags=re.IGNORECASE)
    return clean_paragraph(text)

def _load_json_file(path: Path, default: Any) -> Any:
    try:
        if not path.exists():
            return default
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _normalise_manifest_lastmod(value: Any, fallback: str) -> str:
    text = _first_text(value)
    if not text:
        return normalise_lastmod(fallback)
    date_match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if date_match:
        return normalise_lastmod(date_match.group(0))
    return normalise_lastmod(fallback)


def _site_path_from_url(url: str) -> str:
    parsed = urlparse(url or "")
    path = parsed.path or "/"
    if not path.startswith("/"):
        path = "/" + path
    return path


def _site_url_from_path(path: str) -> str:
    clean = path if path.startswith("/") else f"/{path}"
    return f"{SITE_URL}{clean}"


def _first_party_site_url(value: Any) -> str:
    text = _first_text(value)
    if not text:
        return ""
    if text.startswith("/"):
        return _site_url_from_path(text)
    try:
        parsed = urlparse(text)
    except Exception:
        return ""
    if parsed.scheme in {"http", "https"} and parsed.netloc == urlparse(SITE_URL).netloc:
        return text
    return ""


def _manifest_route(*, family: str, title: str, path: str, lastmod: str, source: str, repo_path: str = "", summary: str = "", entity: str = "") -> Dict[str, str]:
    clean_path = path if path.startswith("/") else f"/{path}"
    return {
        "family": family,
        "title": clean_paragraph(title),
        "path": clean_path,
        "loc": _site_url_from_path(clean_path),
        "lastmod": lastmod,
        "source": source,
        "repo_path": repo_path,
        "summary": brand_safe_discovery_text(summary),
        "entity": brand_safe_discovery_text(entity),
    }


def load_blog_dynamic_routes(generated_lastmod: str) -> List[Dict[str, str]]:
    payload = _load_json_file(ROOT / "blog" / "posts.json", {})
    items = payload.get("items") if isinstance(payload, dict) else []
    routes: List[Dict[str, str]] = []
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        slug = _first_text(item.get("slug"))
        path = _first_text(item.get("path"), f"/blog/posts/{slug}/" if slug else "")
        if not slug or not path:
            continue
        routes.append(_manifest_route(
            family="blog-post",
            title=_first_text(item.get("title"), slug),
            path=path,
            lastmod=_normalise_manifest_lastmod(item.get("published_at") or item.get("datePublished") or item.get("pubDate"), generated_lastmod),
            source="runtime-r2",
            summary=_first_text(item.get("summary"), item.get("excerpt")),
            entity="weekly AI briefing",
        ))
    return routes


def load_podcast_episode_dynamic_routes(generated_lastmod: str) -> List[Dict[str, str]]:
    return []


def load_podcast_route_registry(generated_lastmod: str) -> List[Dict[str, str]]:
    return []


def load_bundle_dynamic_routes(generated_lastmod: str) -> List[Dict[str, str]]:
    path = DATA_DIR / "ebook-bundles.json"
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    bundles = payload.get("bundles", []) if isinstance(payload, dict) else []
    routes = [_manifest_route(
        family="bundle-index",
        title="Curated AI eBook reading paths",
        path="/bundles/",
        lastmod=normalise_lastmod(generated_lastmod),
        source="generated-html",
        repo_path="bundles/index.html",
        summary="Curated three-book AI reading paths from the Jonathan Harris catalogue.",
        entity="book collection",
    )]
    for bundle in bundles if isinstance(bundles, list) else []:
        if not isinstance(bundle, dict):
            continue
        slug = clean_paragraph(bundle.get("slug", ""))
        title = clean_paragraph(bundle.get("title", ""))
        if not slug or not title:
            continue
        routes.append(_manifest_route(
            family="book-bundle",
            title=title,
            path=f"/bundles/{slug}/",
            lastmod=normalise_lastmod(generated_lastmod),
            source="generated-html",
            repo_path=f"bundles/{slug}/index.html",
            summary=clean_paragraph(bundle.get("summary", "")),
            entity="book collection",
        ))
    return routes


def load_book_preview_dynamic_routes(books: List[Dict[str, Any]], generated_lastmod: str) -> List[Dict[str, str]]:
    """Register generated sample chapters for release governance."""
    routes: List[Dict[str, str]] = []
    for book in books:
        slug = clean_paragraph(book.get("slug", ""))
        title = clean_paragraph(book.get("title", ""))
        if not slug or not title:
            continue
        repo_path = f"ebooks/{slug}/sample/index.html"
        sample_file = ROOT / repo_path
        if not sample_file.exists() or html_declares_noindex(sample_file):
            continue
        routes.append(_manifest_route(
            family="book-preview",
            title=f"Sample chapter: {title}",
            path=f"/ebooks/{slug}/sample/",
            lastmod=normalise_lastmod(book.get("dateModified") or book.get("datePublished") or generated_lastmod),
            source="generated-html",
            repo_path=repo_path,
            summary="Genuine manuscript chapter sample for the published eBook page.",
            entity="book sample chapter",
        ))
    return routes


def load_static_discovery_routes(generated_lastmod: str) -> List[Dict[str, str]]:
    candidates = [
        ("site-home", "Jonathan Harris AI ecosystem", "/", "index.html", "Homepage for books, podcast, blog, topics and newsletter.", "person and site"),
        ("person", "Jonathan Harris biography", "/bio/", "bio/index.html", "Author profile and ecosystem authority page.", "person"),
        ("blog-hub", "AI blog hub", "/blog/", "blog/index.html", "Weekly AI analysis and editorial archive.", "blog"),
        ("podcast-hub", "Turing's Torch AI Weekly", "/podcast/", "podcast/index.html", "Podcast hub for AI weekly episodes.", "podcast series"),
        ("transcript-archive", "Podcast transcript archive", "/transcripts/", "transcripts/index.html", "Searchable transcript archive for podcast episodes.", "transcripts"),
        ("topic-index", "AI topic guides", "/topics/", "topics/index.html", "Topic-led entry points into the Jonathan Harris AI library.", "topic index"),
        ("glossary", "AI glossary", "/glossary/", "glossary/index.html", "Plain-English AI glossary for answer engines and readers.", "glossary"),
        ("comparison", "AI book comparison guide", "/compare/", "compare/index.html", "Comparison page for choosing relevant AI books.", "comparison"),
        ("newsletter", "AI Edge", "/newsletter/", "newsletter/index.html", "Sign-up page for AI Edge, the practical AI briefing.", "newsletter"),
        ("lead-magnet", "AI glossary cheat sheet", "/downloads/ai-glossary-cheat-sheet/", "downloads/ai-glossary-cheat-sheet/index.html", "Plain-English AI glossary cheat sheet offered as the newsletter lead magnet.", "lead magnet"),
        ("book-finder", "Find the right AI book", "/book-finder/", "book-finder/index.html", "A deterministic book finder based on reader problem and topic.", "book finder"),
        ("evidence-index", "AI evidence guides", "/evidence/", "evidence/index.html", "Source-backed AI evidence guides designed for useful retrieval and citation.", "evidence index"),
        ("resource-index", "Practical AI checklists", "/resources/", "resources/index.html", "Practical AI checklists and decision resources.", "resource index"),
        ("methodology", "Editorial and evidence methodology", "/methodology/", "methodology/index.html", "How sources are selected, claims are checked, review dates are handled and corrections are made.", "editorial methodology"),
        ("teams", "Practical AI for teams", "/for-teams/", "for-teams/index.html", "AI literacy, reading paths and practical briefing options for managers and teams.", "team AI literacy"),
        ("media", "Media and speaking", "/media/", "media/index.html", "Media, podcast guest and speaking background for Jonathan Harris on practical artificial intelligence.", "media enquiries"),
        ("contributor", "Contribute an AI case study", "/contribute/", "contribute/index.html", "Evidence-first route for submitting a real AI deployment, result or failure for editorial review.", "case study contribution"),
    ]
    routes: List[Dict[str, str]] = []
    for family, title, path, repo_path, summary, entity in candidates:
        if repo_path and not (ROOT / repo_path).exists():
            continue
        routes.append(_manifest_route(
            family=family,
            title=title,
            path=path,
            lastmod="",
            source="repo-html",
            repo_path=repo_path,
            summary=summary,
            entity=entity,
        ))
    return routes


def load_growth_asset_routes(generated_lastmod: str) -> List[Dict[str, str]]:
    routes: List[Dict[str, str]] = []
    for family, data_file, folder in (("evidence-guide", DATA_DIR / "evidence-content.json", "evidence"), ("practical-resource", DATA_DIR / "resource-content.json", "resources")):
        payload = _load_json_file(data_file, {})
        items = payload.get("items", []) if isinstance(payload, dict) else []
        for item in items if isinstance(items, list) else []:
            if not isinstance(item, dict):
                continue
            slug = _first_text(item.get("slug"))
            title = _first_text(item.get("title"), slug)
            if not slug:
                continue
            repo_path = f"{folder}/{slug}/index.html"
            if not (ROOT / repo_path).exists():
                continue
            routes.append(_manifest_route(
                family=family,
                title=title,
                path=f"/{folder}/{slug}/",
                lastmod=_normalise_manifest_lastmod(item.get("last_reviewed"), generated_lastmod),
                source=data_file.relative_to(ROOT).as_posix(),
                repo_path=repo_path,
                summary=_first_text(item.get("summary")),
                entity="AI evidence guide" if family == "evidence-guide" else "AI checklist",
            ))
    return routes


def build_dynamic_route_entries(books: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    generated_lastmod = normalise_lastmod(governed_generated_utc(books))
    routes = [
        *load_static_discovery_routes(generated_lastmod),
        *load_bundle_dynamic_routes(generated_lastmod),
        *load_book_preview_dynamic_routes(books, generated_lastmod),
        *load_growth_asset_routes(generated_lastmod),
        *load_blog_dynamic_routes(generated_lastmod),
        *load_podcast_episode_dynamic_routes(generated_lastmod),
        *load_podcast_route_registry(generated_lastmod),
    ]
    unique: Dict[str, Dict[str, str]] = {}
    for route in routes:
        loc = route.get("loc")
        if loc:
            unique[loc] = route
    return sorted(unique.values(), key=lambda item: (item.get("family", ""), item.get("path", "")))


def build_dynamic_route_manifest(books: List[Dict[str, Any]]) -> Dict[str, Any]:
    routes = build_dynamic_route_entries(books)
    families = Counter(route.get("family", "unknown") for route in routes)
    return {
        "generated_utc": governed_generated_utc(books),
        "base_url": SITE_URL,
        "purpose": "Governed route ledger for generated conversion pages plus blog, podcast, transcript, and LLM discovery surfaces.",
        "route_count": len(routes),
        "families": dict(sorted(families.items())),
        "routes": routes,
    }


def build_search_visibility_surfaces(books: List[Dict[str, Any]]) -> Dict[str, Any]:
    routes = build_dynamic_route_entries(books)
    return {
        "generated_utc": governed_generated_utc(books),
        "lane": "Lane 1 autonomous evidence",
        "seo": {
            "sitemap_includes_dynamic_routes": True,
            "transcript_urls": sum(1 for route in routes if route.get("family") == "podcast-transcript"),
            "podcast_episode_urls": sum(1 for route in routes if route.get("family") == "podcast-episode"),
            "blog_post_urls": sum(1 for route in routes if route.get("family") == "blog-post"),
        },
        "geo": {
            "llms_scope": "full-estate",
            "discovery_families": sorted({route.get("family", "unknown") for route in routes}),
        },
        "aeo": {
            "answer_led_podcast_templates": True,
            "transcript_archive_exposed": any(route.get("family") == "transcript-archive" for route in routes),
        },
    }


def is_site_shell_artifact_path(relative_path: Path) -> bool:
    """Return True for published Site Shell implementation artefacts.

    These files are intentionally fetchable by AIMS but are not standalone
    public pages and therefore must stay out of sitemap/route discovery.
    """
    parts = relative_path.parts
    return len(parts) >= 2 and parts[0] == "assets" and parts[1] == "site-shell"


def build_public_route_registry(books: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    governed_lastmod = normalise_lastmod(governed_generated_utc(books))
    book_paths = {Path("ebooks") / book["slug"] / "index.html": book for book in books}
    excluded_paths = {
        Path("404.html"),
        Path("assets/partials/header.html"),
        Path("assets/partials/footer.html"),
    }

    by_loc: Dict[str, Dict[str, str]] = {}
    for file_path in sorted(ROOT.rglob("*.html")):
        if "node_modules" in file_path.parts:
            continue
        relative_path = file_path.relative_to(ROOT)
        rel_parts = relative_path.parts
        if relative_path in excluded_paths:
            continue
        if is_r2_hosted_podcast_episode_path(relative_path):
            continue
        if rel_parts and rel_parts[0] in {"scripts", "functions"}:
            continue
        # Site Shell header/footer files are published implementation fragments
        # for AIMS-managed R2 content. They are fetchable assets, not public
        # website routes, and must never enter sitemap/crawler route discovery.
        if is_site_shell_artifact_path(relative_path):
            continue
        if html_declares_noindex(file_path):
            continue

        if relative_path in book_paths:
            lastmod = normalise_lastmod(book_paths[relative_path].get("dateModified") or book_paths[relative_path].get("datePublished") or governed_lastmod)
        else:
            lastmod = html_significant_lastmod(file_path)

        loc = path_to_public_url(relative_path)
        by_loc[loc] = {
            "path": f"/{relative_path.as_posix()}",
            "loc": loc,
            "lastmod": lastmod,
        }

    for route in build_dynamic_route_entries(books):
        loc = route.get("loc")
        if not loc:
            continue
        route_lastmod = clean_paragraph(route.get("lastmod", ""))
        if loc in by_loc:
            # A governed route date (book update, evidence review, blog/podcast
            # publication) is stronger than an HTML filesystem timestamp.
            if route_lastmod:
                by_loc[loc]["lastmod"] = route_lastmod
            continue
        by_loc[loc] = {
            "path": route.get("path", _site_path_from_url(loc)),
            "loc": loc,
            "lastmod": route_lastmod,
        }

    return [by_loc[loc] for loc in sorted(by_loc)]


def build_sitemap_xml(books: List[Dict[str, Any]]) -> str:
    urls = []
    for route in build_public_route_registry(books):
        lastmod = clean_paragraph(route.get("lastmod", ""))
        lastmod_xml = f"    <lastmod>{html.escape(lastmod)}</lastmod>\n" if lastmod else ""
        urls.append(
            "  <url>\n"
            f"    <loc>{html.escape(route['loc'])}</loc>\n"
            f"{lastmod_xml}"
            "  </url>"
        )
    return "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n<urlset xmlns=\"http://www.sitemaps.org/schemas/sitemap/0.9\">\n" + "\n".join(urls) + "\n</urlset>\n"



def build_robots_txt() -> str:
    return "\n".join([
        "# Robots rules (generated source snapshot)",
        f"# Canonical publication target: {EXTERNAL_CRAWLER_FILES['robots']}",
        f"# Canonical sitemap target: {EXTERNAL_CRAWLER_FILES['sitemap']}",
        "",
        "User-agent: *",
        "Allow: /",
        "",
        "# Explicit AI crawler allowances",
        "User-agent: GPTBot",
        "Allow: /",
        "User-agent: Google-Extended",
        "Allow: /",
        "User-agent: ClaudeBot",
        "Allow: /",
        "User-agent: CCBot",
        "Allow: /",
        "User-agent: Applebot",
        "Allow: /",
        "",
        "# Sitemap",
        f"Sitemap: {EXTERNAL_CRAWLER_FILES['sitemap']}",
        "",
    ])



def build_llms_txt(books: List[Dict[str, Any]]) -> str:
    dynamic_routes = build_dynamic_route_entries(books)
    routes_by_family: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for route in dynamic_routes:
        routes_by_family[route.get("family", "unknown")].append(route)

    lines = [
        "# Jonathan Harris AI ecosystem",
        f"# Canonical publication target: {EXTERNAL_CRAWLER_FILES['llms']}",
        "# Scope: full estate, including books, topics, blog, podcast episodes and transcript discovery surfaces.",
        f"Homepage: {SITE_URL}/",
        f"Author: {SITE_URL}/bio/",
        f"Books: {SITE_URL}/ebooks/",
        f"Blog: {SITE_URL}/blog/",
        f"Podcast: {SITE_URL}/podcast/",
        f"Transcripts: {SITE_URL}/transcripts/",
        f"Topics: {SITE_URL}/topics/",
        f"Glossary: {SITE_URL}/glossary/",
        "",
        "## Core discovery surfaces",
    ]
    for route in routes_by_family.get("site-home", []) + routes_by_family.get("person", []) + routes_by_family.get("topic-index", []) + routes_by_family.get("glossary", []) + routes_by_family.get("comparison", []):
        lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    lines.extend(["", "## Canonical books"])
    for book in books:
        lines.append(f"- {book['title']}: {book['canonical_url']} — {brand_safe_discovery_text(book['short'])}")

    service_routes = routes_by_family.get("methodology", []) + routes_by_family.get("teams", []) + routes_by_family.get("media", []) + routes_by_family.get("contributor", []) + routes_by_family.get("book-finder", []) + routes_by_family.get("lead-magnet", [])
    if service_routes:
        lines.extend(["", "## Methodology and practical routes"])
        for route in service_routes:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    evidence_routes = routes_by_family.get("evidence-guide", []) + routes_by_family.get("practical-resource", [])
    if evidence_routes:
        lines.extend(["", "## Evidence guides and practical resources"])
        for route in evidence_routes:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    bundle_routes = routes_by_family.get("book-bundle", [])
    if bundle_routes:
        lines.extend(["", "## Curated eBook reading paths"])
        for route in bundle_routes:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    blog_routes = routes_by_family.get("blog-post", [])
    if blog_routes:
        lines.extend(["", "## Blog and weekly AI analysis"])
        for route in blog_routes[:20]:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    podcast_routes = routes_by_family.get("podcast-episode", [])
    if podcast_routes:
        lines.extend(["", "## Podcast episodes"])
        for route in podcast_routes[:40]:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    transcript_routes = routes_by_family.get("podcast-transcript", [])
    if transcript_routes:
        lines.extend(["", "## Podcast transcripts"])
        for route in transcript_routes[:40]:
            lines.append(f"- {route['title']}: {route['loc']} — {brand_safe_discovery_text(route.get('summary', ''))}")

    lines.append("")
    return "\n".join(lines)


def build_crawler_snapshot_payloads(books: List[Dict[str, Any]]) -> Dict[str, str]:
    return {
        CRAWLER_SNAPSHOT_FILENAMES["robots"]: build_robots_txt(),
        CRAWLER_SNAPSHOT_FILENAMES["sitemap"]: build_sitemap_xml(books),
        CRAWLER_SNAPSHOT_FILENAMES["llms"]: build_llms_txt(books),
    }



def build_crawler_snapshot_paths(books: List[Dict[str, Any]]) -> Dict[Path, str]:
    payloads = build_crawler_snapshot_payloads(books)
    return {
        CRAWLER_SNAPSHOTS_DIR / name: content
        for name, content in payloads.items()
    }



def build_published_crawler_paths(books: List[Dict[str, Any]]) -> Dict[Path, str]:
    payloads = build_crawler_snapshot_payloads(books)
    robots_payload = payloads[CRAWLER_SNAPSHOT_FILENAMES["robots"]]
    sitemap_payload = payloads[CRAWLER_SNAPSHOT_FILENAMES["sitemap"]]
    llms_payload = payloads[CRAWLER_SNAPSHOT_FILENAMES["llms"]]
    return {
        ROOT / "robots.txt": robots_payload,
        ROOT / "robot.txt": robots_payload,
        ROOT / "sitemap.xml": sitemap_payload,
        ROOT / "llms.txt": llms_payload,
    }




def build_route_manifest(books: List[Dict[str, Any]]) -> Dict[str, Any]:
    manifest_books = []
    for book in books:
        manifest_books.append({
            "slug": book["slug"],
            "title": book["title"],
            "canonical": {
                "path": f"/ebooks/{book['slug']}/",
                "url": book["canonical_url"],
            },
            "buy": {
                "path": book["buy_route"],
                "full_url": book["buy_route_full"],
                "target": book["buy_url"],
            },
            "topic_page": book["topic_url"],
            "legacy_routes": [
                f"/book/{book['slug']}/",
                f"/book/{book['slug']}/buy-now",
                f"/ebooks/{book['slug']}/detail",
                f"/ebooks/{book['slug']}/detail.html",
                f"/ebooks/{book['slug']}/details.html",
            ],
            "legacy_alias_url": book.get("legacy_alias_url") or "",
            "domain_redirect_families": [
                "books.jonathan-harris.online",
                "ebooks.jonathan-harris.online",
            ],
        })

    return {
        "generated_utc": governed_generated_utc(books),
        "base_url": SITE_URL,
        "ebook_count": len(books),
        "external_crawler_files": EXTERNAL_CRAWLER_FILES,
        "host_redirect_files": {
            "books_domain": "_redirects.books-domain",
            "ebooks_domain": "_redirects.ebooks-domain",
        },
        "malformed_slug_fixes": MALFORMED_SLUG_FIXES,
        "ebooks": manifest_books,
        "dynamic_routes": build_dynamic_route_entries(books),
    }


def build_books_domain_redirects() -> str:
    return "\n".join([
        "# ============================================================",
        "# books.jonathan-harris.online — Domain redirects (generated)",
        "# Redirect governed legacy book paths to the main domain",
        "# ============================================================",
        "",
        "# Canonical host (no www)",
        "https://www.books.jonathan-harris.online/*  https://books.jonathan-harris.online/:splat  301",
        "",
        "# Root → ebooks catalogue",
        "/  https://jonathan-harris.online/ebooks/  301",
        "",
        "# Canonical and legacy ebook families",
        "/ebooks          https://jonathan-harris.online/ebooks/                  301",
        "/ebooks/         https://jonathan-harris.online/ebooks/                  301",
        "/ebooks/*        https://jonathan-harris.online/ebooks/:splat           301",
        "/book            https://jonathan-harris.online/ebooks/                  301",
        "/book/           https://jonathan-harris.online/ebooks/                  301",
        "/book/*          https://jonathan-harris.online/ebooks/:splat           301",
        "/category/*      https://jonathan-harris.online/catalogue/:splat        301",
        "/catalogue/*     https://jonathan-harris.online/catalogue/:splat        301",
        "/topics/*        https://jonathan-harris.online/topics/:splat           301",
        "/author          https://jonathan-harris.online/bio/                    301",
        "/author/         https://jonathan-harris.online/bio/                    301",
        "/author/*        https://jonathan-harris.online/bio/                    301",
        "",
        "# Crawler files resolve on the main domain",
        f"/robots.txt     {EXTERNAL_CRAWLER_FILES['robots']}   301",
        f"/sitemap.xml    {EXTERNAL_CRAWLER_FILES['sitemap']}  301",
        f"/site-map.xml   {EXTERNAL_CRAWLER_FILES['sitemap']}  301",
        "",
        "# Catch-all",
        "/*  https://jonathan-harris.online/:splat  301",
        "",
    ])


def build_ebooks_domain_redirects() -> str:
    return "\n".join([
        "# ============================================================",
        "# ebooks.jonathan-harris.online — Domain redirects (generated)",
        "# Redirect governed ebook and discovery paths to the main domain",
        "# ============================================================",
        "",
        "# Canonical host (no www)",
        "https://www.ebooks.jonathan-harris.online/*  https://ebooks.jonathan-harris.online/:splat  301",
        "",
        "# Root → ebooks catalogue",
        "/          https://jonathan-harris.online/ebooks/  301",
        "/ebooks    https://jonathan-harris.online/ebooks/  301",
        "/ebooks/   https://jonathan-harris.online/ebooks/  301",
        "",
        "# Old paths → new site",
        "/book         https://jonathan-harris.online/ebooks/             301",
        "/book/        https://jonathan-harris.online/ebooks/             301",
        "/book/*       https://jonathan-harris.online/ebooks/:splat       301",
        "/catalogue/*  https://jonathan-harris.online/catalogue/:splat    301",
        "/category/*   https://jonathan-harris.online/catalogue/:splat    301",
        "/author       https://jonathan-harris.online/bio/                301",
        "/author/      https://jonathan-harris.online/bio/                301",
        "/author/*     https://jonathan-harris.online/bio/                301",
        "/topics/*     https://jonathan-harris.online/topics/:splat       301",
        "/glossary/*   https://jonathan-harris.online/glossary/:splat     301",
        "/api/*        https://jonathan-harris.online/api/:splat          301",
        "",
        "# Crawler files resolve on the main domain",
        f"/robots.txt   {EXTERNAL_CRAWLER_FILES['robots']}  301",
        f"/sitemap.xml  {EXTERNAL_CRAWLER_FILES['sitemap']}  301",
        f"/site-map.xml {EXTERNAL_CRAWLER_FILES['sitemap']}  301",
        "",
        "# Catch-all",
        "/*  https://jonathan-harris.online/:splat  301",
        "",
    ])


def build_derivatives(books: List[Dict[str, Any]]) -> None:
    # Compatibility sitemap names are redirects/functions, never duplicate physical
    # canonical XML bodies. Remove stale production artefacts before regeneration.
    for alias in ("Sitemap.xml", "site-map.xml", "sitemap (1).xml"):
        alias_path = ROOT / alias
        if alias_path.exists():
            alias_path.unlink()
    public_records = [book_to_public_record(book) for book in books]
    write_json(EBOOKS_DIR / "books.json", public_records)
    write_json(ROOT / "assets" / "js" / "books.json", public_records)
    write_json(ROOT / "api" / "v1" / "books.json", public_records)
    write_json(ROOT / "api" / "v1" / "featured-book.json", build_featured_book_payload(public_records))

    feed = {
        "version": "https://jsonfeed.org/version/1.1",
        "title": "Jonathan Harris eBooks Catalogue",
        "home_page_url": f"{SITE_URL}/ebooks/",
        "feed_url": f"{SITE_URL}/feed.json",
        "items": [
            {
                "id": book["canonical_url"],
                "url": book["canonical_url"],
                "title": book["title"],
                "summary": book["short"],
                "image": book["cover"],
                "tags": book["tags"],
            }
            for book in books
        ],
    }
    write_json(ROOT / "feed.json", feed)

    generated_utc = governed_generated_utc(books)

    entity_map = {
        "generated_utc": generated_utc,
        "person": {"id": f"{SITE_URL}/#person", "name": SITE_NAME},
        "podcast": {"title": "Turing’s Torch: AI Weekly", "url": f"{SITE_URL}/podcast/"},
        "books": [
            {
                "title": book["title"],
                "slug": book["slug"],
                "url": book["canonical_url"],
                "topic": book["topic"],
                "tags": book["tags"],
                "identifier": book["identifier"],
            }
            for book in books
        ],
    }
    write_json(ROOT / "ai" / "entity-map.json", entity_map)

    llm_index = {
        "generated_utc": generated_utc,
        "books": [
            {
                "title": book["title"],
                "slug": book["slug"],
                "url": f"/ebooks/{book['slug']}/",
                "summary": book["short"],
                "topic": book["topic"],
                "tags": book["tags"],
                "keywords": [clean_paragraph(k).lower() for k in book["keywords"]],
                "buy_url": book["buy_url"],
                "buy_target_url": book["buy_url"],
                "buy_route": book["buy_route"],
                "buy_route_full": book["buy_route_full"],
                "cover": book["cover"],
                "identifier": book["identifier"],
                "entity_id": f"{book['canonical_url']}#book",
                "asin": book["asin"],
                "pages": book["pages"],
                "datePublished": book["datePublished"],
                "dateModified": book.get("dateModified") or infer_build_timestamp(),
                "related_slugs": book.get("related_slugs", []),
            }
            for book in books
        ],
        "topic_authority_pages": [f"/catalogue/{slugify(topic)}/" for topic in sorted({book['topic'] for book in books})],
        "site_sections": [route for route in build_dynamic_route_entries(books) if route.get("family") in {"site-home", "person", "blog-hub", "podcast-hub", "transcript-archive", "topic-index", "glossary", "comparison", "newsletter", "lead-magnet", "book-finder", "evidence-index", "resource-index", "methodology", "teams", "media", "contributor"}],
        "evidence_guides": [route for route in build_dynamic_route_entries(books) if route.get("family") == "evidence-guide"],
        "practical_resources": [route for route in build_dynamic_route_entries(books) if route.get("family") == "practical-resource"],
        "blog_posts": [route for route in build_dynamic_route_entries(books) if route.get("family") == "blog-post"],
        "podcast_episodes": [route for route in build_dynamic_route_entries(books) if route.get("family") == "podcast-episode"],
        "transcripts": [route for route in build_dynamic_route_entries(books) if route.get("family") == "podcast-transcript"],
    }
    write_json(ROOT / "llm-index.json", llm_index)
    write_json(DYNAMIC_ROUTE_MANIFEST_PATH, build_dynamic_route_manifest(books))
    write_json(SEARCH_VISIBILITY_SURFACES_PATH, build_search_visibility_surfaces(books))

    write_json(EBOOKS_DIR / "url-manifest.json", build_route_manifest(books))
    (EBOOKS_DIR / "index.html").write_text(render_ebooks_index(books), encoding="utf-8")

    topic_map: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for book in books:
        topic_map[book["topic"]].append(book)
    for topic, topic_books in topic_map.items():
        topic_dir = CATALOGUE_DIR / slugify(topic)
        topic_dir.mkdir(parents=True, exist_ok=True)
        (topic_dir / "index.html").write_text(render_topic_page(topic, topic_books), encoding="utf-8")
    TOPICS_DIR.mkdir(parents=True, exist_ok=True)
    (TOPICS_DIR / "index.html").write_text(render_topics_index(topic_map), encoding="utf-8")

    for file_path, content in build_crawler_snapshot_paths(books).items():
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(content, encoding="utf-8")
    for file_path, content in build_published_crawler_paths(books).items():
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(content, encoding="utf-8")

    for legacy_path in (
        ROOT / "Sitemap.xml",
        ROOT / "site-map.xml",
        ROOT / "sitemap (1).xml",
        CRAWLER_SNAPSHOTS_DIR / "site-map.xml",
    ):
        try:
            if legacy_path.exists():
                legacy_path.unlink()
        except IsADirectoryError:
            pass

    write_json(CRAWLER_CHECKSUMS_PATH, build_crawler_checksums(books))



def build_book_files(books: List[Dict[str, Any]]) -> None:
    samples = load_book_sample_chapters()
    for book in books:
        book_dir = EBOOKS_DIR / book["slug"]
        book_dir.mkdir(parents=True, exist_ok=True)
        sample_dir = book_dir / "sample"
        sample = samples.get(book["slug"], {})
        sample_available = bool(
            isinstance(sample, dict)
            and sample.get("paragraphs")
            and int(sample.get("word_count") or 0) >= 350
        )
        if sample_available:
            sample_dir.mkdir(parents=True, exist_ok=True)
            (sample_dir / "index.html").write_text(render_book_sample_page(book), encoding="utf-8")
        elif sample_dir.exists():
            # A missing manuscript extraction must never leave a stale placeholder route behind.
            shutil.rmtree(sample_dir)
        metadata = {
            "title": book["title"],
            "slug": book["slug"],
            "description": book["description"],
            "short_description": book["short"],
            "topic": book["topic"],
            "tags": book["tags"],
            "cover": book["cover"],
            "image": book["cover"],
            "buy_url": book["buy_url"],
            "buy_route": book["buy_route"],
            "pages": book["pages"],
            "asin": book["asin"],
            "datePublished": book["datePublished"],
        "dateModified": book.get("dateModified") or infer_build_timestamp(),
            "author": book["author"],
            "tone": book["tone"],
            "audience": book["audience"],
            "canonical_url": book["canonical_url"],
            "identifier": book["identifier"],
        }
        faq_schema = {
            "@context": "https://schema.org",
            "@type": "FAQPage",
            "@id": f"{book['canonical_url']}#faq",
            "mainEntity": book["faq"],
        }
        summary_md = "\n".join([
            f"# {book['title']}",
            "",
            "## Summary",
            "",
            book["summary"],
            "",
            "## What this book covers",
            "",
            book["what_this_book_covers"],
            "",
            "## Who this book is for",
            "",
            book["who_for"],
            "",
            "## What you’ll learn",
            "",
            *[f"- {item}" for item in book["what_youll_learn"]],
            "",
            "## Why it matters",
            "",
            book["why_it_matters"],
            "",
            "## Topics and tags",
            "",
            f"- Topic: {book['topic']}",
            f"- Tags: {', '.join(book['tags'])}",
            f"- Length: {book['pages']} pages",
            "",
            "## Buy",
            "",
            f"- {book['buy_url']}",
            "",
        ])
        write_json(book_dir / "metadata.json", metadata)
        write_json(book_dir / "faq-schema.json", faq_schema)
        (book_dir / "structured-summary.md").write_text(summary_md, encoding="utf-8")
        (book_dir / "index.html").write_text(render_book_page(book, books), encoding="utf-8")



def build_redirect_block(books: List[Dict[str, Any]]) -> str:
    lines = ["# 6A) Branded buy-now redirects"]
    for book in books:
        lines.append(f"{book['buy_route']}   {book['buy_url']}   302")
        lines.append(f"/book/{book['slug']}/buy-now   {book['buy_route']}   301")
    return "\n".join(lines)



def sync_redirects(books: List[Dict[str, Any]]) -> None:
    primary_path = ROOT / "_redirects"
    mirror_path = ROOT / "_redirects.txt"
    block = build_redirect_block(books)
    pattern = re.compile(r"# 6A\) Branded buy-now redirects.*?(?=\n# 7\) CANONICAL: old /book URLs permanently redirect to /ebooks)", re.S)
    legacy_anchor = "# retire legacy detail routes to the canonical ebook page\n"
    malformed_lines = [f"{item['source']}   {item['target']}  301" for item in MALFORMED_SLUG_FIXES]
    crawler_alias_anchor = "# SEO files: serve governed crawler assets from the primary domain root\n"
    crawler_alias_lines = [
        "/robot.txt    /robots.txt   301",
        "/Sitemap.xml  /sitemap.xml  301",
        "/site-map.xml  /sitemap.xml  301",
    ]
    crawler_alias_pattern = re.compile(r"# SEO files: serve governed crawler assets from the primary domain root\n(?:[^#].*\n?)*", re.M)

    text = primary_path.read_text(encoding="utf-8")
    new_text, count = pattern.subn(block + "\n", text)
    if count != 1:
        raise ValueError(f"Could not replace branded redirect block in {primary_path}")
    for line in LEGACY_DETAIL_REDIRECT_LINES + malformed_lines:
        if line not in new_text:
            if legacy_anchor not in new_text:
                raise ValueError(f"Could not find legacy detail redirect anchor in {primary_path}")
            new_text = new_text.replace(legacy_anchor, legacy_anchor + line + "\n", 1)
    alias_block = crawler_alias_anchor + "\n".join(crawler_alias_lines) + "\n"
    if crawler_alias_anchor not in new_text:
        raise ValueError(f"Could not find crawler alias anchor in {primary_path}")
    new_text, alias_count = crawler_alias_pattern.subn(alias_block, new_text, count=1)
    if alias_count != 1:
        raise ValueError(f"Could not refresh crawler alias block in {primary_path}")

    primary_path.write_text(new_text, encoding="utf-8")
    mirror_path.write_text(new_text, encoding="utf-8")
    (ROOT / "_redirects.books-domain").write_text(build_books_domain_redirects(), encoding="utf-8")
    (ROOT / "_redirects.ebooks-domain").write_text(build_ebooks_domain_redirects(), encoding="utf-8")


def ensure_css_file() -> None:
    css = """
.ebook-shell{display:grid;gap:20px}
.ebook-hero .muted{opacity:.88}
.ebook-section h2,.related-books h2,.faq h2,.ebook-index-intro h2{margin-top:0}
.quick-facts ul.ebook-facts{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px 20px;list-style:none;padding:0;margin:0}
.quick-facts li{color:#374151}
.ebook-showcase{display:grid;grid-template-columns:minmax(280px,360px) minmax(0,1fr);gap:20px;align-items:start}
.ebook-showcase__media,.ebook-showcase__content{height:100%}
.ebook-showcase__cover{background:#0D1420}
.ebook-showcase__lead{font-size:1.03rem;line-height:1.75;color:#111827}
.ebook-showcase__subhead,.ebook-showcase__note{color:#4B5563}
.ebook-actions{display:flex;flex-wrap:wrap;gap:10px;margin-top:16px}
.ebook-inline-actions{display:flex;flex-wrap:wrap;gap:12px;margin-top:18px}
.ebook-inline-actions a{display:inline-flex;align-items:center;justify-content:center;min-height:42px;padding:10px 14px;border-radius:999px;text-decoration:none;font-weight:700;border:1px solid rgba(15,23,42,.12);background:#fff;color:#111827}
.ebook-signal-list{list-style:none;padding:0;margin:16px 0 0;display:grid;gap:10px}
.ebook-signal-list li{color:#111827;font-weight:600}
.ebook-theme-pills{display:flex;flex-wrap:wrap;gap:8px;margin-top:14px}
.ebook-pill{display:inline-flex;align-items:center;padding:6px 10px;border-radius:999px;background:#EEF2FF;border:1px solid rgba(79,70,229,.18);color:#312E81;font-size:.9rem;font-weight:700}
.ebook-pill--dark{background:#111827;color:#E5E7EB;border-color:rgba(255,255,255,.12)}
.ebook-learn-list,.ebook-audience-list,.ebook-key-themes{margin:0;padding-left:20px;display:grid;gap:10px}
.ebook-section--accent{background:linear-gradient(180deg,#111827 0%,#0F172A 100%);color:#E5E7EB;border-color:rgba(255,255,255,.08)}
.ebook-section--accent h2{color:#fff}
.ebook-section--accent p,.ebook-section--accent li{color:#E5E7EB}
.related-books ul{list-style:none;padding:0;margin:0;display:grid;gap:12px}
.related-books li{display:flex;justify-content:space-between;gap:16px;align-items:baseline;padding:12px 0;border-bottom:1px solid rgba(15,23,42,.08)}
.related-books li:last-child{border-bottom:none;padding-bottom:0}
.related-books li span{font-size:.92rem;color:#6B7280}
.ebook-faq-list{display:grid;gap:12px}
.ebook-faq-item{border:1px solid rgba(15,23,42,.10);border-radius:14px;padding:0 14px;background:#F8FAFC}
.ebook-faq-item summary{cursor:pointer;font-weight:800;padding:14px 0;color:#111827}
.ebook-faq-item p{margin:0 0 14px;color:#374151}
.ebook-index-intro p{max-width:72ch}
.ebook-catalogue-hero{padding-top:38px;padding-bottom:30px;text-align:center}
.ebook-catalogue-hero h1{margin-bottom:8px}
.ebook-catalogue-hero p{max-width:760px}
.ebook-catalogue-controls{margin:0 0 24px;padding:22px 24px;border:1px solid rgba(15,23,42,.10);border-radius:18px;background:#fff;box-shadow:0 10px 28px rgba(15,23,42,.07)}
.ebook-catalogue-controls__intro{display:flex;align-items:flex-start;justify-content:space-between;gap:24px}
.ebook-catalogue-controls__intro h2{margin:0 0 6px;color:#111827;font-size:1.35rem;line-height:1.25}
.ebook-catalogue-controls__intro p:not(.ebook-count){margin:0;color:#4B5563;max-width:66ch}
.ebook-catalogue-toolbar{display:grid;gap:10px;justify-content:stretch;margin:18px 0 0}
.ebook-search-label{font-size:.82rem;font-weight:800;color:#374151;letter-spacing:.01em}
.ebook-catalogue-toolbar .search{width:100%;max-width:none;box-sizing:border-box;min-height:50px}
.ebook-catalogue-toolbar .chips{justify-content:flex-start;margin:2px 0 0}
.ebook-catalogue-toolbar .chip{background:#F8FAFC;border-color:#E2E8F0;color:#334155}
.ebook-catalogue-toolbar .chip[aria-pressed="true"]{background:#EEF2FF;border-color:rgba(79,70,229,.36);color:#312E81;box-shadow:none}
.ebook-count{margin:3px 0 0;white-space:nowrap;font-weight:700;color:#64748B}
.ebook-topic-directory{margin-top:22px;padding-top:18px;border-top:1px solid rgba(15,23,42,.10)}
.ebook-topic-directory summary{cursor:pointer;font-weight:800;color:#312E81}
.ebook-topic-directory .jh-topic-links{margin-top:14px}
.pager{margin:24px 0 0;display:flex;align-items:center;justify-content:center;gap:14px}
.topic-chip-wrap{margin:2px 0 8px}
.topic-chip{display:inline-flex;align-items:center;padding:6px 10px;border-radius:999px;background:#EEF2FF;color:#312E81;font-size:.88rem;font-weight:700;border:1px solid rgba(79,70,229,.18)}
.book-avail{margin:12px 0 0}
.book-avail__badge{display:inline-flex;align-items:center;gap:8px;padding:8px 12px;border-radius:999px;background:#ECFDF5;color:#166534;border:1px solid rgba(22,101,52,.12);font-size:.88rem;font-weight:700}
.ebook-signal-grid,.topic-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:16px}
.ebook-signal-card h3,.topic-card h2{margin-top:0}
.ebook-signal-card p,.topic-card p{margin:0;color:#4B5563}
.ebooks-catalogue .grid{grid-template-columns:repeat(3,minmax(0,1fr))}
@media (max-width:980px){
  .ebook-showcase{grid-template-columns:1fr}
  .quick-facts ul.ebook-facts,.ebook-signal-grid,.topic-grid{grid-template-columns:1fr}
  .ebooks-catalogue .grid{grid-template-columns:repeat(2,minmax(0,1fr))}
}
@media (max-width:620px){
  .ebooks-catalogue .grid{grid-template-columns:1fr}
  .ebook-catalogue-hero{padding-top:30px;padding-bottom:24px}
  .ebook-catalogue-controls{padding:18px 16px}
  .ebook-catalogue-controls__intro{display:grid;gap:8px}
  .ebook-count{margin:0}
  .ebook-catalogue-toolbar .chips{gap:7px}
  .ebook-catalogue-toolbar .chip{font-size:.84rem;padding:7px 10px}
  .ebook-actions,.ebook-inline-actions{flex-direction:column}
  .ebook-actions .button,.pager .button,.ebook-inline-actions a{width:100%;text-align:center}
  .related-books li{flex-direction:column;align-items:flex-start}
}
""".strip() + "\n"
    EBOOK_TEMPLATE_CSS.write_text(css, encoding="utf-8")



def faq_is_semantically_relevant(book: Dict[str, Any]) -> bool:
    content_tokens = text_tokens(book["summary"]) | text_tokens(book["what_this_book_covers"])
    for item in book.get("what_youll_learn", []):
        content_tokens |= text_tokens(item)
    for item in book.get("faq", []):
        question = clean_paragraph(item.get("name", "")).lower()
        answer = clean_paragraph(item.get("acceptedAnswer", {}).get("text", ""))
        if any(term in question for term in ["how long", "format", "buy", "where can i get", "who is this book for", "who gets the most value", "how detailed is the coverage", "suitable for"]):
            continue
        if not (text_tokens(answer) & content_tokens):
            return False
    return True



def build_crawler_checksums(books: List[Dict[str, Any]]) -> Dict[str, Any]:
    generated_at = governed_generated_utc(books)
    files = build_crawler_snapshot_payloads(books)
    name_to_key = {value: key for key, value in CRAWLER_SNAPSHOT_FILENAMES.items()}
    return {
        "generated_utc": generated_at,
        "files": {
            name: {
                "url": EXTERNAL_CRAWLER_FILES[name_to_key[name]],
                "sha256": sha256_text(content),
            }
            for name, content in files.items()
        },
    }


def build_validation_report(
    errors: List[str],
    books: List[Dict[str, Any]],
    *,
    workbook_title_stats: Dict[str, int] | None = None,
    workbook_content_stats: Dict[str, int] | None = None,
) -> str:
    topics = Counter(book["topic"] for book in books)
    lines = [
        f"Validation run: {governed_generated_utc(books)}",
        f"Book count: {len(books)}",
        f"Topic count: {len(topics)}",
    ]
    if workbook_title_stats:
        lines.append(
            "Workbook title parity: "
            f"{workbook_title_stats['passed']}/{workbook_title_stats['checked']} passed, "
            f"{workbook_title_stats['mismatched']} mismatched"
        )
    if workbook_content_stats:
        lines.append(
            "Workbook content parity: "
            f"{workbook_content_stats['exact_matches']} exact field matches, "
            f"{workbook_content_stats['approved_transformations']} approved transformations, "
            f"{workbook_content_stats['mismatched']} mismatches across "
            f"{workbook_content_stats['checked_fields']} governed field checks"
        )
    lines.extend([
        "",
        "Topics:",
    ])
    for topic, count in sorted(topics.items(), key=lambda item: item[0].lower()):
        lines.append(f"- {topic}: {count}")
    lines.append("")
    if errors:
        lines.append(f"Status: FAILED ({len(errors)} issue(s))")
        lines.extend(f"- {error}" for error in errors)
    else:
        lines.append("Status: PASSED")
        lines.append("- Route registry complete")
        lines.append("- Route manifest and host redirect coverage checks passed")
        lines.append("- Topic discovery pages exist for every book")
        lines.append("- Metadata and FAQ checks passed")
        lines.append("- Generated crawler snapshots and redirect governance checks passed")
    lines.append("")
    return "\n".join(lines)



def workbook_content_parity_audit(
    workbook_path: Path,
    books: List[Dict[str, Any]] | None = None,
) -> Tuple[List[str], Dict[str, int]]:
    if books is None:
        books = load_master()
    stats = {
        "books": 0,
        "checked_fields": 0,
        "exact_matches": 0,
        "approved_transformations": 0,
        "mismatched": 0,
        "missing_books": 0,
    }
    errors: List[str] = []

    order, workbook_map, workbook_content = parse_workbook(workbook_path, sanitise_content=False)
    master_by_slug = {book["slug"]: book for book in books}
    approved_normalisations = load_workbook_normalisations()

    if order != [book["slug"] for book in books]:
        errors.append("Workbook row order does not match the master record order.")

    for slug, workbook in workbook_map.items():
        stats["books"] += 1
        book = master_by_slug.get(slug)
        if not book:
            stats["missing_books"] += 1
            errors.append(f"Workbook contains slug not found in master record: {slug}")
            continue
        for workbook_field, master_field in [
            ("book_url", "canonical_url"),
            ("buy_url", "buy_url"),
            ("buy_route", "buy_route"),
            ("legacy_alias_url", "legacy_alias_url"),
            ("asin", "asin"),
            ("pages", "pages"),
            ("datePublished", "datePublished"),
            ("cover", "cover"),
        ]:
            if clean_paragraph(str(workbook.get(workbook_field, ""))) != clean_paragraph(str(book.get(master_field, ""))):
                errors.append(f"Workbook mismatch for {slug}: {workbook_field} does not match {master_field}.")

        content = workbook_content.get(slug)
        if not content:
            continue
        for workbook_field, master_field in WORKBOOK_GOVERNED_COPY_FIELDS:
            stats["checked_fields"] += 1
            workbook_value = clean_paragraph(str(content.get(workbook_field, "")))
            master_value = clean_paragraph(str(book.get(master_field, "")))
            if workbook_field in {"description", "summary"}:
                workbook_value = strip_pages_from_summary(workbook_value, book.get("pages"))
                master_value = strip_pages_from_summary(master_value, book.get("pages"))
            if workbook_value == master_value:
                stats["exact_matches"] += 1
                continue
            approved_entry = (approved_normalisations.get(slug) or {}).get(workbook_field)
            if approved_entry and approved_entry.get("raw") == workbook_value and approved_entry.get("approved") == master_value:
                stats["approved_transformations"] += 1
                continue
            stats["mismatched"] += 1
            errors.append(f"Workbook content mismatch for {slug}: {workbook_field} does not match {master_field}.")

    return errors, stats


def run_release_checks(books: List[Dict[str, Any]] | None = None, workbook_path: Path | None = None) -> List[str]:
    if books is None:
        books = load_master()
    errors: List[str] = []

    slugs = [book["slug"] for book in books]
    if not books:
        errors.append("Master record is empty.")
    if len(slugs) != len(set(slugs)):
        errors.append("Duplicate slugs found in the master record.")

    asins = [book["asin"] for book in books if book.get("asin")]
    if len(asins) != len(set(asins)):
        errors.append("Duplicate ASIN values found in the master record.")

    identifiers = [book["identifier"] for book in books if book.get("identifier")]
    if len(identifiers) != len(set(identifiers)):
        errors.append("Duplicate internal identifiers found in the master record.")
    errors.extend(content_role_validation_errors(books))
    errors.extend(broken_phrase_errors(books))
    errors.extend(related_book_contract_errors(books))
    for book in books:
        published = clean_paragraph(book.get("datePublished"))
        modified = clean_paragraph(book.get("dateModified"))
        if published and modified and modified[:10] < published[:10]:
            errors.append(f"{book['slug']} has dateModified earlier than datePublished.")

    # Catalogue route/category contracts prevent copy-clone regressions such as
    # Retail being generated with Sports canonical/H1/newsletter metadata.
    category_signatures: Dict[str, str] = {}
    for topic in sorted({clean_paragraph(book.get("topic")) for book in books if clean_paragraph(book.get("topic"))}):
        slug = slugify(topic)
        page = CATALOGUE_DIR / slug / "index.html"
        if not page.exists():
            errors.append(f"Catalogue page missing for {topic}: catalogue/{slug}/index.html")
            continue
        text = page.read_text(encoding="utf-8", errors="ignore")
        expected_url = f"{SITE_URL}/catalogue/{slug}/"
        for marker, label in ((f'<link href="{expected_url}" rel="canonical"/>', "canonical"), (f'<h1>{html.escape(topic)} AI Books</h1>', "H1"), (f'/newsletter/?source=topic%3A{slug}&amp;placement=inline', "newsletter source")):
            if marker not in text:
                errors.append(f"catalogue/{slug}/ {label} does not align with route/category metadata.")
        sig = re.sub(r"\s+", " ", re.sub(r"<footer.*", "", text, flags=re.I | re.S)).strip()
        digest = hashlib.sha256(sig.encode("utf-8")).hexdigest()
        if digest in category_signatures and category_signatures[digest] != slug:
            errors.append(f"Cross-category duplicate detected: {slug} duplicates {category_signatures[digest]}.")
        category_signatures[digest] = slug

    physical_sitemaps = [name for name in ("sitemap.xml", "Sitemap.xml", "site-map.xml", "sitemap (1).xml") if (ROOT / name).exists()]
    if physical_sitemaps != ["sitemap.xml"]:
        errors.append(f"Production root must contain one physical sitemap.xml; found {physical_sitemaps}.")

    legacy_data_files = [ROOT / "data" / "ebook-content-source.json", ROOT / "data" / "ebook-source-overrides.json"]
    for legacy_path in legacy_data_files:
        if legacy_path.exists():
            errors.append(f"Legacy data source still exists: {legacy_path.relative_to(ROOT)}")

    detail_files = list(EBOOKS_DIR.glob("**/detail.html")) + list(EBOOKS_DIR.glob("**/details.html"))
    if detail_files:
        errors.append("Legacy detail source files still exist under /ebooks/.")

    books_index_path = EBOOKS_DIR / "index.html"
    if not books_index_path.exists():
        errors.append("ebooks/index.html is missing.")
    else:
        index_html = books_index_path.read_text(encoding="utf-8")
        for book in books:
            if f'/ebooks/{book["slug"]}/' not in index_html:
                errors.append(f"ebooks/index.html does not link to /ebooks/{book['slug']}/.")
        if index_html.count('data-placement="catalogue_card"') < len(books) * 2 or index_html.count('>View book</a>') < len(books) or index_html.count('>Buy on Amazon</a>') < len(books):
            errors.append("Catalogue cards must expose View book and Buy on Amazon outside <details> for every title.")

    seen_serp_titles: Dict[str, str] = {}
    duplicate_serp_titles: Dict[str, List[str]] = defaultdict(list)
    for book in books:
        page_path = EBOOKS_DIR / book["slug"] / "index.html"
        metadata_path = EBOOKS_DIR / book["slug"] / "metadata.json"
        faq_path = EBOOKS_DIR / book["slug"] / "faq-schema.json"
        if not page_path.exists():
            errors.append(f"Book page missing: {page_path}")
            continue
        text = page_path.read_text(encoding="utf-8")
        if f'<link href="{book["canonical_url"]}" rel="canonical"/>' not in text:
            errors.append(f"{book['slug']} is missing the expected canonical link.")
        if text.count('"@type":"FAQPage"') != 1:
            errors.append(f"{book['slug']} should emit exactly one FAQPage JSON-LD block.")
        if text.count('"@type":"Book"') != 1:
            errors.append(f"{book['slug']} should emit exactly one Book JSON-LD block.")
        person_ref = f'"author":{{"@id":"{SITE_URL}/#person"}}'
        publisher_ref = f'"publisher":{{"@id":"{SITE_URL}/#person"}}'
        if person_ref not in text or publisher_ref not in text:
            errors.append(f"{book['slug']} Book JSON-LD must reference the canonical #person author/publisher node.")
        if "Before you buy" not in text or 'href="#deeper-overview">See exactly what this book covers' not in text:
            errors.append(f"{book['slug']} is missing the buying-confidence strip.")
        title_match = re.search(r"<title>(.*?)</title>", text, re.I | re.S)
        if not title_match:
            errors.append(f"{book['slug']} page head title is missing.")
        else:
            actual_title = clean_paragraph(html.unescape(re.sub(r"\s+", " ", title_match.group(1))))
            if not actual_title.endswith(f" | {SITE_NAME}"):
                errors.append(f"{book['slug']} page title is missing the brand suffix.")
            # March 2026 title policy: canonical ebook pages keep full workbook-governed titles.
            # Length is advisory here because the workbook remains the governing source of truth.
            if actual_title in seen_serp_titles and seen_serp_titles[actual_title] != book['slug']:
                duplicate_serp_titles[actual_title].extend([seen_serp_titles[actual_title], book['slug']])
            else:
                seen_serp_titles[actual_title] = book['slug']
        expected_social_title = html.escape(f"{book_meta_title(book)} | {SITE_NAME}", quote=False)
        for marker in [
            f'<meta content="{expected_social_title}" property="og:title"/>',
            f'<meta content="{expected_social_title}" name="twitter:title"/>',
        ]:
            if marker not in text:
                errors.append(f"{book['slug']} page social title drift detected.")
                break
        expected_desc = html.escape(book_meta_description(book), quote=True)
        for marker in [
            f'<meta content="{expected_desc}" name="description"/>',
            f'<meta content="{expected_desc}" property="og:description"/>',
            f'<meta content="{expected_desc}" name="twitter:description"/>',
        ]:
            if marker not in text:
                errors.append(f"{book['slug']} page head description drift detected.")
                break
        errors.extend(metadata_budget_errors(f"ebooks/{book['slug']}/index.html", text, max_description=155))
        errors.extend(template_contract_errors(text, book))
        if f'href="/book/{book["slug"]}/' in text or f'href="/ebooks/{book["slug"]}/detail' in text:
            errors.append(f"{book['slug']} still links to a retired route.")
        if not faq_is_semantically_relevant(book):
            errors.append(f"{book['slug']} has FAQ content that looks off-topic.")
        if not metadata_path.exists() or not faq_path.exists():
            errors.append(f"{book['slug']} is missing metadata.json or faq-schema.json.")
        else:
            metadata = read_json(metadata_path, default={}) or {}
            faq_schema = read_json(faq_path, default={}) or {}
            if metadata.get("canonical_url") != book["canonical_url"]:
                errors.append(f"{book['slug']} metadata.json canonical_url does not match the master record.")
            if metadata.get("buy_url") != book["buy_url"]:
                errors.append(f"{book['slug']} metadata.json buy_url does not match the master record.")
            meta_published = clean_paragraph(metadata.get("datePublished"))
            meta_modified = clean_paragraph(metadata.get("dateModified"))
            if meta_published and meta_modified and meta_modified[:10] < meta_published[:10]:
                errors.append(f"{book['slug']} metadata.json has dateModified earlier than datePublished.")
            book_schema = None
            for script_body in re.findall(r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>', text, flags=re.I | re.S):
                try:
                    candidate = json.loads(html.unescape(script_body))
                except (TypeError, ValueError, json.JSONDecodeError):
                    continue
                if isinstance(candidate, dict) and candidate.get("@type") == "Book":
                    book_schema = candidate
                    break
            if isinstance(book_schema, dict):
                schema_published = clean_paragraph(book_schema.get("datePublished"))
                schema_modified = clean_paragraph(book_schema.get("dateModified"))
                if schema_published and schema_modified and schema_modified[:10] < schema_published[:10]:
                    errors.append(f"{book['slug']} generated Book JSON-LD has dateModified earlier than datePublished.")
            else:
                errors.append(f"{book['slug']} generated Book JSON-LD could not be parsed for chronology validation.")
            if faq_schema.get("mainEntity") != book.get("faq"):
                errors.append(f"{book['slug']} faq-schema.json does not match the master record.")

        topic_page = ROOT / book["topic_url"].strip("/") / "index.html"
        if not book.get("topic_url") or not topic_page.exists():
            errors.append(f"{book['slug']} is missing a governed topic discovery page.")

    for title, slugs in duplicate_serp_titles.items():
        unique_slugs = sorted(set(slugs))
        if len(unique_slugs) > 1:
            errors.append(f"Duplicate ebook SERP title '{title}' used by: {', '.join(unique_slugs)}")

    discovery_pages = [
        {
            "path": EBOOKS_DIR / "index.html",
            "label": "ebooks/index.html",
            "required": [
                '<meta content="index,follow" name="robots"/>',
                '<meta content="website" property="og:type"/>',
                f'<meta content="{SITE_URL}/ebooks/" property="og:url"/>',
                '<meta content="AI eBooks Catalogue | Jonathan Harris" property="og:title"/>',
                '<meta content="summary_large_image" name="twitter:card"/>',
                '<meta content="AI eBooks Catalogue | Jonathan Harris" name="twitter:title"/>',
            ],
            "description": f"Ebook catalogue: {len(books)} AI titles by Jonathan Harris covering industries, ethics, safety, and practical adoption.",
            "max_description": 155,
        },
        {
            "path": TOPICS_DIR / "index.html",
            "label": "topics/index.html",
            "required": [
                '<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>',
                '<link href="https://images.jonathan-harris.online" rel="preconnect"/>',
                '<link href="https://assets.jonathan-harris.online" rel="preconnect"/>',
                '<meta content="#0D1420" name="theme-color"/>',
                '<meta content="index,follow" name="robots"/>',
                '<meta content="website" property="og:type"/>',
                f'<meta content="{SITE_URL}/topics/" property="og:url"/>',
                '<meta content="AI Topics | Jonathan Harris" property="og:title"/>',
                '<meta content="summary_large_image" name="twitter:card"/>',
                '<meta content="AI Topics | Jonathan Harris" name="twitter:title"/>',
            ],
            "description": topics_index_meta_description(),
            "min_description": 120,
            "max_description": 155,
        },
    ]
    for topic in sorted({book["topic"] for book in books}, key=str.lower):
        topic_slug = slugify(topic)
        discovery_pages.append({
            "path": CATALOGUE_DIR / topic_slug / "index.html",
            "label": f"catalogue/{topic_slug}/index.html",
            "required": [
                '<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>',
                '<link href="https://images.jonathan-harris.online" rel="preconnect"/>',
                '<link href="https://assets.jonathan-harris.online" rel="preconnect"/>',
                '<meta content="#0D1420" name="theme-color"/>',
                '<meta content="index,follow" name="robots"/>',
                '<meta content="website" property="og:type"/>',
                f'<meta content="{SITE_URL}/catalogue/{topic_slug}/" property="og:url"/>',
                f'<meta content="{html.escape(topic)} AI Books | Jonathan Harris" property="og:title"/>',
                '<meta content="summary_large_image" name="twitter:card"/>',
                f'<meta content="{html.escape(topic)} AI Books | Jonathan Harris" name="twitter:title"/>',
            ],
            "description": catalogue_meta_description(topic),
            "min_description": 120,
            "max_description": 155,
        })

    for page in discovery_pages:
        if not page["path"].exists():
            errors.append(f"Discovery page missing: {page['label']}")
            continue
        page_text = page["path"].read_text(encoding="utf-8")
        canonical_match = re.search(r'<link[^>]+href="([^"]+)"[^>]+rel="canonical"', page_text, re.I)
        canonical_href = canonical_match.group(1) if canonical_match else ""
        if canonical_href:
            expected_hreflang = [
                f'<link href="{canonical_href}" hreflang="en" rel="alternate"/>',
                f'<link href="{canonical_href}" hreflang="x-default" rel="alternate"/>',
            ]
            for marker in expected_hreflang:
                if marker not in page_text:
                    errors.append(f"Discovery hreflang metadata missing from {page['label']}: {marker}")
        for marker in page["required"]:
            if marker not in page_text:
                errors.append(f"Discovery metadata missing from {page['label']}: {marker}")
        expected_description = page.get("description")
        if expected_description:
            description_variants = {
                expected_description,
                html.escape(expected_description, quote=False),
                html.escape(expected_description, quote=True),
            }
            required_tags = ['name="description"', 'property="og:description"', 'name="twitter:description"']
            for tag in required_tags:
                if not any(f'<meta content="{variant}" {tag}/>' in page_text for variant in description_variants):
                    errors.append(f"Discovery description drift detected for {page['label']}: {tag}")
                    break
        errors.extend(metadata_budget_errors(page["label"], page_text, min_description=page.get("min_description"), max_description=page.get("max_description")))

    redirects_text = (ROOT / "_redirects").read_text(encoding="utf-8")
    redirects_mirror_path = ROOT / "_redirects.txt"
    if not redirects_mirror_path.exists():
        errors.append("_redirects.txt is missing. Run scripts/sync_redirects.py to regenerate the governed mirror.")
        redirects_mirror = ""
    else:
        redirects_mirror = redirects_mirror_path.read_text(encoding="utf-8")
        if redirects_text != redirects_mirror:
            errors.append("_redirects.txt is not an exact generated mirror of _redirects.")
    redirect_block_pattern = re.compile(r"# 6A\) Branded buy-now redirects.*?(?=\n# 7\) CANONICAL: old /book URLs permanently redirect to /ebooks)", re.S)
    existing_redirect_block = redirect_block_pattern.search(redirects_text)
    expected_redirect_block = build_redirect_block(books).strip()
    if not existing_redirect_block or existing_redirect_block.group(0).strip() != expected_redirect_block:
        errors.append("The governed branded buy-now redirect block in _redirects has drifted from the generated source.")
    for line in LEGACY_DETAIL_REDIRECT_LINES:
        if line not in redirects_text:
            errors.append(f"Redirect family missing from _redirects: {line}")
        if line not in redirects_mirror:
            errors.append(f"Redirect family missing from _redirects.txt: {line}")
    for legacy_alias in ("/en-gb/*  /:splat  200", "/en-us/*  /:splat  200", "/en-ca/*  /:splat  200", "/en-au/*  /:splat  200"):
        if legacy_alias in redirects_text or legacy_alias in redirects_mirror:
            errors.append(f"Locale alias pass-through must not remain live: {legacy_alias}")
    for line in LOCALE_ALIAS_REDIRECT_LINES:
        if line not in redirects_text:
            errors.append(f"Locale alias redirect missing from _redirects: {line}")
        if line not in redirects_mirror:
            errors.append(f"Locale alias redirect missing from _redirects.txt: {line}")
    malformed_lines = [f"{item['source']}   {item['target']}  301" for item in MALFORMED_SLUG_FIXES]
    for line in malformed_lines:
        if line not in redirects_text:
            errors.append(f"Malformed slug fix missing from _redirects: {line}")
        if line not in redirects_mirror:
            errors.append(f"Malformed slug fix missing from _redirects.txt: {line}")
    for book in books:
        required_lines = [
            f"{book['buy_route']}   {book['buy_url']}   302",
            f"/book/{book['slug']}/buy-now   {book['buy_route']}   301",
        ]
        for line in required_lines:
            if line not in redirects_text:
                errors.append(f"Redirect rule missing from _redirects: {line}")
            if line not in redirects_mirror:
                errors.append(f"Redirect rule missing from _redirects.txt: {line}")

    forbidden_main_redirects = [
        f"/robots.txt    {EXTERNAL_CRAWLER_FILES['robots']}   301",
        f"/sitemap.xml   {EXTERNAL_CRAWLER_FILES['sitemap']}  301",
        f"/site-map.xml {EXTERNAL_CRAWLER_FILES['sitemap']} 301",
    ]
    for snippet in forbidden_main_redirects:
        if snippet in redirects_text:
            errors.append(f"Main redirect file still redirects a governed crawler asset instead of serving it directly: {snippet}")
        if snippet in redirects_mirror:
            errors.append(f"Main redirect mirror still redirects a governed crawler asset instead of serving it directly: {snippet}")

    legacy_typo_alias = "/robot.txt    /robots.txt   301"
    legacy_sitemap_alias = "/Sitemap.xml  /sitemap.xml  301"
    legacy_site_map_alias = "/site-map.xml  /sitemap.xml  301"
    if legacy_typo_alias not in redirects_text or legacy_typo_alias not in redirects_mirror:
        errors.append("Legacy /robot.txt crawler alias is missing from the main redirect files.")
    if legacy_sitemap_alias not in redirects_text or legacy_sitemap_alias not in redirects_mirror:
        errors.append("Legacy /Sitemap.xml crawler alias is missing from the main redirect files.")
    if legacy_site_map_alias not in redirects_text or legacy_site_map_alias not in redirects_mirror:
        errors.append("Legacy /site-map.xml crawler alias is missing from the main redirect files.")

    books_domain = (ROOT / "_redirects.books-domain").read_text(encoding="utf-8")
    ebooks_domain = (ROOT / "_redirects.ebooks-domain").read_text(encoding="utf-8")
    if books_domain != build_books_domain_redirects():
        errors.append("_redirects.books-domain has drifted from the generated host redirect source.")
    if ebooks_domain != build_ebooks_domain_redirects():
        errors.append("_redirects.ebooks-domain has drifted from the generated host redirect source.")
    host_expectations = [
        f"/robots.txt   {EXTERNAL_CRAWLER_FILES['robots']}",
        f"/sitemap.xml  {EXTERNAL_CRAWLER_FILES['sitemap']}",
        f"/site-map.xml {EXTERNAL_CRAWLER_FILES['sitemap']}",
        "/book/*",
        "/catalogue/*",
    ]
    for snippet in host_expectations:
        if snippet not in ebooks_domain:
            errors.append(f"Host redirect coverage missing from _redirects.ebooks-domain: {snippet}")
    books_domain_expectations = ["/book/*", "/ebooks/*", "/catalogue/*"]
    for snippet in books_domain_expectations:
        if snippet not in books_domain:
            errors.append(f"Host redirect coverage missing from _redirects.books-domain: {snippet}")

    generated_crawler_files = build_crawler_snapshot_paths(books)
    for file_path, expected in generated_crawler_files.items():
        if not file_path.exists():
            errors.append(f"Generated crawler snapshot missing: {file_path.relative_to(ROOT)}")
            continue
        actual = file_path.read_text(encoding="utf-8")
        if actual != expected:
            errors.append(f"Generated crawler snapshot drift detected: {file_path.relative_to(ROOT)}")

    published_crawler_files = build_published_crawler_paths(books)
    for file_path, expected in published_crawler_files.items():
        if not file_path.exists():
            errors.append(f"Published crawler file missing: {file_path.relative_to(ROOT)}")
            continue
        actual = file_path.read_text(encoding="utf-8")
        if actual != expected:
            errors.append(f"Published crawler file drift detected: {file_path.relative_to(ROOT)}")

    forbidden_legacy_crawler_duplicates = [
        ROOT / "sitemap (1).xml",
        CRAWLER_SNAPSHOTS_DIR / "site-map.xml",
    ]
    for legacy_path in forbidden_legacy_crawler_duplicates:
        if legacy_path.exists():
            errors.append(
                f"Legacy duplicate crawler file must be deleted so sitemap.xml remains the single source of truth: {legacy_path.relative_to(ROOT)}"
            )

    sitemap_routes = build_public_route_registry(books)
    expected_sitemap_locations = {route["loc"] for route in sitemap_routes}
    actual_sitemap = build_crawler_snapshot_payloads(books)[CRAWLER_SNAPSHOT_FILENAMES["sitemap"]]
    actual_sitemap_locations = set(re.findall(r"<loc>([^<]+)</loc>", actual_sitemap))
    missing_sitemap_locations = sorted(expected_sitemap_locations - actual_sitemap_locations)
    unexpected_sitemap_locations = sorted(actual_sitemap_locations - expected_sitemap_locations)
    if missing_sitemap_locations:
        errors.append(f"Sitemap coverage missing {len(missing_sitemap_locations)} public route(s); first missing: {missing_sitemap_locations[0]}")
    if unexpected_sitemap_locations:
        errors.append(f"Sitemap includes unexpected route(s); first unexpected: {unexpected_sitemap_locations[0]}")

    sitemap_text = actual_sitemap
    for book in books:
        book_loc = html.escape(book["canonical_url"])
        pattern = re.compile(rf"<loc>{re.escape(book_loc)}</loc>\s*<lastmod>([^<]+)</lastmod>")
        match = pattern.search(sitemap_text)
        expected_lastmod = normalise_lastmod(book.get("dateModified") or book.get("datePublished"))
        if not match:
            errors.append(f"Sitemap entry missing for canonical book page {book['slug']}.")
            continue
        if match.group(1) != expected_lastmod:
            errors.append(f"Sitemap lastmod drift detected for {book['slug']}: expected {expected_lastmod}, found {match.group(1)}.")

    stray_catalogue_file = CATALOGUE_DIR / "cyber-security" / "1"
    if stray_catalogue_file.exists():
        errors.append("Stray catalogue artefact still exists: catalogue/cyber-security/1")

    manifest = read_json(EBOOKS_DIR / "url-manifest.json", default={}) or {}
    manifest_books = manifest.get("ebooks", [])
    if len(manifest_books) != len(books):
        errors.append("ebooks/url-manifest.json does not contain one route record per book.")
    if manifest.get("malformed_slug_fixes") != MALFORMED_SLUG_FIXES:
        errors.append("ebooks/url-manifest.json does not fully declare malformed slug fixes.")
    manifest_by_slug = {item.get("slug"): item for item in manifest_books}
    for book in books:
        entry = manifest_by_slug.get(book["slug"])
        if not entry:
            errors.append(f"Route manifest missing slug {book['slug']}.")
            continue
        legacy_routes = entry.get("legacy_routes", [])
        expected_legacy = {f"/book/{book['slug']}/", f"/book/{book['slug']}/buy-now", f"/ebooks/{book['slug']}/detail", f"/ebooks/{book['slug']}/detail.html", f"/ebooks/{book['slug']}/details.html"}
        if set(legacy_routes) != expected_legacy:
            errors.append(f"Route manifest legacy route coverage is incomplete for {book['slug']}.")
        if entry.get("buy", {}).get("target") != book["buy_url"]:
            errors.append(f"Route manifest buy target mismatch for {book['slug']}.")

    html_files = [
        p for p in ROOT.rglob("*.html")
        if "node_modules" not in p.parts
        and not p.relative_to(ROOT).as_posix().startswith("assets/site-shell/")
        and not is_r2_hosted_podcast_episode_path(p.relative_to(ROOT))
    ]
    css_bundle = "\n".join(css_path.read_text(encoding="utf-8", errors="ignore") for css_path in (ROOT / "assets" / "css").glob("*.css"))
    if any(re.search(r"class=[\"']([^\"']*\bsr-only\b[^\"']*)[\"']", p.read_text(encoding="utf-8", errors="ignore"), re.I) for p in html_files):
        if not re.search(r"(^|[\s,{])\.sr-only\b", css_bundle, re.M):
            errors.append("Shared CSS is missing a governed .sr-only utility while HTML still references sr-only.")
    for file_path in html_files:
        text = file_path.read_text(encoding="utf-8", errors="ignore")
        if file_path == ROOT / "index.html" and "\"@type\": \"Organization\"" not in text:
            errors.append("Homepage Organisation schema is missing from index.html.")
        if re.search(r'href="/book/', text):
            errors.append(f"Retired /book/ link found in HTML: {file_path.relative_to(ROOT)}")
        if re.search(r'href="[^"]*/detail(?:\.html)?"', text):
            errors.append(f"Retired /detail link found in HTML: {file_path.relative_to(ROOT)}")
        if has_disallowed_inline_style(text):
            errors.append(f"Inline style attribute found in HTML: {file_path.relative_to(ROOT)}")

    static_metadata_pages = [
        (ROOT / "index.html", "index.html", None, None, 155),
        (ROOT / "bio" / "index.html", "bio/index.html", None, None, 155),
        (ROOT / "podcast" / "index.html", "podcast/index.html", 60, None, 155),
        (ROOT / "blog" / "weekly" / "index.html", "blog/weekly/index.html", 60, None, 155),
    ]
    for page_path, label, max_title, min_description, max_description in static_metadata_pages:
        if not page_path.exists():
            errors.append(f"Static page missing: {label}")
            continue
        page_text = page_path.read_text(encoding="utf-8", errors="ignore")
        errors.extend(
            metadata_budget_errors(
                label,
                page_text,
                max_title=max_title,
                min_description=min_description,
                max_description=max_description,
            )
        )

    for topic in sorted({book["topic"] for book in books}, key=str.lower):
        topic_slug = slugify(topic)
        page_path = CATALOGUE_DIR / topic_slug / "index.html"
        if not page_path.exists():
            errors.append(f"Catalogue topic page missing for breadcrumb validation: {page_path.relative_to(ROOT)}")
            continue
        page_text = page_path.read_text(encoding="utf-8", errors="ignore")
        expected_nav = f'<nav aria-label="Breadcrumb" class="breadcrumbs"><a href="/">Home</a><span aria-hidden="true">›</span><a href="/topics/">Topics</a><span aria-hidden="true">›</span><span>{html.escape(topic)}</span></nav>'
        if expected_nav not in page_text:
            errors.append(f"Catalogue breadcrumb trail drift detected for {page_path.relative_to(ROOT)}.")
        breadcrumb_payload = json.dumps(build_topic_breadcrumb_schema(topic), ensure_ascii=False, separators=(",", ":"))
        if breadcrumb_payload not in page_text:
            errors.append(f"Catalogue breadcrumb schema missing or drifted for {page_path.relative_to(ROOT)}.")
        for cover_match in re.finditer(r'<img\b[^>]*class="([^"]*\bcover\b[^"]*)"[^>]*>', page_text, re.I):
            tag = cover_match.group(0)
            src = extract_img_src(tag)
            if is_remote_image_src(src) and not src.startswith("https://images.jonathan-harris.online/"):
                errors.append(f"Unapproved remote cover host in {page_path.relative_to(ROOT)}.")
                break
            # Absolute covers are served directly from the governed image host.
            # Do not wrap them in same-origin /cdn-cgi/image transforms because that
            # path is not reliable when the image host redirects upstream.
            if is_remote_image_src(src):
                if '/cdn-cgi/image/' in tag:
                    errors.append(f"Remote cover is incorrectly wrapped in Cloudflare image resizing: {page_path.relative_to(ROOT)}.")
                    break
            else:
                if 'srcset="' not in tag or 'sizes="' not in tag:
                    errors.append(f"Same-origin responsive cover markup missing from {page_path.relative_to(ROOT)}.")
                    break
                if not all(f" {width}w" in tag for width in (400, 800, 1200)):
                    errors.append(f"Same-origin responsive cover widths drift detected for {page_path.relative_to(ROOT)}.")
                    break

    homepage_text = (ROOT / "index.html").read_text(encoding="utf-8", errors="ignore")
    featured_cover_match = re.search(r'<img\b[^>]*id="featuredEbookCover"[^>]*>', homepage_text, re.I)
    if not featured_cover_match:
        errors.append("Homepage featured cover image is missing.")
    else:
        featured_tag = featured_cover_match.group(0)
        featured_src = extract_img_src(featured_tag)
        if is_remote_image_src(featured_src) and not featured_src.startswith("https://images.jonathan-harris.online/"):
            errors.append("Homepage featured cover uses an unapproved remote image host.")
        if is_remote_image_src(featured_src):
            if '/cdn-cgi/image/' in featured_tag:
                errors.append("Homepage remote featured cover is incorrectly wrapped in Cloudflare image resizing.")
        elif 'srcset="' not in featured_tag or 'sizes="' not in featured_tag:
            errors.append("Homepage same-origin featured cover is missing responsive srcset/sizes markup.")
        elif not all(f" {width}w" in featured_tag for width in (400, 800, 1200)):
            errors.append("Homepage same-origin featured cover responsive widths drifted from the governed 400/800/1200 contract.")

    for book in books:
        page_path = EBOOKS_DIR / book["slug"] / "index.html"
        if not page_path.exists():
            continue
        page_text = page_path.read_text(encoding="utf-8", errors="ignore")
        cover_match = re.search(r'<img\b[^>]*class="([^"]*\bebook-showcase__cover\b[^"]*)"[^>]*>', page_text, re.I)
        if not cover_match:
            errors.append(f"Book cover block missing from {page_path.relative_to(ROOT)}.")
            continue
        cover_tag = cover_match.group(0)
        cover_src = extract_img_src(cover_tag)
        if is_remote_image_src(cover_src) and not cover_src.startswith("https://images.jonathan-harris.online/"):
            errors.append(f"Unapproved remote book cover host in {page_path.relative_to(ROOT)}.")
            continue
        if is_remote_image_src(cover_src):
            if '/cdn-cgi/image/' in cover_tag:
                errors.append(f"Remote book cover is incorrectly wrapped in Cloudflare image resizing: {page_path.relative_to(ROOT)}.")
        else:
            if 'srcset="' not in cover_tag or 'sizes="' not in cover_tag:
                errors.append(f"Same-origin responsive book cover markup missing from {page_path.relative_to(ROOT)}.")
                continue
            if not all(f" {width}w" in cover_tag for width in (400, 800, 1200)):
                errors.append(f"Same-origin responsive book cover widths drift detected for {page_path.relative_to(ROOT)}.")

    js_image_checks = {
        ROOT / "assets" / "js" / "featured-book.min.js": r'if(!e||!Array.isArray(t)||/^https?:\/\//i.test(e))return""',
        ROOT / "assets" / "js" / "books.min.js": r'if(/^https?:\/\//i.test(e))return""',
    }
    for file_path, required_snippet in js_image_checks.items():
        file_text = file_path.read_text(encoding="utf-8", errors="ignore")
        compact_text = re.sub(r"\s+", "", file_text)
        if required_snippet not in compact_text:
            errors.append(f"Remote-image safety guard drift detected in {file_path.relative_to(ROOT)}")

    removal_plan_path = ROOT / "docs" / "search-console-stale-url-removal-plan.md"
    if not removal_plan_path.exists():
        errors.append("Search Console stale URL remediation plan is missing: docs/search-console-stale-url-removal-plan.md")
    else:
        removal_plan_text = removal_plan_path.read_text(encoding="utf-8", errors="ignore")
        for snippet in ["/en-gb/", "/en-au/", "/book/", "Search Console", "Request indexing", "Removals"]:
            if snippet not in removal_plan_text:
                errors.append(f"Search Console stale URL remediation plan is missing required guidance: {snippet}")

    static_topic_pages = [
        ROOT / "topics" / "ai-for-beginners" / "index.html",
        ROOT / "topics" / "ai-in-business" / "index.html",
        ROOT / "topics" / "ai-in-healthcare" / "index.html",
        ROOT / "topics" / "generative-ai" / "index.html",
        ROOT / "topics" / "robotics-automation" / "index.html",
    ]
    for page_path in static_topic_pages:
        if not page_path.exists():
            errors.append(f"Static topic page missing: {page_path.relative_to(ROOT)}")
            continue
        page_text = page_path.read_text(encoding="utf-8", errors="ignore")
        for marker in [
            '<link href="https://assets.jonathan-harris.online/favicon.ico" rel="icon" type="image/x-icon"/>',
            '<link href="https://images.jonathan-harris.online" rel="preconnect"/>',
            '<link href="https://assets.jonathan-harris.online" rel="preconnect"/>',
            '<meta content="#0D1420" name="theme-color"/>',
        ]:
            if marker not in page_text:
                errors.append(f"Static topic page head baseline drift detected for {page_path.relative_to(ROOT)}: {marker}")

    checksum_payload = read_json(CRAWLER_CHECKSUMS_PATH, default={}) or {}
    checksum_files = checksum_payload.get("files", {})
    expected_crawler_content = build_crawler_snapshot_payloads(books)
    if set(checksum_files.keys()) != set(expected_crawler_content.keys()):
        errors.append("config/crawler-checksums.json does not declare the governed crawler snapshots.")
    for name, expected_text in expected_crawler_content.items():
        file_payload = checksum_files.get(name, {})
        expected_hash = sha256_text(expected_text)
        if file_payload.get("sha256") != expected_hash:
            errors.append(f"Crawler checksum drift detected for {name}.")

    # Blog publication content is externally governed in R2. The repository must not
    # carry or validate a committed blog/posts.json snapshot, because that could
    # resurrect stale posts after the source object has been removed. Runtime
    # functions below remain responsible for exposing the live R2 manifest.
    committed_blog_manifest = ROOT / "blog" / "posts.json"
    if committed_blog_manifest.exists():
        errors.append("blog/posts.json must not be committed; the live blog manifest is governed in R2.")

    weekly_archive_html = (ROOT / "blog" / "weekly" / "index.html").read_text(encoding="utf-8")
    site_ui_js = (ROOT / "assets" / "js" / "site-ui.min.js").read_text(encoding="utf-8")
    blog_js = (ROOT / "assets" / "js" / "blog.bundle.min.js").read_text(encoding="utf-8")
    weekly_archive_runtime_js = ROOT / "functions" / "blog" / "weekly" / "index.js"
    sitemap_runtime_js = ROOT / "functions" / "sitemap.xml.js"
    blog_manifest_runtime_js = ROOT / "functions" / "blog" / "posts.json.js"
    blog_post_runtime_js = ROOT / "functions" / "blog" / "posts" / "[[slug]].js"
    blog_image_runtime_js = ROOT / "functions" / "blog" / "images" / "[[slug]].js"
    if "manifest stack" in weekly_archive_html or "published manifest" in weekly_archive_html or "falls back to" in weekly_archive_html:
        errors.append("blog/weekly/index.html still exposes runtime publication language instead of deterministic archive copy.")
    if "cfg.R2_PUBLIC_BASE_URL_BLOG" in blog_js or "cfg.RSS_URL" in blog_js:
        errors.append("assets/js/blog.bundle.min.js still depends on remote manifest or RSS fallback instead of the same-origin weekly publication surface.")
    if "createElement(\"style\")" in site_ui_js or "createElement('style')" in site_ui_js:
        errors.append("assets/js/site-ui.min.js still injects runtime style tags instead of relying on governed CSS.")
    if not weekly_archive_runtime_js.exists():
        errors.append("functions/blog/weekly/index.js is missing; weekly archive robots state would drift from the live publication manifest.")
    else:
        weekly_runtime_text = weekly_archive_runtime_js.read_text(encoding="utf-8")
        if "/blog/posts.json" not in weekly_runtime_text:
            errors.append("functions/blog/weekly/index.js must read the same-origin /blog/posts.json publication manifest.")
    if not sitemap_runtime_js.exists():
        errors.append("functions/sitemap.xml.js is missing; sitemap visibility would drift from the live publication manifest.")
    else:
        sitemap_runtime_text = sitemap_runtime_js.read_text(encoding="utf-8")
        if "/blog/posts.json" not in sitemap_runtime_text:
            errors.append("functions/sitemap.xml.js must read the same-origin /blog/posts.json publication manifest.")

    if not blog_manifest_runtime_js.exists():
        errors.append("functions/blog/posts.json.js is missing; /blog/posts.json would stay pinned to the stale committed snapshot instead of the live R2 manifest.")
    if not blog_post_runtime_js.exists():
        errors.append("functions/blog/posts/[[slug]].js is missing; same-origin blog post URLs cannot proxy published posts from R2.")
    if not blog_image_runtime_js.exists():
        errors.append("functions/blog/images/[[slug]].js is missing; same-origin blog image URLs cannot proxy published artwork from R2.")

    llm_index_payload = read_json(ROOT / "llm-index.json", default={}) or {}
    llm_index_books = llm_index_payload.get("books") if isinstance(llm_index_payload, dict) else None
    if not isinstance(llm_index_books, list) or len(llm_index_books) != len(books):
        errors.append("llm-index.json is missing book records or is out of sync with the master record.")
    else:
        llm_index_by_slug = {clean_paragraph(item.get("slug")): item for item in llm_index_books if isinstance(item, dict) and clean_paragraph(item.get("slug"))}
        for book in books:
            indexed = llm_index_by_slug.get(book["slug"])
            if not indexed:
                errors.append(f"llm-index.json is missing record for {book['slug']}")
                continue
            if indexed.get("related_slugs", []) != book.get("related_slugs", []):
                errors.append(f"llm-index.json related_slugs drift detected for {book['slug']}")

    errors.extend(static_blog_archive_errors())
    errors.extend(blog_related_link_errors())
    errors.extend(json_ld_validation_errors())

    featured_payload = read_json(ROOT / "api" / "v1" / "featured-book.json", default={}) or {}
    expected_featured_payload = build_featured_book_payload([book_to_public_record(book) for book in books])
    if featured_payload != expected_featured_payload:
        errors.append("Derivative featured-book API payload drift detected for api/v1/featured-book.json.")

    page_404 = ROOT / "404.html"
    if not page_404.exists():
        errors.append("404.html is missing.")
    else:
        not_found_text = page_404.read_text(encoding="utf-8")
        title_match = re.search(r"<title>(.*?)</title>", not_found_text, re.I | re.S)
        og_title_match = re.search(r'<meta[^>]+(?:property="og:title"[^>]+content="([^"]+)"|content="([^"]+)"[^>]+property="og:title")', not_found_text, re.I)
        canonical_match = re.search(r'<link[^>]+href="([^"]+)"[^>]+rel="canonical"', not_found_text, re.I)
        title_value = clean_paragraph(html.unescape(title_match.group(1))) if title_match else ""
        og_title_value = clean_paragraph(html.unescape(og_title_match.group(1) or og_title_match.group(2))) if og_title_match else ""
        canonical_href = canonical_match.group(1) if canonical_match else ""
        if not title_value:
            errors.append("404.html title tag is missing.")
        if not og_title_value:
            errors.append("404.html og:title is missing.")
        if title_value and og_title_value and title_value != og_title_value:
            errors.append("404.html og:title does not match the title tag.")
        if canonical_href:
            expected_hreflang = [
                f'<link href="{canonical_href}" hreflang="en" rel="alternate"/>',
                f'<link href="{canonical_href}" hreflang="x-default" rel="alternate"/>',
            ]
            for marker in expected_hreflang:
                if marker not in not_found_text:
                    errors.append(f"404.html is missing hreflang metadata: {marker}")

    if workbook_path and workbook_path.exists():
        errors.extend(validate_pages_sheet_operational_view(workbook_path))
        errors.extend(workbook_static_route_contract_errors(workbook_path))
        workbook_content_errors, _ = workbook_content_parity_audit(workbook_path, books)
        errors.extend(workbook_content_errors)

    return errors


def run_import_command(args: argparse.Namespace) -> int:
    try:
        workbook_path = Path(args.workbook).expanduser().resolve()
        order, workbook_map, _ = parse_workbook(workbook_path)
    except Exception as exc:
        print(f"Import check failed: {exc}")
        return 1

    if args.check:
        if not workbook_map:
            print("Workbook check failed: no ebook rows found.")
            return 1
        print(f"Workbook check passed: {len(order)} ebook rows found.")
        return 0

    records = build_master_from_workbook(workbook_path)
    save_master(records)
    print(f"Wrote {len(records)} records to {MASTER_PATH.relative_to(ROOT)}")
    return 0



def run_fix_pages_command(args: argparse.Namespace) -> int:
    books = load_master()
    ensure_css_file()
    build_book_files(books)
    if args.check:
        errors = [
            err
            for err in run_release_checks(books, Path(args.workbook).resolve() if getattr(args, "workbook", None) else None)
            if "canonical link" in err or "FAQPage" in err or "Book JSON-LD" in err or "Book page missing" in err
        ]
        if errors:
            for error in errors:
                print(error)
            return 1
        print("Page metadata check passed.")
        return 0
    print(f"Generated {len(books)} canonical ebook pages.")
    return 0



def run_build_derivatives_command(args: argparse.Namespace) -> int:
    books = load_master()
    ensure_css_file()
    build_derivatives(books)
    if args.check:
        errors = [
            err
            for err in run_release_checks(books, Path(args.workbook).resolve() if getattr(args, "workbook", None) else None)
            if "Derivative" in err or "topic discovery" in err or "url-manifest" in err or "host redirect" in err
        ]
        if errors:
            for error in errors:
                print(error)
            return 1
        print("Derivative build check passed.")
        return 0
    print("Derivative JSON, manifests, crawler files, and governed discovery pages rebuilt.")
    return 0



def run_sync_redirects_command(args: argparse.Namespace) -> int:
    books = load_master()
    sync_redirects(books)
    if args.check:
        errors = [err for err in run_release_checks(books) if "Redirect rule missing" in err]
        if errors:
            for error in errors:
                print(error)
            return 1
        print("Redirect sync check passed.")
        return 0
    print("Redirect files synchronised.")
    return 0





def json_ld_validation_errors() -> List[str]:
    errors: List[str] = []
    for path in sorted(ROOT.rglob("*.html")):
        relative = path.relative_to(ROOT)
        if "node_modules" in path.parts or "assets" in path.parts or "templates" in path.parts or "partials" in path.parts:
            continue
        if is_r2_hosted_podcast_episode_path(relative):
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        blocks = re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', text, re.I | re.S)
        for index, raw in enumerate(blocks, start=1):
            payload = clean_paragraph(raw)
            if not payload:
                continue
            try:
                json.loads(payload)
            except Exception as exc:
                errors.append(f"Invalid JSON-LD in {relative} block #{index}: {exc}")
    return errors


def static_blog_archive_errors() -> List[str]:
    errors: List[str] = []
    manifest = read_json(ROOT / "blog" / "posts.json", default={}) or {}
    items = manifest.get("items") if isinstance(manifest, dict) else None
    if not isinstance(items, list) or not items:
        return errors

    expected_paths = []
    for item in items:
        if not isinstance(item, dict):
            continue
        slug = clean_paragraph(item.get("slug"))
        if slug:
            expected_paths.append(f"/blog/posts/{slug}/")

    contracts = {
        Path("blog/index.html"): 1,
        Path("blog/weekly/index.html"): 4,
    }
    for rel, required_count in contracts.items():
        text = (ROOT / rel).read_text(encoding="utf-8")
        if "Published briefings appear here" in text or "The latest briefings appear here" in text:
            errors.append(f"{rel} still ships placeholder archive copy instead of static post discovery.")
        missing = [path for path in expected_paths[:required_count] if path not in text]
        if missing:
            qualifier = "latest published briefing" if rel == Path("blog/index.html") else "published briefings"
            errors.append(f"{rel} is missing static links for {qualifier}. First missing: {missing[0]}")
    return errors


def blog_related_link_errors() -> List[str]:
    errors: List[str] = []
    for path in sorted((ROOT / "blog" / "posts").glob("*/index.html")):
        text = path.read_text(encoding="utf-8", errors="ignore")
        for href, label in re.findall(r'<a href="([^"]+)">([^<]+)</a>', text):
            href_clean = clean_paragraph(href)
            label_clean = clean_paragraph(html.unescape(label))
            if href_clean and not href_clean.startswith(("/", "http://", "https://", "#", "mailto:", "tel:")):
                errors.append(f"Blog post malformed link href in {path.relative_to(ROOT)}: {href_clean}")
                break
            if label_clean.startswith("/ebooks/") and href_clean and not href_clean.startswith("/ebooks/"):
                errors.append(f"Blog post related-book anchor text/href reversal detected in {path.relative_to(ROOT)}")
                break
    return errors

def rebuild_all(workbook_path: Path) -> List[str]:
    books = build_master_from_workbook(workbook_path)
    save_master(books)
    ensure_css_file()
    build_book_files(books)
    build_derivatives(books)
    sync_redirects(books)
    errors = run_release_checks(books, workbook_path)
    VALIDATION_REPORT.write_text(build_validation_report(errors, books), encoding="utf-8")
    return errors



def detect_governed_workbook_path() -> Path | None:
    candidates = sorted(ROOT.glob("*.xlsx")) + sorted(ROOT.glob("*.xlsm"))
    if len(candidates) == 1:
        return candidates[0].resolve()

    xlsx_candidates = {path.stem: path for path in ROOT.glob("*.xlsx")}
    xlsm_candidates = {path.stem: path for path in ROOT.glob("*.xlsm")}
    shared_stems = sorted(set(xlsx_candidates) & set(xlsm_candidates))
    if len(candidates) == 2 and len(shared_stems) == 1:
        return xlsx_candidates[shared_stems[0]].resolve()

    return None

def ebook_title_length_warnings(books: List[Dict[str, Any]]) -> List[str]:
    warnings: List[str] = []
    for book in books:
        page_path = EBOOKS_DIR / book["slug"] / "index.html"
        if not page_path.exists():
            continue
        text = page_path.read_text(encoding="utf-8", errors="ignore")
        title_match = re.search(r"<title>(.*?)</title>", text, re.I | re.S)
        if not title_match:
            continue
        actual_title = clean_paragraph(html.unescape(re.sub(r"\s+", " ", title_match.group(1))))
        if 90 < len(actual_title) <= 120:
            warnings.append(f"{book['slug']} page title is {len(actual_title)} characters; warning threshold is 90.")
    return warnings



def has_disallowed_inline_style(text: str) -> bool:
    """Return True when HTML contains inline style outside the approved GTM noscript iframe."""
    cleaned = re.sub(
        r'<iframe\b(?=[^>]*https://www\.googletagmanager\.com/ns\.html\?id=GTM-PC4K9KRK)(?=[^>]*style=["\']display:none;visibility:hidden["\'])[^>]*></iframe>',
        '<iframe data-approved-gtm-noscript></iframe>',
        text,
        flags=re.I | re.S,
    )
    return bool(re.search(r'\sstyle\s*=', cleaned, re.I))

def run_validate_command(workbook_path: Path | None = None) -> int:
    effective_workbook = workbook_path or detect_governed_workbook_path()
    errors = run_release_checks(workbook_path=effective_workbook)
    books = load_master()
    warnings = ebook_title_length_warnings(books)
    workbook_title_stats = None
    workbook_content_stats = None
    if effective_workbook and effective_workbook.exists():
        _, workbook_title_stats = workbook_title_parity_audit(effective_workbook)
        _, workbook_content_stats = workbook_content_parity_audit(effective_workbook, books)
    VALIDATION_REPORT.write_text(
        build_validation_report(
            errors,
            books,
            workbook_title_stats=workbook_title_stats,
            workbook_content_stats=workbook_content_stats,
        ),
        encoding="utf-8",
    )
    if workbook_title_stats:
        print(f"Workbook title parity checked {workbook_title_stats['checked']} routes; {workbook_title_stats['passed']} passed.")
    if workbook_content_stats:
        print(
            "Workbook content parity checked "
            f"{workbook_content_stats['checked_fields']} governed fields across {workbook_content_stats['books']} books; "
            f"{workbook_content_stats['exact_matches']} exact matches, "
            f"{workbook_content_stats['approved_transformations']} approved transformations, "
            f"{workbook_content_stats['mismatched']} mismatches."
        )
    for warning in warnings:
        print(f"[WARN] {warning}")
    if errors:
        for error in errors:
            print(error)
        print(f"Validation failed with {len(errors)} issue(s).")
        return 1
    print("Release validation passed.")
    return 0
